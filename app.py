"""
HSL Solutions ERP v3.0 - Unified Application
=============================================
Integrated flow: CRM → Ofertare (Configurator) → Comandă → Factură → WMS

Modules:
  /admin                - Dashboard (overview)
  /admin/clienti        - CRM: Client management  
  /admin/oferte         - Vânzări: Oferte list + detail
  /admin/comenzi        - Vânzări: Comenzi list + detail
  /admin/facturi        - Facturare: Facturi list
  /admin/wms            - WMS: Mișcări stoc
  /configurator         - Ofertare: Product configurator
  /configurator/admin   - Admin: Manage configurable products
  
API endpoints:
  /api/anaf/lookup/<cui> - ANAF CUI lookup
  /api/cfg/*             - Configurator CRUD (products, accessories, categories)
  /api/oferta/*          - Save/convert/export oferte
  /api/comanda/*         - Order management
  /api/factura/*         - Invoice generation
"""

import os, json, io
from datetime import datetime, date, timedelta, timezone
from functools import wraps
from flask import (Flask, render_template, redirect, url_for, request, flash,
                   jsonify, send_file, abort)
from flask_login import (LoginManager, login_user, logout_user,
                         login_required, current_user)
from models import (db, Utilizator, Rol, MODULES, MODULE_GROUPS, Client,
                    CategorieProdus, CategorieAccesoriu, ProdusConfig, ProdusCategorie,
                    Accesoriu, AccesoriuCompat, ProdusLegatura, SablonListaPreturi,
                    Oferta, LinieOferta, FollowUpOferta, Comanda, LinieComanda,
                    Factura, LinieFactura,
                    Furnizor, CelulaDepozit, MapareCod, NIR, LinieNIR, VerificareNIR, StocProdus, StocMinim, MiscareStoc,
                    Picking, LiniePicking, NotaLivrare,
                    TipActivitate, SablonActivitate, LinieSablon,
                    Activitate, ComentariuActivitate, Setari, AuditLog,
                    Conversatie, Mesaj, MesajCitire, chat_members)

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from googleapiclient.discovery import build as google_build
    build = google_build
    HAS_GMAIL = True
except ImportError:
    HAS_GMAIL = False
    build = None


def create_app():
    import os
    os.environ['OAUTHLIB_INSECURE_TRANSPORT'] = '1'  # Allow OAuth over HTTP for dev
    app = Flask(__name__, template_folder='templates', static_folder='static')
    app.config['SECRET_KEY'] = 'hsl-erp-v3-unified-2025'
    app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///hsl_erp.db'
    app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
    app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
        'connect_args': {'timeout': 30},
        'pool_pre_ping': True,
    }

    db.init_app(app)

    # Enable WAL mode for better concurrent access with SQLite
    with app.app_context():
        from sqlalchemy import event as sa_event
        @sa_event.listens_for(db.engine, 'connect')
        def _set_sqlite_pragma(dbapi_conn, connection_record):
            cursor = dbapi_conn.cursor()
            cursor.execute("PRAGMA journal_mode=WAL")
            cursor.execute("PRAGMA busy_timeout=5000")
            cursor.close()

    login_manager = LoginManager(app)
    login_manager.login_view = 'login'

    @login_manager.user_loader
    def load_user(uid):
        return db.session.get(Utilizator, int(uid))

    @app.context_processor
    def inject_globals():
        return {'now': datetime.now(timezone.utc), 'today': date.today(), 'Setari': Setari, 'MODULES': MODULES, 'MODULE_GROUPS': MODULE_GROUPS, 'Activitate': Activitate}

    # Track user activity
    PAGE_MODULE_MAP = {
        '/admin/clienti': 'CRM Clienți',
        '/admin/client': 'CRM Client',
        '/configurator/admin': 'Admin Configurator',
        '/configurator': 'Configurator Oferte',
        '/admin/oferte': 'Oferte',
        '/admin/comenzi': 'Comenzi',
        '/admin/facturi': 'Facturi',
        '/admin/wms': 'WMS Stoc',
        '/admin/activitati': 'Activități',
        '/admin/utilizatori': 'Utilizatori',
        '/admin/setari': 'Setări',
        '/admin/mail': 'Email',
        '/admin/roluri': 'Roluri',
        '/admin/nomenclator': 'Nomenclator',
        '/admin': 'Dashboard',
    }

    @app.before_request
    def track_user_activity():
        if current_user.is_authenticated and request.endpoint and not request.path.startswith('/static'):
            # Don't track API calls to avoid excessive DB writes
            if '/api/' not in request.path:
                now = datetime.now(timezone.utc)
                # Only update if last_seen is older than 10 seconds (throttle)
                last = current_user.last_seen
                if last and last.tzinfo is None:
                    last = last.replace(tzinfo=timezone.utc)
                if not last or (now - last).total_seconds() > 3:
                    # Determine current module from URL
                    page = request.path
                    module_name = ''
                    for prefix, name in PAGE_MODULE_MAP.items():
                        if page.startswith(prefix):
                            module_name = name
                            break
                    if not module_name and page.startswith('/admin/'):
                        module_name = page.split('/')[2].capitalize() if len(page.split('/')) > 2 else 'Admin'
                    current_user.last_seen = now
                    current_user.current_page = module_name or page
                    try:
                        db.session.commit()
                    except:
                        db.session.rollback()

    def module_required(modul):
        """Decorator: requires user to have access to the given module"""
        def decorator(f):
            @wraps(f)
            @login_required
            def decorated(*args, **kwargs):
                if not current_user.has_access(modul):
                    flash(f'Nu ai acces la acest modul.', 'error')
                    return redirect(url_for('dashboard'))
                return f(*args, **kwargs)
            return decorated
        return decorator

    # ════════════════════════════════════════════════════════════
    # AUTH
    # ════════════════════════════════════════════════════════════
    @app.route('/')
    def index():
        return redirect(url_for('dashboard')) if current_user.is_authenticated else redirect(url_for('login'))

    @app.route('/login', methods=['GET', 'POST'])
    def login():
        if current_user.is_authenticated:
            return redirect(url_for('dashboard'))
        if request.method == 'POST':
            u = Utilizator.query.filter_by(username=request.form.get('username', '').strip()).first()
            if u and u.check_password(request.form.get('password', '')):
                if not u.activ:
                    flash('Contul este dezactivat.', 'error')
                    return render_template('login.html')
                login_user(u)
                flash(f'Bine ai venit, {u.nume_complet}!', 'success')
                nxt = request.args.get('next')
                if nxt:
                    return redirect(nxt)
                # Agents go to configurator, admins to dashboard
                if u.is_admin:
                    return redirect(url_for('dashboard'))
                return redirect(url_for('configurator_page'))
            flash('Username sau parolă incorecte!', 'error')
        return render_template('login.html')

    @app.route('/logout')
    @login_required
    def logout():
        logout_user()
        return redirect(url_for('login'))

    # ════════════════════════════════════════════════════════════
    # DASHBOARD (MODULAR)
    # ════════════════════════════════════════════════════════════

    # Available widgets with their required permissions
    DASHBOARD_WIDGETS = [
        {'id': 'stat_vanzari_firma', 'name': 'Vânzări Firmă Luna', 'icon': 'bi-building', 'size': 'stat', 'requires': None},
        {'id': 'stat_vanzari_mele', 'name': 'Vânzările Mele Luna', 'icon': 'bi-person-check', 'size': 'stat', 'requires': None},
        {'id': 'stat_comision', 'name': 'Comision', 'icon': 'bi-piggy-bank', 'size': 'stat', 'requires': None},
        {'id': 'stat_comenzi', 'name': 'Comenzi Active', 'icon': 'bi-cart-check', 'size': 'stat', 'requires': 'comenzi'},
        {'id': 'stat_activitati', 'name': 'Activități Deschise', 'icon': 'bi-kanban', 'size': 'stat', 'requires': 'activitati'},
        {'id': 'stat_facturi', 'name': 'Facturi Neîncasate', 'icon': 'bi-receipt', 'size': 'stat', 'requires': 'facturi'},
        {'id': 'stat_marja_firma', 'name': 'Marjă Firmă', 'icon': 'bi-graph-up-arrow', 'size': 'stat', 'requires': 'facturi'},
        {'id': 'stat_marja_mea', 'name': 'Marja Mea', 'icon': 'bi-person-lines-fill', 'size': 'stat', 'requires': None},
        {'id': 'stat_curs', 'name': 'Curs EUR/RON', 'icon': 'bi-currency-exchange', 'size': 'stat', 'requires': None},
        {'id': 'chart_vanzari', 'name': 'Grafic Vânzări', 'icon': 'bi-graph-up', 'size': 'wide', 'requires': None},
        {'id': 'followups', 'name': 'Follow-ups de Făcut', 'icon': 'bi-bell', 'size': 'half', 'requires': 'oferte'},
        {'id': 'activitati_mele', 'name': 'Activitățile Mele', 'icon': 'bi-list-task', 'size': 'half', 'requires': 'activitati'},
        {'id': 'pipeline', 'name': 'Pipeline Oferte', 'icon': 'bi-funnel', 'size': 'half', 'requires': 'oferte'},
        {'id': 'oferte_recente', 'name': 'Oferte Recente', 'icon': 'bi-file-text', 'size': 'half', 'requires': 'oferte'},
        {'id': 'comenzi_recente', 'name': 'Comenzi Recente', 'icon': 'bi-cart', 'size': 'half', 'requires': 'comenzi'},
        {'id': 'stoc_overview', 'name': 'Stoc Depozit', 'icon': 'bi-box-seam', 'size': 'half', 'requires': 'wms'},
        {'id': 'alerte_stoc', 'name': 'Alerte Stoc Minim', 'icon': 'bi-exclamation-triangle', 'size': 'half', 'requires': 'wms'},
    ]

    DEFAULT_WIDGETS = ['stat_vanzari_firma', 'stat_vanzari_mele', 'stat_comision', 'stat_comenzi',
                       'stat_activitati', 'stat_facturi', 'stat_curs', 'chart_vanzari', 'followups',
                       'activitati_mele', 'pipeline']

    @app.route('/admin')
    @login_required
    def dashboard():
        from sqlalchemy import func
        filter_own = current_user.doar_proprii

        # === PERIOD FILTER ===
        period = request.args.get('period', 'luna_curenta')
        custom_start = request.args.get('start', '')
        custom_end = request.args.get('end', '')
        today_date = date.today()

        if period == 'luna_curenta':
            p_start = today_date.replace(day=1)
            p_end = (today_date.replace(day=28) + timedelta(days=4)).replace(day=1)  # first of next month
            p_label = today_date.strftime('%B %Y')
        elif period == 'saptamana_curenta':
            p_start = today_date - timedelta(days=today_date.weekday())
            p_end = p_start + timedelta(days=7)
            p_label = f'{p_start.strftime("%d.%m")} – {(p_end - timedelta(days=1)).strftime("%d.%m.%Y")}'
        elif period == 'ultima_saptamana':
            p_end = today_date - timedelta(days=today_date.weekday())
            p_start = p_end - timedelta(days=7)
            p_label = f'{p_start.strftime("%d.%m")} – {(p_end - timedelta(days=1)).strftime("%d.%m.%Y")}'
        elif period == 'ultima_luna':
            first_this = today_date.replace(day=1)
            p_end = first_this
            p_start = (first_this - timedelta(days=1)).replace(day=1)
            p_label = p_start.strftime('%B %Y')
        elif period == 'an_curent':
            p_start = date(today_date.year, 1, 1)
            p_end = date(today_date.year + 1, 1, 1)
            p_label = str(today_date.year)
        elif period == 'an_trecut':
            p_start = date(today_date.year - 1, 1, 1)
            p_end = date(today_date.year, 1, 1)
            p_label = str(today_date.year - 1)
        elif period == 'custom' and custom_start and custom_end:
            try:
                p_start = datetime.strptime(custom_start, '%Y-%m-%d').date()
                p_end = datetime.strptime(custom_end, '%Y-%m-%d').date() + timedelta(days=1)
                p_label = f'{p_start.strftime("%d.%m.%Y")} – {(p_end - timedelta(days=1)).strftime("%d.%m.%Y")}'
            except:
                p_start = today_date.replace(day=1)
                p_end = (today_date.replace(day=28) + timedelta(days=4)).replace(day=1)
                p_label = today_date.strftime('%B %Y')
                period = 'luna_curenta'
        else:
            p_start = today_date.replace(day=1)
            p_end = (today_date.replace(day=28) + timedelta(days=4)).replace(day=1)
            p_label = today_date.strftime('%B %Y')
            period = 'luna_curenta'

        # Get user's widget config or defaults
        user_config = current_user.dashboard_config
        if user_config:
            active_widget_ids = user_config.get('widgets', DEFAULT_WIDGETS)
        else:
            active_widget_ids = list(DEFAULT_WIDGETS)

        # Filter widgets by permissions
        available_widgets = []
        for w in DASHBOARD_WIDGETS:
            if w['requires'] is None or current_user.has_access(w['requires']):
                available_widgets.append(w)
        available_ids = {w['id'] for w in available_widgets}
        active_widget_ids = [wid for wid in active_widget_ids if wid in available_ids]

        # Build data for active widgets
        data = {}
        active_set = set(active_widget_ids)

        of_query = Oferta.query
        cmd_query = Comanda.query
        if filter_own:
            of_query = of_query.filter_by(creat_de_id=current_user.id)
            cmd_query = cmd_query.filter_by(creat_de_id=current_user.id)

        # Stats
        if 'stat_vanzari_firma' in active_set:
            data['vanzari_luna'] = db.session.query(func.sum(Comanda.total)).filter(
                Comanda.status != 'anulat', Comanda.data_comanda >= p_start, Comanda.data_comanda < p_end
            ).scalar() or 0
        if 'stat_vanzari_mele' in active_set or 'stat_comision' in active_set:
            data['vanzari_proprii'] = db.session.query(func.coalesce(func.sum(Comanda.total), 0)).filter(
                Comanda.status != 'anulat', Comanda.creat_de_id == current_user.id,
                Comanda.data_comanda >= p_start, Comanda.data_comanda < p_end
            ).scalar() or 0
            data['comision_procent'] = current_user.comision_procent or 0
            data['comision_luna'] = round(float(data['vanzari_proprii']) * data['comision_procent'] / 100, 2)
        if 'stat_comenzi' in active_set:
            data['comenzi_active'] = cmd_query.filter(
                Comanda.status.notin_(['finalizata', 'anulat']),
                Comanda.data_comanda >= p_start, Comanda.data_comanda < p_end
            ).count()
        if 'stat_activitati' in active_set:
            aq_stat = Activitate.query.filter(
                Activitate.status.in_(['de_facut', 'in_lucru', 'in_asteptare']),
                *([Activitate.asignat_id == current_user.id] if filter_own else [])
            )
            if period != 'luna_curenta':  # filter by creation date for non-default periods
                aq_stat = aq_stat.filter(Activitate.data_creare >= datetime.combine(p_start, datetime.min.time()),
                                         Activitate.data_creare < datetime.combine(p_end, datetime.min.time()))
            data['activitati_deschise'] = aq_stat.count()
        if 'stat_facturi' in active_set:
            fq_pf = Factura.query.filter(Factura.tip == 'proforma', Factura.status.in_(['emisa', 'trimisa']))
            fq_fc = Factura.query.filter(Factura.tip == 'fiscala', Factura.status.in_(['emisa', 'trimisa']))
            if filter_own:
                data['facturi_neincasate'] = 0
                data['proforme_neplatite'] = 0
            else:
                data['facturi_neincasate'] = fq_fc.count()
                data['proforme_neplatite'] = fq_pf.count()

        # Curs valutar
        if 'stat_curs' in active_set:
            try:
                from curs_service import get_curs_for_date
                from models import CursValutar
                # Check if single day selected (p_end == p_start + 1 day)
                is_single_day = (p_end - p_start).days == 1
                target_date = p_start if is_single_day else date.today()
                
                curs, curs_bnr_val = get_curs_for_date(target_date, 'EUR')
                cached = CursValutar.query.filter_by(data=target_date, moneda='EUR').first()
                
                data['curs_bt'] = curs or 0
                data['curs_bnr'] = curs_bnr_val or 0
                data['curs_data'] = target_date.strftime('%d.%m.%Y')
                data['curs_sursa'] = cached.sursa if cached else ('bnr' if curs else 'n/a')
            except Exception as ex:
                log.error(f'Curs error: {ex}')
                data['curs_bt'] = 0
                data['curs_bnr'] = 0
                data['curs_sursa'] = 'eroare'
                data['curs_data'] = ''

        # Marja — calc from comenzi in period
        if 'stat_marja_firma' in active_set or 'stat_marja_mea' in active_set:
            # Build acquisition price map
            _pa_map = {}
            for row in db.session.query(StocProdus.cod_intern, func.avg(StocProdus.pret_achizitie_mediu)).group_by(StocProdus.cod_intern).all():
                _pa_map[row[0]] = row[1] or 0
            for row in db.session.query(LinieNIR.cod_intern, func.avg(LinieNIR.pret_achizitie)).group_by(LinieNIR.cod_intern).all():
                if row[0] not in _pa_map: _pa_map[row[0]] = row[1] or 0

            _cmds = Comanda.query.filter(Comanda.status != 'anulat',
                                          Comanda.data_comanda >= p_start, Comanda.data_comanda < p_end).all()
            mf_vanzare, mf_cost = 0, 0
            mm_vanzare, mm_cost = 0, 0
            for _c in _cmds:
                for _l in _c.linii:
                    v = _l.valoare_linie
                    c_u = _pa_map.get(_l.cod, 0)
                    c_t = _l.cantitate * c_u
                    mf_vanzare += v; mf_cost += c_t
                    if _c.creat_de_id == current_user.id:
                        mm_vanzare += v; mm_cost += c_t

            if 'stat_marja_firma' in active_set:
                data['marja_firma'] = round(mf_vanzare - mf_cost, 2)
                data['marja_firma_pct'] = round((mf_vanzare - mf_cost) / mf_vanzare * 100, 1) if mf_vanzare else 0
            if 'stat_marja_mea' in active_set:
                data['marja_mea'] = round(mm_vanzare - mm_cost, 2)
                data['marja_mea_pct'] = round((mm_vanzare - mm_cost) / mm_vanzare * 100, 1) if mm_vanzare else 0

        # Chart — always shows 6 months relative to period end
        if 'chart_vanzari' in active_set:
            vanzari_lunare = []
            # Generate months that fall within [p_start, p_end)
            # Start from the first of p_start's month
            cursor = p_start.replace(day=1)
            while cursor < p_end:
                m_start = cursor
                if cursor.month == 12:
                    m_end = date(cursor.year + 1, 1, 1)
                else:
                    m_end = date(cursor.year, cursor.month + 1, 1)
                # Clamp to period
                q_start = max(m_start, p_start)
                q_end = min(m_end, p_end)
                
                total = db.session.query(func.coalesce(func.sum(Comanda.total), 0)).filter(
                    Comanda.status != 'anulat', Comanda.data_comanda >= q_start, Comanda.data_comanda < q_end
                ).scalar() or 0
                proprii = db.session.query(func.coalesce(func.sum(Comanda.total), 0)).filter(
                    Comanda.status != 'anulat', Comanda.creat_de_id == current_user.id,
                    Comanda.data_comanda >= q_start, Comanda.data_comanda < q_end
                ).scalar() or 0
                vanzari_lunare.append({
                    'luna': cursor.strftime('%b %Y'),
                    'total': round(float(total), 2),
                    'proprii': round(float(proprii), 2)
                })
                cursor = m_end
            data['vanzari_lunare'] = vanzari_lunare

        # Follow-ups — always today (not affected by period)
        if 'followups' in active_set:
            fq = db.session.query(FollowUpOferta, Oferta).join(Oferta, FollowUpOferta.oferta_id == Oferta.id).filter(
                FollowUpOferta.next_date != None, FollowUpOferta.next_date <= date.today(),
                Oferta.status.in_(['draft', 'trimisa']))
            if filter_own: fq = fq.filter(Oferta.creat_de_id == current_user.id)
            data['followups_due'] = fq.order_by(FollowUpOferta.next_date.asc()).limit(10).all()

        # Activitati — always current open (not affected by period)
        if 'activitati_mele' in active_set:
            aq = Activitate.query.filter(Activitate.status.in_(['de_facut', 'in_lucru', 'in_asteptare']))
            if filter_own: aq = aq.filter(db.or_(Activitate.asignat_id == current_user.id, Activitate.creat_de_id == current_user.id))
            data['activitati'] = aq.order_by(
                db.case((Activitate.prioritate == 'urgenta', 0), (Activitate.prioritate == 'ridicata', 1),
                         (Activitate.prioritate == 'normala', 2), else_=3),
                Activitate.deadline.asc().nullslast()
            ).limit(8).all()

        # Pipeline — filtered by period
        if 'pipeline' in active_set:
            pq = db.session.query(Oferta.status, func.count(Oferta.id), func.coalesce(func.sum(Oferta.total), 0)).filter(
                Oferta.data_creare >= datetime.combine(p_start, datetime.min.time()),
                Oferta.data_creare < datetime.combine(p_end, datetime.min.time()))
            if filter_own: pq = pq.filter(Oferta.creat_de_id == current_user.id)
            pipeline_data = pq.group_by(Oferta.status).all()
            data['pipeline'] = {s: {'count': c, 'total': t} for s, c, t in pipeline_data}

        # Oferte recente — filtered by period
        if 'oferte_recente' in active_set:
            oq = of_query.filter(Oferta.data_creare >= datetime.combine(p_start, datetime.min.time()),
                                  Oferta.data_creare < datetime.combine(p_end, datetime.min.time()))
            data['oferte_recente'] = oq.order_by(Oferta.data_creare.desc()).limit(5).all()

        # Comenzi recente — filtered by period
        if 'comenzi_recente' in active_set:
            cq = cmd_query.filter(Comanda.data_comanda >= p_start, Comanda.data_comanda < p_end)
            data['comenzi_recente'] = cq.order_by(Comanda.data_creare.desc()).limit(5).all()

        # Stoc overview
        if 'stoc_overview' in active_set:
            data['stoc_top'] = db.session.query(
                StocProdus.cod_intern, db.func.max(StocProdus.denumire).label('den'),
                db.func.sum(StocProdus.cantitate).label('qty'),
                db.func.sum(StocProdus.cantitate * StocProdus.pret_achizitie_mediu).label('val')
            ).group_by(StocProdus.cod_intern).order_by(db.text('val DESC')).limit(10).all()

        # Alerte stoc minim
        if 'alerte_stoc' in active_set:
            praguri = StocMinim.query.filter_by(activ=True).all()
            data['alerte_stoc'] = [p for p in praguri if p.sub_prag]

        return render_template('admin/dashboard.html',
                             active_widgets=active_widget_ids,
                             available_widgets=available_widgets,
                             data=data,
                             period=period, p_label=p_label,
                             p_start=p_start.strftime('%Y-%m-%d'),
                             p_end=(p_end - timedelta(days=1)).strftime('%Y-%m-%d'))

    @app.route('/api/dashboard/save-config', methods=['POST'])
    @login_required
    def api_dashboard_save_config():
        d = request.get_json()
        widgets = d.get('widgets', [])
        current_user.dashboard_config = {'widgets': widgets}
        db.session.commit()
        return jsonify({'success': True})

    # ════════════════════════════════════════════════════════════
    # GLOBAL SEARCH
    # ════════════════════════════════════════════════════════════
    @app.route('/api/search')
    @login_required
    def api_global_search():
        q = request.args.get('q', '').strip()
        if len(q) < 2:
            return jsonify({'results': []})
        results = []
        like = f'%{q}%'
        limit = 5

        # Clients
        if current_user.has_access('crm'):
            for c in Client.query.filter(
                db.or_(Client.nume.ilike(like), Client.cui.ilike(like), Client.email.ilike(like), Client.telefon.ilike(like))
            ).limit(limit).all():
                results.append({'type': 'client', 'title': c.nume, 'subtitle': f'{c.cui or ""} · {c.email or ""}',
                               'url': f'/admin/clienti/{c.id}/detalii'})

        # Oferte
        if current_user.has_access('oferte'):
            of_q = Oferta.query.filter(db.or_(Oferta.numar.ilike(like), Oferta.observatii.ilike(like)))
            if current_user.doar_proprii: of_q = of_q.filter(Oferta.creat_de_id == current_user.id)
            for o in of_q.limit(limit).all():
                results.append({'type': 'oferta', 'title': o.numar, 'subtitle': f'{o.client.nume if o.client else ""} · {dict(Oferta.STATUSES).get(o.status, o.status)} · {"{:,.0f}".format(o.total)}€',
                               'url': f'/admin/oferte/{o.id}'})

        # Comenzi
        if current_user.has_access('comenzi'):
            cm_q = Comanda.query.filter(db.or_(Comanda.numar.ilike(like), Comanda.observatii.ilike(like)))
            if current_user.doar_proprii: cm_q = cm_q.filter(Comanda.creat_de_id == current_user.id)
            for c in cm_q.limit(limit).all():
                results.append({'type': 'comanda', 'title': c.numar, 'subtitle': f'{c.client.nume if c.client else ""} · {c.status_display} · {"{:,.0f}".format(c.total)}€',
                               'url': f'/admin/comenzi/{c.id}'})

        # Produse
        if current_user.has_access('nomenclator'):
            for p in ProdusConfig.query.filter(
                db.or_(ProdusConfig.cod.ilike(like), ProdusConfig.denumire.ilike(like))
            ).limit(limit).all():
                results.append({'type': 'produs', 'title': p.cod, 'subtitle': p.denumire,
                               'url': f'/admin/nomenclator/produs/{p.id}'})

        # NIR-uri
        if current_user.has_access('wms'):
            for n in NIR.query.filter(db.or_(NIR.numar_factura_furnizor.ilike(like), NIR.numar.ilike(like))).limit(limit).all():
                results.append({'type': 'nir', 'title': n.numar, 'subtitle': f'{n.furnizor.nume if n.furnizor else ""} · {n.numar_factura_furnizor or ""}',
                               'url': f'/admin/wms/nir/{n.id}'})

        # Facturi
        if current_user.has_access('facturi'):
            for f in Factura.query.filter(
                db.or_(Factura.serie.ilike(like), db.cast(Factura.numar, db.String).ilike(like))
            ).limit(limit).all():
                results.append({'type': 'factura', 'title': f'{"PF" if f.tip=="proforma" else "FC"} {f.numar_complet}',
                               'subtitle': f'{f.client.nume if f.client else ""} · {f.status}',
                               'url': f'/admin/facturi/{f.id}'})

        # Activități
        for a in Activitate.query.filter(
            db.or_(Activitate.titlu.ilike(like), Activitate.descriere.ilike(like))
        ).limit(limit).all():
            results.append({'type': 'activitate', 'title': a.titlu,
                           'subtitle': f'{a.asignat.nume_complet if a.asignat else ""} · {a.status or ""}',
                           'url': f'/admin/activitati/{a.id}'})

        # Also search clients by name in oferte/comenzi
        if current_user.has_access('oferte') and len(results) < 15:
            for o in Oferta.query.join(Client).filter(Client.nume.ilike(like)).limit(3).all():
                results.append({'type': 'oferta', 'title': o.numar, 'subtitle': f'{o.client.nume} · {dict(Oferta.STATUSES).get(o.status, o.status)}',
                               'url': f'/admin/oferte/{o.id}'})

        return jsonify({'results': results[:20]})

    # ════════════════════════════════════════════════════════════
    # AUDIT LOG
    # ════════════════════════════════════════════════════════════
    @app.route('/admin/audit-log')
    @login_required
    def audit_log_page():
        filter_tip = request.args.get('tip', '')
        q = AuditLog.query
        if filter_tip:
            q = q.filter_by(tip_document=filter_tip)
        entries = q.order_by(AuditLog.data.desc()).limit(200).all()
        return render_template('admin/audit_log.html', entries=entries, filter_tip=filter_tip)

    @app.route('/api/audit-log/<tip>/<int:doc_id>')
    @login_required
    def api_audit_log(tip, doc_id):
        entries = AuditLog.get_for(tip, doc_id)
        return jsonify({'entries': [
            {'actiune': e.actiune, 'detalii': e.detalii, 'utilizator': e.utilizator.nume_complet if e.utilizator else '?',
             'data': e.data.strftime('%d.%m.%Y %H:%M')} for e in entries
        ]})

    # ════════════════════════════════════════════════════════════
    # CHAT SYSTEM
    # ════════════════════════════════════════════════════════════

    @app.route('/admin/chat')
    @login_required
    def chat_page():
        """Full chat page (WhatsApp-style)"""
        users = Utilizator.query.filter_by(activ=True).filter(Utilizator.id != current_user.id).all()
        users_json = [{'id': u.id, 'nume_complet': u.nume_complet} for u in users]
        return render_template('admin/chat.html', users=users_json)

    @app.route('/api/chat/conversatii')
    @login_required
    def api_chat_conversatii():
        """List user's conversations with unread counts"""
        convs = Conversatie.query.filter(
            Conversatie.activ == True,
            Conversatie.membri.any(id=current_user.id)
        ).all()
        convs.sort(key=lambda c: c.ultimul_mesaj.data_trimitere if c.ultimul_mesaj else c.data_creare, reverse=True)
        result = []
        for c in convs:
            um = c.ultimul_mesaj
            result.append({
                'id': c.id, 'tip': c.tip, 'nume': c.display_name_for(current_user),
                'necitite': c.necitite_pentru(current_user.id),
                'ultimul_mesaj': um.text[:50] if um and um.text else ('📎 ' + um.fisier_nume[:30] if um and um.fisier_nume else ''),
                'ultimul_autor': um.autor.nume_complet if um and um.autor else '',
                'ultimul_timp': um.data_trimitere.strftime('%H:%M') if um else '',
                'doc_tip': c.doc_tip, 'doc_id': c.doc_id, 'doc_numar': c.doc_numar,
            })
        total_necitite = sum(r['necitite'] for r in result)
        return jsonify({'conversatii': result, 'total_necitite': total_necitite})

    @app.route('/api/chat/mesaje/<int:conv_id>')
    @login_required
    def api_chat_mesaje(conv_id):
        """Get messages for a conversation, optionally only new ones after a given ID"""
        conv = Conversatie.query.get_or_404(conv_id)
        if current_user not in conv.membri:
            return jsonify({'error': 'Nu ești membru'}), 403
        after_id = request.args.get('after', 0, type=int)
        # Mark as read
        for m in Mesaj.query.filter_by(conversatie_id=conv_id).filter(Mesaj.autor_id != current_user.id).all():
            m.marcheaza_citit(current_user.id)
        db.session.commit()
        query = Mesaj.query.filter_by(conversatie_id=conv_id)
        if after_id:
            query = query.filter(Mesaj.id > after_id)
        mesaje = query.order_by(Mesaj.data_trimitere.asc()).all()

        # Build read receipts summary per message
        def get_reads(msg):
            if msg.autor_id != current_user.id:
                return None  # Only show read info for own messages
            reads = MesajCitire.query.filter_by(mesaj_id=msg.id).filter(
                MesajCitire.utilizator_id != msg.autor_id
            ).all()
            if not reads:
                return {'status': 'sent', 'readers': []}
            return {
                'status': 'read',
                'readers': [{'nume': r.utilizator.nume_complet, 'ora': r.data_citire.strftime('%H:%M'),
                             'data': r.data_citire.strftime('%d.%m.%Y')} for r in reads]
            }

        return jsonify({'mesaje': [
            {'id': m.id, 'autor': m.autor.nume_complet, 'autor_id': m.autor_id,
             'text': m.text or '', 'ora': m.data_trimitere.strftime('%H:%M'),
             'data': m.data_trimitere.strftime('%d.%m.%Y'), 'mine': m.autor_id == current_user.id,
             'fisier': {'nume': m.fisier_nume, 'path': m.fisier_path, 'size': m.fisier_size,
                        'tip': m.fisier_tip} if m.fisier_path else None,
             'citit': get_reads(m),
             'reply': {'id': m.reply_to.id, 'autor': m.reply_to.autor.nume_complet,
                        'text': (m.reply_to.text or '')[:80],
                        'fisier': m.reply_to.fisier_nume} if m.reply_to else None}
            for m in mesaje
        ], 'conv_nume': conv.display_name_for(current_user), 'conv_tip': conv.tip})

    @app.route('/api/chat/citiri/<int:msg_id>')
    @login_required
    def api_chat_citiri(msg_id):
        """Detailed read receipts for a specific message"""
        msg = Mesaj.query.get_or_404(msg_id)
        conv = Conversatie.query.get(msg.conversatie_id)
        if current_user not in conv.membri:
            return jsonify({'error': 'Acces interzis'}), 403
        reads = MesajCitire.query.filter_by(mesaj_id=msg_id).filter(
            MesajCitire.utilizator_id != msg.autor_id
        ).order_by(MesajCitire.data_citire.asc()).all()
        # Who hasn't read yet
        all_members = [m for m in conv.membri if m.id != msg.autor_id]
        read_ids = {r.utilizator_id for r in reads}
        unread = [m for m in all_members if m.id not in read_ids]
        return jsonify({
            'citit': [{'nume': r.utilizator.nume_complet, 'data': r.data_citire.strftime('%d.%m.%Y %H:%M')} for r in reads],
            'necitit': [{'nume': u.nume_complet} for u in unread],
            'text_preview': (msg.text or '')[:50]
        })

    @app.route('/api/chat/trimite', methods=['POST'])
    @login_required
    def api_chat_trimite():
        """Send a message (text or file)"""
        # Handle both JSON and multipart/form-data
        if request.content_type and 'json' in request.content_type:
            d = request.get_json() or {}
            conv_id = d.get('conversatie_id')
            text = d.get('text', '').strip()
            reply_to_id = d.get('reply_to_id')
            file = None
        else:
            conv_id = request.form.get('conversatie_id')
            text = request.form.get('text', '').strip()
            reply_to_id = request.form.get('reply_to_id')
            file = request.files.get('fisier')

        if not conv_id:
            return jsonify({'error': 'Lipsă conversatie_id'}), 400
        conv = Conversatie.query.get_or_404(int(conv_id))
        if current_user not in conv.membri:
            return jsonify({'error': 'Nu ești membru'}), 403

        fisier_nume = fisier_path = fisier_tip = None
        fisier_size = 0

        if file and file.filename:
            import os, uuid
            upload_dir = os.path.join(app.root_path, 'static', 'chat_files')
            os.makedirs(upload_dir, exist_ok=True)
            ext = os.path.splitext(file.filename)[1].lower()
            safe_name = f'{uuid.uuid4().hex[:12]}{ext}'
            path = os.path.join(upload_dir, safe_name)
            file.save(path)
            fisier_nume = file.filename
            fisier_path = f'/static/chat_files/{safe_name}'
            fisier_size = os.path.getsize(path)
            img_exts = {'.jpg','.jpeg','.png','.gif','.webp','.bmp','.svg'}
            doc_exts = {'.pdf','.doc','.docx','.xls','.xlsx','.ppt','.pptx','.csv','.txt'}
            if ext in img_exts:
                fisier_tip = 'image'
            elif ext in doc_exts:
                fisier_tip = 'document'
            else:
                fisier_tip = 'other'

        if not text and not fisier_path:
            return jsonify({'error': 'Mesaj gol'}), 400

        msg = Mesaj(conversatie_id=int(conv_id), autor_id=current_user.id, text=text or '',
                    reply_to_id=int(reply_to_id) if reply_to_id else None,
                    fisier_nume=fisier_nume, fisier_path=fisier_path,
                    fisier_size=fisier_size, fisier_tip=fisier_tip)
        msg.marcheaza_citit(current_user.id)
        db.session.add(msg)
        db.session.commit()
        return jsonify({'success': True, 'mesaj_id': msg.id,
                        'ora': msg.data_trimitere.strftime('%H:%M')})

    @app.route('/api/chat/nou', methods=['POST'])
    @login_required
    def api_chat_nou():
        """Create new conversation"""
        d = request.get_json()
        tip = d.get('tip', 'direct')
        membri_ids = d.get('membri', [])

        if tip == 'direct' and len(membri_ids) == 1:
            # Check if direct conversation already exists
            other_id = membri_ids[0]
            for conv in current_user.conversatii_chat:
                if conv.tip == 'direct' and conv.activ:
                    member_ids = [m.id for m in conv.membri]
                    if other_id in member_ids and current_user.id in member_ids and len(member_ids) == 2:
                        return jsonify({'success': True, 'conversatie_id': conv.id, 'existent': True})

        conv = Conversatie(
            tip=tip, creat_de_id=current_user.id,
            nume=d.get('nume', ''),
            doc_tip=d.get('doc_tip'), doc_id=d.get('doc_id'), doc_numar=d.get('doc_numar')
        )
        db.session.add(conv)
        db.session.flush()
        # Add members
        conv.membri.append(current_user)
        for uid in membri_ids:
            u = Utilizator.query.get(uid)
            if u and u not in conv.membri:
                conv.membri.append(u)
        db.session.commit()
        return jsonify({'success': True, 'conversatie_id': conv.id})

    @app.route('/api/chat/document', methods=['POST'])
    @login_required
    def api_chat_document():
        """Get or create conversation for a document"""
        d = request.get_json()
        doc_tip = d.get('doc_tip')
        doc_id = d.get('doc_id')
        doc_numar = d.get('doc_numar', '')
        # Find existing
        conv = Conversatie.query.filter_by(doc_tip=doc_tip, doc_id=doc_id, activ=True).first()
        if conv:
            if current_user not in conv.membri:
                conv.membri.append(current_user)
                db.session.commit()
            return jsonify({'success': True, 'conversatie_id': conv.id, 'existent': True})
        # Create new
        conv = Conversatie(tip='document', creat_de_id=current_user.id,
                          doc_tip=doc_tip, doc_id=doc_id, doc_numar=doc_numar,
                          nume=f'{doc_tip.title()} {doc_numar}')
        db.session.add(conv)
        db.session.flush()
        conv.membri.append(current_user)
        # Add the document's creator too
        if doc_tip == 'comanda':
            cmd = Comanda.query.get(doc_id)
            if cmd and cmd.creat_de and cmd.creat_de not in conv.membri:
                conv.membri.append(cmd.creat_de)
        elif doc_tip == 'oferta':
            ofe = Oferta.query.get(doc_id)
            if ofe and ofe.creat_de and ofe.creat_de not in conv.membri:
                conv.membri.append(ofe.creat_de)
        db.session.commit()
        return jsonify({'success': True, 'conversatie_id': conv.id})

    @app.route('/api/chat/necitite')
    @login_required
    def api_chat_necitite():
        """Quick unread count for badge"""
        total = 0
        for conv in current_user.conversatii_chat:
            if conv.activ:
                total += conv.necitite_pentru(current_user.id)
        return jsonify({'total': total})

    @app.route('/api/chat/conversatie/<int:conv_id>/sterge', methods=['POST'])
    @login_required
    def api_chat_sterge_conv(conv_id):
        """Delete (deactivate) a conversation"""
        conv = Conversatie.query.get_or_404(conv_id)
        if current_user not in conv.membri:
            return jsonify({'error': 'Nu ești membru'}), 403
        conv.activ = False
        db.session.commit()
        return jsonify({'success': True})

    @app.route('/api/chat/mesaj/<int:msg_id>/sterge', methods=['POST'])
    @login_required
    def api_chat_sterge_mesaj(msg_id):
        """Delete a message (only own messages or admin)"""
        msg = Mesaj.query.get_or_404(msg_id)
        conv = Conversatie.query.get(msg.conversatie_id)
        if current_user not in conv.membri:
            return jsonify({'error': 'Nu ești membru'}), 403
        if msg.autor_id != current_user.id and current_user.username != 'admin':
            return jsonify({'error': 'Poți șterge doar mesajele tale'}), 403
        # Delete file if exists
        if msg.fisier_path:
            import os
            fpath = os.path.join(app.root_path, msg.fisier_path.lstrip('/'))
            if os.path.exists(fpath):
                os.remove(fpath)
        db.session.delete(msg)
        db.session.commit()
        return jsonify({'success': True})

    @app.route('/api/chat/resolve', methods=['POST'])
    @login_required
    def api_chat_resolve():
        """Resolve document references by type + search key"""
        d = request.get_json()
        refs = d.get('refs', [])
        results = []
        for ref in refs:
            if isinstance(ref, str):
                # Legacy format: skip
                continue
            tip = ref.get('type', '')
            key = ref.get('title', '').strip()
            if not tip or not key:
                continue
            card = None
            if tip == 'comanda':
                obj = Comanda.query.filter_by(numar=key).first()
                if not obj:
                    obj = Comanda.query.filter(Comanda.numar.ilike(f'%{key}%')).first()
                if obj:
                    card = {'type': 'comanda', 'icon': 'bi-cart-check', 'color': '#0d6efd',
                            'titlu': f'Comandă {obj.numar}',
                            'subtitlu': obj.client.nume if obj.client else '',
                            'status': obj.status.title(), 'total': f'{obj.total:,.0f} {obj.moneda}',
                            'url': f'/admin/comenzi/{obj.id}'}
            elif tip == 'oferta':
                obj = Oferta.query.filter_by(numar=key).first()
                if not obj:
                    obj = Oferta.query.filter(Oferta.numar.ilike(f'%{key}%')).first()
                if obj:
                    card = {'type': 'oferta', 'icon': 'bi-file-earmark-text', 'color': '#fd7e14',
                            'titlu': f'Ofertă {obj.numar}',
                            'subtitlu': obj.client.nume if obj.client else '',
                            'status': obj.status.title() if obj.status else '',
                            'total': f'{obj.total:,.0f} {obj.moneda}' if obj.total else '',
                            'url': f'/admin/oferte/{obj.id}'}
            elif tip == 'produs':
                obj = ProdusConfig.query.filter_by(cod=key).first()
                if not obj:
                    obj = ProdusConfig.query.filter(
                        db.or_(ProdusConfig.cod.ilike(f'%{key}%'), ProdusConfig.denumire.ilike(f'%{key}%'))
                    ).first()
                if obj:
                    card = {'type': 'produs', 'icon': 'bi-box', 'color': '#6f42c1',
                            'titlu': obj.denumire,
                            'subtitlu': f'Cod: {obj.cod}',
                            'status': 'Activ' if obj.activ else 'Inactiv',
                            'total': f'{obj.pret:,.2f} €' if obj.pret else '',
                            'url': f'/admin/nomenclator/produs/{obj.id}'}
            elif tip == 'client':
                obj = Client.query.filter(
                    db.or_(Client.nume.ilike(f'%{key}%'), Client.cui.ilike(f'%{key}%'))
                ).first()
                if obj:
                    card = {'type': 'client', 'icon': 'bi-people', 'color': '#61993B',
                            'titlu': obj.nume,
                            'subtitlu': obj.cui or obj.email or '',
                            'status': obj.tara or '',
                            'total': '',
                            'url': f'/admin/clienti/{obj.id}/detalii'}
            elif tip == 'activitate':
                try:
                    obj = Activitate.query.get(int(key))
                except (ValueError, TypeError):
                    obj = Activitate.query.filter(Activitate.titlu.ilike(f'%{key}%')).first()
                if obj:
                    card = {'type': 'activitate', 'icon': 'bi-list-check', 'color': '#20c997',
                            'titlu': obj.titlu,
                            'subtitlu': obj.asignat.nume_complet if obj.asignat else '',
                            'status': obj.status.title() if obj.status else '',
                            'total': '',
                            'url': f'/admin/activitati/{obj.id}'}
            elif tip == 'factura':
                obj = Factura.query.filter(
                    db.or_(Factura.serie.ilike(f'%{key}%'), db.cast(Factura.numar, db.String).ilike(f'%{key}%'))
                ).first()
                if obj:
                    card = {'type': 'factura', 'icon': 'bi-receipt', 'color': '#dc3545',
                            'titlu': f'Factură {obj.serie}/{obj.numar}',
                            'subtitlu': obj.client.nume if obj.client else '',
                            'status': obj.status.title() if obj.status else '',
                            'total': f'{obj.total:,.0f} RON' if obj.total else '',
                            'url': f'/admin/facturi/{obj.id}'}
            elif tip == 'nir':
                obj = NIR.query.filter_by(numar=key).first()
                if not obj:
                    obj = NIR.query.filter(NIR.numar.ilike(f'%{key}%')).first()
                if obj:
                    card = {'type': 'nir', 'icon': 'bi-box-arrow-in-down', 'color': '#0dcaf0',
                            'titlu': f'NIR {obj.numar}',
                            'subtitlu': obj.furnizor.nume if obj.furnizor else '',
                            'status': obj.status.title() if obj.status else '',
                            'total': '',
                            'url': f'/admin/wms/nir/{obj.id}'}
            if card:
                card['search_key'] = key
                results.append(card)
        return jsonify({'cards': results})

    # ════════════════════════════════════════════════════════════
    # MAIL INTEGRATION
    # ════════════════════════════════════════════════════════════
    from models import ContMail, MailThread, MailMesaj

    @app.route('/admin/mail')
    @login_required
    def mail_page():
        """Main mail interface"""
        conturi = ContMail.query.filter_by(utilizator_id=current_user.id, activ=True).all()
        users = Utilizator.query.filter_by(activ=True).all()
        users_json = [{'id': u.id, 'nume_complet': u.nume_complet} for u in users]
        return render_template('admin/mail.html', conturi=conturi, users=users_json)

    @app.route('/admin/mail/connect')
    @login_required
    def mail_connect():
        """Start OAuth2 flow to connect Gmail account"""
        import os
        os.environ['OAUTHLIB_INSECURE_TRANSPORT'] = '1'  # Allow HTTP for localhost dev
        try:
            from mail_service import get_oauth_flow
            # Use X-Forwarded-Proto for Railway/proxy, fallback to request scheme
            scheme = request.headers.get('X-Forwarded-Proto', request.scheme)
            redirect_uri = f"{scheme}://{request.host}/admin/mail/callback"
            flow = get_oauth_flow(app, redirect_uri)
            auth_url, state = flow.authorization_url(
                access_type='offline', prompt='consent',
                login_hint=current_user.email
            )
            from flask import session as flask_session
            flask_session['oauth_state'] = state
            flask_session['oauth_redirect_uri'] = redirect_uri  # Save for callback
            return redirect(auth_url)
        except FileNotFoundError as e:
            flash(str(e), 'danger')
            return redirect('/admin/mail')
        except Exception as e:
            flash(f'Eroare la conectarea mail: {e}. Verificați dacă serverul are acces la internet.', 'danger')
            return redirect('/admin/mail')

    @app.route('/admin/mail/callback')
    @login_required
    def mail_callback():
        """OAuth2 callback from Google"""
        import os
        os.environ['OAUTHLIB_INSECURE_TRANSPORT'] = '1'
        try:
            from mail_service import get_oauth_flow
            from flask import session as flask_session
            # Use saved redirect_uri to match exactly what was sent to Google
            scheme = request.headers.get('X-Forwarded-Proto', request.scheme)
            redirect_uri = flask_session.pop('oauth_redirect_uri', f"{scheme}://{request.host}/admin/mail/callback")
            flow = get_oauth_flow(app, redirect_uri)
            # Rewrite authorization_response to use correct scheme
            auth_response = request.url
            if scheme == 'https' and auth_response.startswith('http://'):
                auth_response = 'https://' + auth_response[7:]
            flow.fetch_token(authorization_response=auth_response)
            creds = flow.credentials
            # Get user email
            service = build('gmail', 'v1', credentials=creds)
            profile = service.users().getProfile(userId='me').execute()
            email_addr = profile.get('emailAddress', '')
            # Save or update
            cont = ContMail.query.filter_by(utilizator_id=current_user.id, email=email_addr).first()
            if not cont:
                tip = 'personal'
                if 'office' in email_addr.lower():
                    tip = 'office'
                elif 'vanzari' in email_addr.lower():
                    tip = 'monitorizare'
                cont = ContMail(utilizator_id=current_user.id, email=email_addr, tip=tip)
                db.session.add(cont)
            cont.access_token = creds.token
            cont.refresh_token = creds.refresh_token
            cont.token_expiry = creds.expiry
            cont.activ = True
            cont.data_conectare = datetime.now(timezone.utc)
            db.session.commit()
            flash(f'Cont {email_addr} conectat cu succes!', 'success')
        except Exception as e:
            flash(f'Eroare la autentificare Gmail: {e}', 'danger')
        return redirect('/admin/mail')

    @app.route('/admin/mail/disconnect/<int:cont_id>', methods=['POST'])
    @login_required
    def mail_disconnect(cont_id):
        """Disconnect a mail account"""
        cont = ContMail.query.get_or_404(cont_id)
        if cont.utilizator_id != current_user.id and current_user.username != 'admin':
            return jsonify({'error': 'Nu ai permisiuni'}), 403
        cont.activ = False
        db.session.commit()
        return jsonify({'success': True})

    @app.route('/api/mail/sync', methods=['POST'])
    @login_required
    def api_mail_sync():
        """Trigger manual sync - runs in background thread.
        DB locks are minimal because mail_service commits per message."""
        import threading
        user_id = current_user.id
        
        def _do_sync(app, user_id, deep):
            with app.app_context():
                try:
                    from mail_service import sync_inbox
                    conturi = ContMail.query.filter_by(utilizator_id=user_id, activ=True).all()
                    office = ContMail.query.filter_by(tip='office', activ=True).first()
                    if office and office not in conturi:
                        conturi.append(office)
                    for cont in conturi:
                        try:
                            pages = 3 if deep else 1
                            sync_inbox(cont, max_results=50, max_pages=pages)
                        except Exception as e:
                            log.error(f'Sync error for {cont.email}: {e}')
                            db.session.rollback()
                except Exception as e:
                    log.error(f'Background sync error: {e}')
        
        d = request.get_json(silent=True) or {}
        deep = d.get('deep', False)
        t = threading.Thread(target=_do_sync, args=(app, user_id, deep), daemon=True)
        t.start()
        return jsonify({'success': True, 'message': 'Sincronizare pornită în fundal'})

    @app.route('/api/mail/threads')
    @login_required
    def api_mail_threads():
        """List mail threads for current user"""
        folder = request.args.get('folder', 'inbox')  # inbox, atribuite, toate
        page = request.args.get('page', 1, type=int)
        per_page = 30
        
        q = MailThread.query
        
        if folder == 'atribuite':
            q = q.filter_by(atribuit_id=current_user.id)
        elif folder == 'toate':
            pass  # All threads (for admin/office)
        else:
            # Inbox: threads from user's accounts OR assigned to user
            user_conturi = [c.id for c in ContMail.query.filter_by(utilizator_id=current_user.id, activ=True).all()]
            q = q.filter(
                db.or_(
                    MailThread.cont_mail_id.in_(user_conturi) if user_conturi else False,
                    MailThread.atribuit_id == current_user.id
                )
            )
        
        # Filter by status
        status = request.args.get('status')
        if status:
            q = q.filter_by(status=status)
        
        # Search
        search = request.args.get('q', '').strip()
        if search:
            like = f'%{search}%'
            q = q.filter(
                db.or_(
                    MailThread.subiect.ilike(like),
                    MailThread.ultimul_mesaj_de_la.ilike(like)
                )
            )
        
        q = q.filter(MailThread.status != 'arhivat')
        q = q.order_by(MailThread.ultimul_mesaj_data.desc().nullslast())
        
        total = q.count()
        threads = q.offset((page - 1) * per_page).limit(per_page).all()
        
        return jsonify({
            'threads': [{
                'id': t.id,
                'subiect': t.subiect or '(fără subiect)',
                'de_la': t.ultimul_mesaj_de_la or '',
                'data': t.ultimul_mesaj_data.strftime('%d.%m.%Y %H:%M') if t.ultimul_mesaj_data else '',
                'data_scurta': _mail_date_short(t.ultimul_mesaj_data),
                'nr_mesaje': t.nr_mesaje or 0,
                'status': t.status,
                'citit': t.citit,
                'are_atasamente': t.are_atasamente,
                'client': t.client.nume if t.client else None,
                'atribuit': t.atribuit.nume_complet if t.atribuit else None,
                'atribuit_id': t.atribuit_id,
                'snippet': (t.mesaje[-1].snippet if t.mesaje else '')[:120],
                'prioritate': t.prioritate,
                'etichete': [e.strip() for e in (t.etichete or '').split(',') if e.strip()],
            } for t in threads],
            'total': total,
            'pages': (total + per_page - 1) // per_page
        })

    def _mail_date_short(dt):
        if not dt:
            return ''
        now = datetime.now(timezone.utc)
        if dt.date() == now.date():
            return dt.strftime('%H:%M')
        elif dt.year == now.year:
            return dt.strftime('%d %b')
        return dt.strftime('%d.%m.%Y')

    @app.route('/api/mail/thread/<int:tid>')
    @login_required
    def api_mail_thread_detail(tid):
        """Get full thread with all messages"""
        thread = MailThread.query.get_or_404(tid)
        # Mark as read
        if not thread.citit:
            thread.citit = True
            db.session.commit()
        
        return jsonify({
            'thread': {
                'id': thread.id,
                'subiect': thread.subiect,
                'status': thread.status,
                'prioritate': thread.prioritate,
                'client_id': thread.client_id,
                'client': thread.client.nume if thread.client else None,
                'atribuit_id': thread.atribuit_id,
                'atribuit': thread.atribuit.nume_complet if thread.atribuit else None,
                'oferta_id': thread.oferta_id,
                'comanda_id': thread.comanda_id,
                'activitate_id': thread.activitate_id,
                'etichete': [e.strip() for e in (thread.etichete or '').split(',') if e.strip()],
            },
            'mesaje': [{
                'id': m.id,
                'gmail_msg_id': m.gmail_msg_id,
                'de_la': m.de_la,
                'de_la_email': m.de_la_email,
                'catre': m.catre,
                'cc': m.cc,
                'subiect': m.subiect,
                'body_html': m.body_html,
                'body_text': m.body_text,
                'data': m.data_trimitere.strftime('%d.%m.%Y %H:%M') if m.data_trimitere else '',
                'directie': m.directie,
                'atasamente': m.atasamente,
                'snippet': m.snippet
            } for m in thread.mesaje]
        })

    @app.route('/api/mail/thread/<int:tid>/atribuie', methods=['POST'])
    @login_required
    def api_mail_atribuie(tid):
        """Assign thread to a user + auto-forward via Gmail"""
        thread = MailThread.query.get_or_404(tid)
        d = request.get_json()
        user_id = d.get('utilizator_id')
        old_atribuit = thread.atribuit_id
        
        if user_id:
            user_id = int(user_id)
            thread.atribuit_id = user_id
            thread.status = 'atribuit'
        else:
            thread.atribuit_id = None
            thread.status = 'nou'
        db.session.commit()
        
        user = Utilizator.query.get(user_id) if user_id else None
        forwarded = False
        
        # Auto-forward: send the original mail to the assigned agent's email
        if user and user.email and user_id != old_atribuit:
            office_cont = ContMail.query.filter_by(tip='office', activ=True).first()
            if office_cont and thread.mesaje:
                try:
                    from mail_service import send_mail
                    # Build forward body from original messages
                    orig = thread.mesaje[0]  # First message in thread
                    fwd_body = f'''<div style="font-family:Arial,sans-serif;font-size:13px">
                        <p><b>Atribuit de {current_user.nume_complet}</b> — <a href="{request.host_url}admin/mail">Deschide în ERP</a></p>
                        <hr style="border:none;border-top:1px solid #ddd">
                        <p><b>De la:</b> {orig.de_la} &lt;{orig.de_la_email}&gt;<br>
                        <b>Către:</b> {orig.catre or ""}<br>
                        <b>Data:</b> {orig.data_trimitere.strftime("%d.%m.%Y %H:%M") if orig.data_trimitere else ""}<br>
                        <b>Subiect:</b> {orig.subiect or ""}</p>
                        <hr style="border:none;border-top:1px solid #ddd">
                        {orig.body_html or (orig.body_text or "").replace(chr(10), "<br>")}
                    </div>'''
                    
                    gmail_id = send_mail(
                        office_cont,
                        to=user.email,
                        subject=f'[Atribuit] {thread.subiect or "(fără subiect)"}',
                        body_html=fwd_body,
                        reply_to_msg_id=orig.gmail_msg_id
                    )
                    forwarded = bool(gmail_id)
                except Exception as e:
                    print(f'Auto-forward error: {e}')
        
        return jsonify({'success': True, 'atribuit': user.nume_complet if user else None, 'forwarded': forwarded})

    @app.route('/api/mail/thread/<int:tid>/client', methods=['POST'])
    @login_required
    def api_mail_link_client(tid):
        """Link thread to a client"""
        thread = MailThread.query.get_or_404(tid)
        d = request.get_json()
        thread.client_id = d.get('client_id')
        db.session.commit()
        return jsonify({'success': True})

    @app.route('/api/mail/thread/<int:tid>/status', methods=['POST'])
    @login_required
    def api_mail_thread_status(tid):
        """Update thread status"""
        thread = MailThread.query.get_or_404(tid)
        d = request.get_json()
        thread.status = d.get('status', thread.status)
        if d.get('prioritate'):
            thread.prioritate = d['prioritate']
        db.session.commit()
        return jsonify({'success': True})

    @app.route('/api/mail/thread/<int:tid>/creaza', methods=['POST'])
    @login_required
    def api_mail_creaza_document(tid):
        """Create oferta/comanda/activitate from mail thread"""
        thread = MailThread.query.get_or_404(tid)
        d = request.get_json()
        tip = d.get('tip')  # oferta, comanda, activitate
        
        if tip == 'oferta':
            oferta = Oferta(
                numar=f'OFR-{datetime.now().strftime("%Y%m%d-%H%M%S")}',
                client_id=thread.client_id,
                status='draft',
                data_creare=datetime.now(timezone.utc),
                observatii=f'Creat din mail: {thread.subiect}',
                moneda='EUR', subtotal=0, tva_procent=19, tva_valoare=0, total=0,
                creat_de_id=current_user.id
            )
            db.session.add(oferta)
            db.session.flush()
            thread.oferta_id = oferta.id
            db.session.commit()
            return jsonify({'success': True, 'url': f'/admin/oferte/{oferta.id}', 'numar': oferta.numar})
        
        elif tip == 'activitate':
            act = Activitate(
                titlu=thread.subiect or 'Activitate din mail',
                descriere=f'Creat din thread mail #{thread.id}\n{thread.mesaje[0].snippet if thread.mesaje else ""}',
                status='de_facut',
                asignat_id=thread.atribuit_id or current_user.id,
                data_start=date.today()
            )
            db.session.add(act)
            db.session.flush()
            thread.activitate_id = act.id
            db.session.commit()
            return jsonify({'success': True, 'url': f'/admin/activitati/{act.id}', 'titlu': act.titlu})
        
        return jsonify({'error': 'Tip necunoscut'}), 400

    @app.route('/api/mail/send', methods=['POST'])
    @login_required
    def api_mail_send():
        """Send email from user's connected account"""
        from mail_service import send_mail
        d = request.get_json()
        
        # Find user's personal mail account
        cont = ContMail.query.filter_by(
            utilizator_id=current_user.id, activ=True
        ).filter(ContMail.tip.in_(['personal', 'office'])).first()
        if not cont:
            return jsonify({'error': 'Nu ai cont de mail conectat'}), 400
        
        to = d.get('to', '')
        subject = d.get('subject', '')
        body = d.get('body', '')
        cc = d.get('cc', '')
        reply_to = d.get('reply_to_gmail_id')
        thread_id = d.get('thread_id')
        
        if not to or not body:
            return jsonify({'error': 'Completează destinatar și mesaj'}), 400
        
        gmail_id = send_mail(cont, to, subject, body, cc=cc, reply_to_msg_id=reply_to)
        
        if gmail_id:
            # Sync the sent message back
            from mail_service import sync_inbox
            sync_inbox(cont, max_results=5)
            return jsonify({'success': True, 'gmail_id': gmail_id})
        return jsonify({'error': 'Eroare la trimitere'}), 500

    @app.route('/api/mail/attachment/<int:msg_id>/<int:att_idx>')
    @login_required
    def api_mail_download_att(msg_id, att_idx):
        """Download an email attachment"""
        from mail_service import download_attachment
        msg = MailMesaj.query.get_or_404(msg_id)
        atts = msg.atasamente
        if att_idx >= len(atts):
            return 'Atașament negăsit', 404
        att = atts[att_idx]
        if not att.get('gmail_att_id'):
            return 'Atașament inline (nu poate fi descărcat separat)', 400
        cont = msg.thread.cont_mail
        if not cont or not cont.activ:
            return 'Cont mail inactiv', 500
        try:
            data = download_attachment(cont, msg.gmail_msg_id, att.get('gmail_att_id'))
        except Exception as e:
            return f'Eroare: {e}', 500
        if data:
            from flask import send_file
            import io
            return send_file(
                io.BytesIO(data), download_name=att['name'],
                mimetype=att.get('mime', 'application/octet-stream')
            )
        return 'Eroare download atașament', 500

    # ════════════════════════════════════════════════════════════
    # ÎNCASĂRI BANCARE
    # ════════════════════════════════════════════════════════════
    from models import Incasare

    @app.route('/admin/incasari')
    @login_required
    def incasari_page():
        return render_template('admin/incasari.html')

    @app.route('/api/incasari')
    @login_required
    def api_incasari_list():
        """List all payments with filters"""
        page = request.args.get('page', 1, type=int)
        status = request.args.get('status', '')
        search = request.args.get('q', '').strip()
        per_page = 30

        q = Incasare.query
        if status:
            q = q.filter_by(status=status)
        if search:
            like = f'%{search}%'
            q = q.filter(db.or_(
                Incasare.platitor_nume.ilike(like),
                Incasare.referinta.ilike(like),
                Incasare.detalii.ilike(like)
            ))
        q = q.order_by(Incasare.data_tranzactie.desc())
        total = q.count()
        items = q.offset((page - 1) * per_page).limit(per_page).all()

        return jsonify({
            'incasari': [{
                'id': i.id,
                'data': i.data_tranzactie.strftime('%d.%m.%Y') if i.data_tranzactie else '',
                'suma': i.suma,
                'moneda': i.moneda,
                'platitor': i.platitor_nume or '',
                'referinta': i.referinta or '',
                'status': i.status,
                'factura': f'{"PF" if i.factura.tip=="proforma" else "FC"} {i.factura.numar_complet}' if i.factura else None,
                'factura_id': i.factura_id,
                'client': i.client.nume if i.client else None,
                'sursa': i.sursa,
            } for i in items],
            'total': total,
            'pages': (total + per_page - 1) // per_page,
            'stats': _incasari_stats()
        })

    def _incasari_stats():
        nereconciliat = Incasare.query.filter_by(status='nereconciliat').count()
        automat = Incasare.query.filter_by(status='automat').count()
        manual = Incasare.query.filter_by(status='manual').count()
        total_nereconciliat = db.session.query(db.func.sum(Incasare.suma)).filter_by(status='nereconciliat').scalar() or 0
        return {
            'nereconciliat': nereconciliat,
            'automat': automat,
            'manual': manual,
            'total_nereconciliat': round(total_nereconciliat, 2)
        }

    @app.route('/api/incasari/import-csv', methods=['POST'])
    @login_required
    def api_incasari_import_csv():
        """Import bank transactions from CSV"""
        from bank_service import parse_bt_csv, reconcile_batch
        file = request.files.get('fisier')
        if not file:
            return jsonify({'error': 'Niciun fișier selectat'}), 400

        content = file.read()
        # Try multiple encodings
        transactions = parse_bt_csv(content, 'utf-8')
        if not transactions:
            transactions = parse_bt_csv(content, 'latin-1')
        if not transactions:
            transactions = parse_bt_csv(content, 'cp1250')

        imported = 0
        skipped = 0
        new_ids = []
        for t in transactions:
            # Skip duplicates
            existing = Incasare.query.filter_by(referinta_banca=t['referinta_banca']).first()
            if existing:
                skipped += 1
                continue
            inc = Incasare(
                data_tranzactie=t['data_tranzactie'],
                suma=t['suma'], moneda=t.get('moneda', 'RON'),
                platitor_nume=t['platitor_nume'],
                platitor_iban=t.get('platitor_iban', ''),
                platitor_cui=t.get('platitor_cui', ''),
                referinta=t['referinta'],
                detalii=t.get('detalii', ''),
                referinta_banca=t['referinta_banca'],
                sursa='csv'
            )
            db.session.add(inc)
            db.session.flush()
            new_ids.append(inc.id)
            imported += 1

        db.session.commit()

        # Auto-match
        stats = reconcile_batch(new_ids) if new_ids else {'matched': 0}

        return jsonify({
            'success': True,
            'imported': imported,
            'skipped': skipped,
            'matched': stats.get('matched', 0),
            'match_types': stats.get('types', {})
        })

    @app.route('/api/incasari/mock', methods=['POST'])
    @login_required
    def api_incasari_mock():
        """Generate mock transactions for testing"""
        from bank_service import generate_mock_transactions, reconcile_batch
        transactions = generate_mock_transactions(15)
        new_ids = []
        for t in transactions:
            existing = Incasare.query.filter_by(referinta_banca=t['referinta_banca']).first()
            if existing:
                continue
            inc = Incasare(
                data_tranzactie=t['data_tranzactie'],
                suma=t['suma'], moneda=t.get('moneda', 'RON'),
                platitor_nume=t['platitor_nume'],
                platitor_iban=t.get('platitor_iban', ''),
                platitor_cui=t.get('platitor_cui', ''),
                referinta=t['referinta'],
                detalii=t.get('detalii', ''),
                referinta_banca=t['referinta_banca'],
                sursa='mock'
            )
            db.session.add(inc)
            db.session.flush()
            new_ids.append(inc.id)
        db.session.commit()
        stats = reconcile_batch(new_ids) if new_ids else {'matched': 0}
        return jsonify({'success': True, 'count': len(new_ids), 'matched': stats.get('matched', 0)})

    @app.route('/api/incasari/<int:iid>/reconciliaza', methods=['POST'])
    @login_required
    def api_incasari_reconciliaza(iid):
        """Manually reconcile a payment to an invoice"""
        inc = Incasare.query.get_or_404(iid)
        d = request.get_json()
        factura_id = d.get('factura_id')

        if factura_id:
            factura = Factura.query.get_or_404(factura_id)
            inc.factura_id = factura.id
            inc.client_id = factura.client_id
            inc.status = 'manual'
            inc.reconciliat_de_id = current_user.id
            inc.data_reconciliere = datetime.now(timezone.utc)
            # Update invoice
            total_incasat = sum(i.suma for i in factura.incasari if i.status in ('automat', 'manual'))
            if total_incasat >= factura.total - 0.01:
                factura.status = 'incasata'
                # Cross-update related invoices
                from bank_service import _sync_related_invoices
                _sync_related_invoices(factura)
            else:
                factura.status = 'partial'
        else:
            # Unreconcile
            if inc.factura:
                inc.factura.status = 'emisa'
            inc.factura_id = None
            inc.client_id = None
            inc.status = 'nereconciliat'
            inc.reconciliat_de_id = None
            inc.data_reconciliere = None

        db.session.commit()
        return jsonify({'success': True})

    @app.route('/api/incasari/<int:iid>/ignora', methods=['POST'])
    @login_required
    def api_incasari_ignora(iid):
        """Mark payment as ignored"""
        inc = Incasare.query.get_or_404(iid)
        inc.status = 'ignorat'
        db.session.commit()
        return jsonify({'success': True})

    @app.route('/api/incasari/suggest/<int:iid>')
    @login_required
    def api_incasari_suggest(iid):
        """Suggest matching invoices for a payment"""
        inc = Incasare.query.get_or_404(iid)
        tolerance = 0.5
        unpaid = Factura.query.filter(Factura.status.in_(['emisa', 'trimisa', 'partial'])).all()

        suggestions = []
        platitor = (inc.platitor_nume or '').lower()
        for f in unpaid:
            score = 0
            reasons = []
            # Amount match
            diff = abs(f.total - inc.suma)
            if diff < 0.01:
                score += 50
                reasons.append('sumă exactă')
            elif diff < tolerance:
                score += 30
                reasons.append('sumă apropiată')
            elif diff / max(f.total, 1) < 0.05:
                score += 10
                reasons.append('sumă ~5%')
            # Name match
            if f.client and f.client.nume.lower() in platitor:
                score += 30
                reasons.append('nume client')
            elif f.client and any(w in platitor for w in f.client.nume.lower().split() if len(w) > 3):
                score += 15
                reasons.append('nume parțial')
            # CUI match
            if inc.platitor_cui and f.client and f.client.cui:
                if inc.platitor_cui.replace('RO', '') == f.client.cui.replace('RO', ''):
                    score += 40
                    reasons.append('CUI')

            if score > 0:
                suggestions.append({
                    'factura_id': f.id,
                    'numar': f.numar_complet,
                    'client': f.client.nume if f.client else '',
                    'total': f.total,
                    'moneda': f.moneda,
                    'data': f.data_factura.strftime('%d.%m.%Y') if f.data_factura else '',
                    'score': score,
                    'reasons': reasons
                })

        suggestions.sort(key=lambda x: x['score'], reverse=True)
        return jsonify({'suggestions': suggestions[:10]})

    @app.route('/api/incasari/re-match', methods=['POST'])
    @login_required
    def api_incasari_rematch():
        """Re-parse payer info and re-run auto-matching"""
        from bank_service import reconcile_batch, _extract_payer_info
        # Re-parse all payer info from detalii
        reparsed = 0
        for inc in Incasare.query.all():
            if inc.detalii:
                name, iban, cui = _extract_payer_info(inc.detalii)
                if name and name != inc.platitor_nume:
                    inc.platitor_nume = name
                    reparsed += 1
                if iban and not inc.platitor_iban:
                    inc.platitor_iban = iban
                if cui and not inc.platitor_cui:
                    inc.platitor_cui = cui
        db.session.commit()
        # Then re-match
        stats = reconcile_batch()
        stats['reparsed'] = reparsed
        return jsonify({'success': True, **stats})

    # ════════════════════════════════════════════════════════════
    # CURS VALUTAR
    # ════════════════════════════════════════════════════════════
    @app.route('/api/curs-valutar')
    @login_required
    def api_curs_valutar():
        """Get today's exchange rate"""
        from curs_service import get_curs_today
        curs, curs_bnr = get_curs_today('EUR')
        from models import CursValutar
        cached = CursValutar.query.filter_by(data=date.today(), moneda='EUR').first()
        return jsonify({
            'curs': curs, 'curs_bnr': curs_bnr,
            'sursa': cached.sursa if cached else None,
            'multiplicator': cached.multiplicator if cached else 1.01,
            'data': str(date.today()),
        })

    @app.route('/api/curs-valutar/manual', methods=['POST'])
    @login_required
    def api_curs_manual():
        """Set manual exchange rate"""
        from curs_service import set_manual_rate
        d = request.get_json()
        curs = float(d.get('curs', 0))
        data_str = d.get('data', str(date.today()))
        data_curs = date.fromisoformat(data_str)
        if curs <= 0:
            return jsonify({'error': 'Cursul trebuie să fie pozitiv'}), 400
        set_manual_rate(data_curs, 'EUR', curs)
        return jsonify({'success': True, 'curs': curs, 'data': str(data_curs)})

    # ════════════════════════════════════════════════════════════
    # NOTIFICĂRI IN-APP
    # ════════════════════════════════════════════════════════════
    @app.route('/api/notifications')
    @login_required
    def api_notifications():
        notifs = []
        today_date = date.today()

        # Follow-ups restante
        if current_user.has_access('oferte'):
            followups = FollowUpOferta.query.filter(
                FollowUpOferta.next_date <= today_date,
                FollowUpOferta.next_date != None
            ).join(Oferta).filter(Oferta.status.in_(['draft','trimisa'])).all()
            if current_user.doar_proprii:
                followups = [f for f in followups if f.oferta.creat_de_id == current_user.id]
            for f in followups[:5]:
                notifs.append({'type': 'warning', 'icon': 'bi-telephone', 'text': f'Follow-up restant: {f.oferta.numar}',
                              'url': f'/admin/oferte/{f.oferta_id}', 'time': f.next_date.strftime('%d.%m')})

        # Alerte stoc minim
        if current_user.has_access('wms'):
            praguri = StocMinim.query.filter_by(activ=True).all()
            alerte = [p for p in praguri if p.sub_prag]
            for a in alerte[:5]:
                notifs.append({'type': 'danger', 'icon': 'bi-exclamation-triangle', 'text': f'Stoc sub minim: {a.cod_intern} ({a.stoc_actual:.0f}/{a.prag_minim:.0f})',
                              'url': '/admin/wms/alerte-stoc'})

        # Picking-uri de pregătit
        if current_user.has_access('wms'):
            pickings_noi = Picking.query.filter(Picking.status.in_(['nou', 'in_pregatire'])).count()
            if pickings_noi:
                notifs.append({'type': 'info', 'icon': 'bi-cart-check', 'text': f'{pickings_noi} picking-uri de pregătit',
                              'url': '/admin/wms/pickings'})

        # Activități restante
        if current_user.has_access('activitati'):
            activitati_q = Activitate.query.filter(
                Activitate.status.in_(['noua', 'in_lucru']),
                Activitate.deadline <= today_date
            )
            if current_user.doar_proprii:
                activitati_q = activitati_q.filter_by(asignat_id=current_user.id)
            restante = activitati_q.count()
            if restante:
                notifs.append({'type': 'warning', 'icon': 'bi-list-check', 'text': f'{restante} activități restante',
                              'url': '/admin/activitati'})

        # Mesaje necitite
        total_chat = 0
        for conv in current_user.conversatii_chat:
            if conv.activ:
                total_chat += conv.necitite_pentru(current_user.id)
        if total_chat:
            notifs.append({'type': 'primary', 'icon': 'bi-chat-dots', 'text': f'{total_chat} mesaje necitite',
                          'url': '/admin/chat'})

        return jsonify({'notifications': notifs, 'count': len(notifs)})

    # ════════════════════════════════════════════════════════════
    # CRM - CLIENTI
    # ════════════════════════════════════════════════════════════
    @app.route('/admin/clienti')
    @module_required('crm')
    def clienti_list():
        q = request.args.get('q', '')
        query = Client.query
        if q:
            query = query.filter(
                db.or_(Client.nume.ilike(f'%{q}%'), Client.cui.ilike(f'%{q}%'),
                       Client.email.ilike(f'%{q}%'), Client.telefon.ilike(f'%{q}%'))
            )
        clienti = query.order_by(Client.nume).all()
        return render_template('admin/clienti.html', clienti=clienti, q=q)

    @app.route('/admin/clienti/nou', methods=['GET','POST'])
    @module_required('crm')
    def client_nou():
        if request.method == 'POST':
            c = Client(
                cui=request.form.get('cui','').strip() or None,
                nume=request.form.get('nume','').strip(),
                tip=request.form.get('tip','companie'),
                nr_reg_com=request.form.get('nr_reg_com',''),
                email=request.form.get('email',''),
                telefon=request.form.get('telefon',''),
                telefon_secundar=request.form.get('telefon_secundar',''),
                persoana_contact=request.form.get('persoana_contact',''),
                adresa=request.form.get('adresa',''),
                oras=request.form.get('oras',''),
                judet=request.form.get('judet',''),
                cod_postal=request.form.get('cod_postal',''),
                tara=request.form.get('tara','România'),
                banca=request.form.get('banca',''),
                iban=request.form.get('iban',''),
                observatii=request.form.get('observatii',''),
                sablon_pret_id=int(request.form.get('sablon_pret_id')) if request.form.get('sablon_pret_id') else None,
            )
            db.session.add(c)
            db.session.commit()
            flash(f'Client "{c.nume}" adăugat!', 'success')
            return redirect(url_for('clienti_list'))
        sabloane = SablonListaPreturi.query.filter_by(activ=True).order_by(SablonListaPreturi.nume).all()
        return render_template('admin/client_form.html', client=None, sabloane_preturi=sabloane)

    @app.route('/admin/clienti/<int:cid>', methods=['GET','POST'])
    @login_required
    def client_edit(cid):
        c = Client.query.get_or_404(cid)
        if request.method == 'POST':
            for f in ['cui','nume','tip','nr_reg_com','email','telefon','telefon_secundar',
                       'persoana_contact','adresa','oras','judet','cod_postal','tara','banca','iban','observatii']:
                setattr(c, f, request.form.get(f, '').strip() or getattr(c, f))
            c.sablon_pret_id = int(request.form.get('sablon_pret_id')) if request.form.get('sablon_pret_id') else None
            c.activ = 'activ' in request.form
            db.session.commit()
            flash('Client actualizat!', 'success')
            return redirect(url_for('client_detail', cid=cid))
        sabloane = SablonListaPreturi.query.filter_by(activ=True).order_by(SablonListaPreturi.nume).all()
        return render_template('admin/client_form.html', client=c, sabloane_preturi=sabloane)

    @app.route('/admin/clienti/<int:cid>/detalii')
    @module_required('crm')
    def client_detail(cid):
        c = Client.query.get_or_404(cid)
        return render_template('admin/client_detail.html', client=c)

    # ════════════════════════════════════════════════════════════
    # ANAF LOOKUP
    # ════════════════════════════════════════════════════════════
    @app.route('/api/anaf/lookup/<cui>')
    @login_required
    def anaf_lookup(cui):
        import requests as req
        cui_clean = cui.strip().upper().replace('RO', '').strip()
        try:
            cui_int = int(cui_clean)
        except ValueError:
            return jsonify({'error': 'CUI invalid', 'success': False}), 400
        try:
            payload = json.dumps([{'cui': cui_int, 'data': date.today().strftime('%Y-%m-%d')}])
            for url in ['https://webservicesp.anaf.ro/api/PlatitorTvaRest/v9/tva',
                        'https://webservicesp.anaf.ro/PlatitorTvaRest/api/v8/ws/tva']:
                try:
                    r = req.post(url, data=payload, headers={'Content-Type':'application/json'}, timeout=15)
                    if r.status_code == 200: break
                except: continue
            else:
                return jsonify({'error': 'ANAF nu răspunde', 'success': False}), 503
            data = json.loads(r.text)
            if data.get('found') and len(data['found']) > 0:
                f = data['found'][0]
                gen = f.get('date_generale', f)
                sed = f.get('adresa_sediu_social', {})
                return jsonify({'success': True, 'cui': str(cui_int),
                    'denumire': gen.get('denumire',''), 'adresa': gen.get('adresa',''),
                    'nr_reg_com': gen.get('nrRegCom',''), 'telefon': gen.get('telefon',''),
                    'oras': sed.get('sdenumire_Localitate',''), 'judet': sed.get('sdenumire_Judet',''),
                    'cod_postal': gen.get('codPostal','') or sed.get('scod_Postal','')})
            return jsonify({'error': f'CUI {cui_int} negăsit', 'success': False}), 404
        except Exception as e:
            return jsonify({'error': str(e), 'success': False}), 500

    # ════════════════════════════════════════════════════════════
    # VÂNZĂRI - OFERTE (list, detail, status)
    # ════════════════════════════════════════════════════════════
    @app.route('/admin/oferte')
    @module_required('oferte')
    def oferte_list():
        status = request.args.get('status', '')
        client_id = request.args.get('client_id', '')
        period = request.args.get('period', '')
        date_start = request.args.get('date_start', '')
        date_end = request.args.get('date_end', '')
        query = Oferta.query
        if current_user.doar_proprii:
            query = query.filter_by(creat_de_id=current_user.id)
        if status:
            query = query.filter_by(status=status)
        if client_id:
            query = query.filter_by(client_id=int(client_id))
        if date_start and date_end:
            try:
                ds = datetime.strptime(date_start, '%Y-%m-%d').date()
                de = datetime.strptime(date_end, '%Y-%m-%d').date() + timedelta(days=1)
                query = query.filter(Oferta.data_creare >= datetime(ds.year, ds.month, ds.day),
                                      Oferta.data_creare < datetime(de.year, de.month, de.day))
            except: pass
        elif period:
            today_date = date.today()
            if period == 'saptamana':
                query = query.filter(Oferta.data_creare >= datetime(today_date.year, today_date.month, today_date.day) - timedelta(days=7))
            elif period == 'luna':
                query = query.filter(Oferta.data_creare >= datetime(today_date.year, today_date.month, 1))
            elif period == 'an':
                query = query.filter(Oferta.data_creare >= datetime(today_date.year, 1, 1))
        oferte = query.order_by(Oferta.data_creare.desc()).all()
        clienti = Client.query.filter_by(activ=True).order_by(Client.nume).all()
        return render_template('admin/oferte.html', oferte=oferte, status_filter=status,
                             client_filter=client_id, period_filter=period,
                             date_start=date_start, date_end=date_end, clienti=clienti)

    @app.route('/admin/oferte/<int:oid>')
    @module_required('oferte')
    def oferta_detail(oid):
        o = Oferta.query.get_or_404(oid)
        # Agents can only see their own offers
        if current_user.doar_proprii and o.creat_de_id != current_user.id:
            flash('Nu ai acces la această ofertă.', 'error')
            return redirect(url_for('oferte_list'))
        return render_template('admin/oferta_detail.html', oferta=o)

    @app.route('/api/oferta/<int:oid>/status', methods=['POST'])
    @login_required
    def api_oferta_status(oid):
        o = Oferta.query.get_or_404(oid)
        new_status = request.json.get('status')
        if new_status in dict(Oferta.STATUSES):
            old_status = o.status
            o.status = new_status
            AuditLog.log('oferta', o.id, o.numar, 'status_schimbat',
                         f'{dict(Oferta.STATUSES).get(old_status,old_status)} → {dict(Oferta.STATUSES).get(new_status,new_status)}',
                         current_user.id)
            db.session.commit()
            return jsonify({'success': True})
        return jsonify({'error': 'Status invalid'}), 400

    # ════════════════════════════════════════════════════════════
    # KEY FLOW: OFERTĂ → COMANDĂ (conversion)
    # ════════════════════════════════════════════════════════════
    @app.route('/api/oferta/<int:oid>/convert', methods=['POST'])
    @login_required
    def api_oferta_to_comanda(oid):
        """Convert an accepted offer into an order - THE KEY BUSINESS FLOW
        
        Rules:
        - Must have at least one proforma
        - Proforma incasata → comanda status 'noua' (direct)
        - Proforma confirmata → comanda status 'pending' (needs admin/director approval)
        - Proforma only emisa/trimisa → blocked
        """
        oferta = Oferta.query.get_or_404(oid)

        if oferta.comanda:
            return jsonify({'error': 'Oferta are deja o comandă asociată', 'comanda_id': oferta.comanda.id}), 400

        # Check proforma exists
        proforme = Factura.query.filter_by(oferta_id=oferta.id, tip='proforma').all()
        if not proforme:
            return jsonify({'error': 'Nu se poate crea comanda fără proformă. Generați mai întâi o proformă.'}), 400

        # Determine proforma status
        pf_incasata = any(p.status == 'incasata' for p in proforme)
        pf_confirmata = any(p.status == 'confirmata' for p in proforme)

        if not pf_incasata and not pf_confirmata:
            return jsonify({'error': 'Proforma trebuie să fie cel puțin confirmată de client înainte de a crea comanda.'}), 400

        # Determine order status
        if pf_incasata:
            order_status = 'noua'
        else:
            # Confirmata but not paid - pending approval
            order_status = 'pending'

        # Generate order number
        nr = f"CMD-{datetime.now().strftime('%Y%m%d-%H%M%S')}"

        comanda = Comanda(
            numar=nr, client_id=oferta.client_id, oferta_id=oferta.id,
            status=order_status, data_comanda=date.today(),
            subtotal=oferta.subtotal, tva_procent=oferta.tva_procent,
            tva_valoare=oferta.tva_valoare, total=oferta.total,
            moneda=oferta.moneda, observatii=f'Din oferta {oferta.numar}',
            creat_de_id=current_user.id,
        )
        db.session.add(comanda)
        db.session.flush()

        # Copy ALL lines from offer to order
        for lo in oferta.linii:
            lc = LinieComanda(
                comanda_id=comanda.id, ordine=lo.ordine, tip=lo.tip,
                cod=lo.cod, denumire=lo.denumire, dimensiune=lo.dimensiune,
                um=lo.um, cantitate=lo.cantitate, pret_unitar=lo.pret_final,
            )
            lc.parametri = lo.parametri
            lc.accesorii = lo.accesorii
            db.session.add(lc)

        # Update offer status
        oferta.status = 'comanda'

        # Apply any activity templates triggered by oferta->comanda
        sabloane = SablonActivitate.query.filter_by(trigger='oferta_comanda', activ=True).all()
        for s in sabloane:
            s.aplica(comanda_id=comanda.id, client_id=oferta.client_id, creat_de_id=current_user.id)

        db.session.commit()

        status_msg = 'în așteptare aprobare (proformă confirmată, neplătită)' if order_status == 'pending' else 'creată'
        AuditLog.log('oferta', oferta.id, oferta.numar, 'convertit', f'Convertit în comandă {comanda.numar} ({status_msg})', current_user.id)
        AuditLog.log('comanda', comanda.id, comanda.numar, 'creat', f'Creat din oferta {oferta.numar} - {status_msg}', current_user.id)
        db.session.commit()

        return jsonify({
            'success': True, 'comanda_id': comanda.id, 'numar': comanda.numar,
            'pending': order_status == 'pending'
        })

    # ════════════════════════════════════════════════════════════
    # OFERTE - FOLLOW-UPS
    # ════════════════════════════════════════════════════════════
    @app.route('/api/oferta/<int:oid>/followup', methods=['POST'])
    @login_required
    def api_oferta_followup(oid):
        oferta = Oferta.query.get_or_404(oid)
        d = request.get_json()
        fu = FollowUpOferta(
            oferta_id=oid,
            metoda=d.get('metoda', 'telefon'),
            rezultat=d.get('rezultat', 'interesat'),
            note=d.get('note', '').strip(),
            next_date=datetime.strptime(d['next_date'], '%Y-%m-%d').date() if d.get('next_date') else None,
            creat_de_id=current_user.id,
        )
        db.session.add(fu)
        db.session.commit()
        return jsonify({'success': True, 'id': fu.id})

    @app.route('/api/followup/<int:fid>', methods=['DELETE'])
    @login_required
    def api_followup_delete(fid):
        fu = FollowUpOferta.query.get_or_404(fid)
        db.session.delete(fu)
        db.session.commit()
        return jsonify({'success': True})

    # ════════════════════════════════════════════════════════════
    # OFERTE - REVIZII
    # ════════════════════════════════════════════════════════════
    @app.route('/api/oferta/<int:oid>/revizie', methods=['POST'])
    @login_required
    def api_oferta_revizie(oid):
        """Create a new revision by copying the offer"""
        original = Oferta.query.get_or_404(oid)
        # Find the root offer (for chain tracking)
        root = original.parinte if original.parinte else original
        # Next version number
        max_v = max([r.versiune for r in root.revizii] + [root.versiune])
        new_v = max_v + 1

        nr = f"OF-{datetime.now().strftime('%Y%m%d-%H%M%S')}-{str(root.id).zfill(4)}"
        revizie = Oferta(
            numar=nr, versiune=new_v, parinte_id=root.id,
            client_id=original.client_id, status='draft',
            data_oferta=date.today(),
            data_expirare=date.today() + timedelta(days=original.valabilitate_zile),
            valabilitate_zile=original.valabilitate_zile,
            discount_mode=original.discount_mode, discount_global=original.discount_global,
            subtotal=original.subtotal, tva_procent=original.tva_procent,
            tva_valoare=original.tva_valoare, total=original.total,
            moneda=original.moneda,
            observatii=f'Revizie v{new_v} din {original.numar_display}',
            creat_de_id=current_user.id,
        )
        db.session.add(revizie)
        db.session.flush()

        # Copy all lines
        for lo in original.linii:
            nl = LinieOferta(
                oferta_id=revizie.id, ordine=lo.ordine, tip=lo.tip,
                cod=lo.cod, denumire=lo.denumire, dimensiune=lo.dimensiune,
                um=lo.um, cantitate=lo.cantitate, pret_catalog=lo.pret_catalog,
                discount_adaos=lo.discount_adaos, pret_final=lo.pret_final,
            )
            nl.parametri = lo.parametri
            nl.accesorii = lo.accesorii
            db.session.add(nl)

        db.session.commit()
        return jsonify({'success': True, 'revizie_id': revizie.id, 'versiune': new_v})

    @app.route('/api/oferta/<int:oid>/edit-data', methods=['GET'])
    @login_required
    def api_oferta_edit_data(oid):
        """Get offer data for editing in configurator"""
        o = Oferta.query.get_or_404(oid)
        items = []
        for l in o.linii:
            items.append({
                'tip': l.tip, 'cod': l.cod, 'cod_baza': l.cod, 'denumire': l.denumire,
                'dimensiune': l.dimensiune, 'um': l.um, 'cantitate': l.cantitate,
                'pret_catalog': l.pret_catalog, 'discount_adaos': l.discount_adaos,
                'pret_final': l.pret_final, 'parametri': l.parametri, 'accesorii': l.accesorii,
                'is_sub_line': l.is_sub_line or False, 'parent_cod': l.parent_cod or '',
            })
        return jsonify({
            'oferta_id': o.id, 'client_id': o.client_id,
            'valabilitate_zile': o.valabilitate_zile,
            'discount_mode': o.discount_mode, 'discount_global': o.discount_global,
            'items': items,
        })

    # ════════════════════════════════════════════════════════════
    # VÂNZĂRI - COMENZI
    # ════════════════════════════════════════════════════════════
    @app.route('/admin/comenzi')
    @module_required('comenzi')
    def comenzi_list():
        status = request.args.get('status', '')
        client_id = request.args.get('client_id', '')
        period = request.args.get('period', '')
        date_start = request.args.get('date_start', '')
        date_end = request.args.get('date_end', '')
        query = Comanda.query
        if current_user.doar_proprii:
            query = query.filter_by(creat_de_id=current_user.id)
        if status:
            query = query.filter_by(status=status)
        if client_id:
            query = query.filter_by(client_id=int(client_id))
        if date_start and date_end:
            try:
                ds = datetime.strptime(date_start, '%Y-%m-%d').date()
                de = datetime.strptime(date_end, '%Y-%m-%d').date() + timedelta(days=1)
                query = query.filter(Comanda.data_comanda >= ds, Comanda.data_comanda < de)
            except Exception:
                pass
        elif period:
            today_date = date.today()
            if period == 'saptamana':
                query = query.filter(Comanda.data_comanda >= today_date - timedelta(days=7))
            elif period == 'luna':
                query = query.filter(Comanda.data_comanda >= today_date.replace(day=1))
            elif period == 'an':
                query = query.filter(Comanda.data_comanda >= date(today_date.year, 1, 1))
        comenzi = query.order_by(Comanda.data_creare.desc()).all()
        clienti = Client.query.filter_by(activ=True).order_by(Client.nume).all()
        return render_template('admin/comenzi.html', comenzi=comenzi, status_filter=status,
                             client_filter=client_id, period_filter=period,
                             date_start=date_start, date_end=date_end, clienti=clienti)

    @app.route('/admin/comenzi/<int:cid>')
    @module_required('comenzi')
    def comanda_detail(cid):
        c = Comanda.query.get_or_404(cid)
        # Calculate margin per line
        from sqlalchemy import func as sqfunc
        pa_map = {}
        for row in db.session.query(StocProdus.cod_intern, sqfunc.avg(StocProdus.pret_achizitie_mediu)).group_by(StocProdus.cod_intern).all():
            pa_map[row[0]] = row[1] or 0
        for row in db.session.query(LinieNIR.cod_intern, sqfunc.avg(LinieNIR.pret_achizitie)).group_by(LinieNIR.cod_intern).all():
            if row[0] not in pa_map: pa_map[row[0]] = row[1] or 0
        marja_linii = []
        total_vanzare, total_cost = 0, 0
        for l in c.linii:
            vanz = l.valoare_linie
            cost_u = pa_map.get(l.cod, 0)
            cost = l.cantitate * cost_u
            marja = vanz - cost
            marja_pct = (marja / vanz * 100) if vanz else 0
            marja_linii.append({'cost_u': cost_u, 'cost': cost, 'marja': marja, 'marja_pct': marja_pct})
            total_vanzare += vanz; total_cost += cost
        marja_totala = total_vanzare - total_cost
        marja_totala_pct = (marja_totala / total_vanzare * 100) if total_vanzare else 0
        return render_template('admin/comanda_detail.html', comanda=c,
                             marja_linii=marja_linii, marja_totala=marja_totala,
                             marja_totala_pct=marja_totala_pct, total_cost=total_cost)

    @app.route('/api/comanda/<int:cid>/status', methods=['POST'])
    @login_required
    def api_comanda_status(cid):
        c = Comanda.query.get_or_404(cid)
        new_status = request.json.get('status')
        if new_status in dict(Comanda.STATUSES):
            old_status = c.status

            # Pending → noua requires admin
            if old_status == 'pending' and new_status in ('noua', 'anulat'):
                if not current_user.is_admin:
                    return jsonify({'error': 'Doar un administrator poate aproba sau respinge comenzile în așteptare.'}), 403

            # Block any non-admin action on pending orders
            if old_status == 'pending' and new_status not in ('noua', 'anulat'):
                return jsonify({'error': 'Comanda trebuie aprobată mai întâi.'}), 400

            c.status = new_status

            # WMS: When order is confirmed, create stock reservation
            if new_status == 'confirmata' and old_status == 'noua':
                for linie in c.linii:
                    ms = MiscareStoc(comanda_id=c.id, tip='rezervare',
                        cod_produs=linie.cod, denumire_produs=linie.denumire,
                        cantitate=linie.cantitate, numar_document=c.numar,
                        utilizator_id=current_user.id)
                    db.session.add(ms)
                # Trigger activity templates
                for s in SablonActivitate.query.filter_by(trigger='comanda_confirmata', activ=True):
                    s.aplica(comanda_id=c.id, client_id=c.client_id, creat_de_id=current_user.id)

            # Trigger on production
            if new_status == 'productie':
                for s in SablonActivitate.query.filter_by(trigger='comanda_productie', activ=True):
                    s.aplica(comanda_id=c.id, client_id=c.client_id, creat_de_id=current_user.id)

            # WMS: When delivered, convert reservation to exit
            if new_status == 'livrata':
                for linie in c.linii:
                    ms = MiscareStoc(comanda_id=c.id, tip='iesire',
                        cod_produs=linie.cod, denumire_produs=linie.denumire,
                        cantitate=linie.cantitate, numar_document=c.numar,
                        utilizator_id=current_user.id)
                    db.session.add(ms)
                # Trigger activity templates
                for s in SablonActivitate.query.filter_by(trigger='comanda_livrata', activ=True):
                    s.aplica(comanda_id=c.id, client_id=c.client_id, creat_de_id=current_user.id)

            AuditLog.log('comanda', c.id, c.numar, 'status_schimbat',
                         f'{dict(Comanda.STATUSES).get(old_status,old_status)} → {dict(Comanda.STATUSES).get(new_status,new_status)}',
                         current_user.id)
            db.session.commit()
            return jsonify({'success': True})
        return jsonify({'error': 'Status invalid'}), 400

    # ════════════════════════════════════════════════════════════
    # KEY FLOW: OFERTĂ → PROFORMĂ
    # ════════════════════════════════════════════════════════════
    @app.route('/api/oferta/<int:oid>/proforma', methods=['POST'])
    @login_required
    def api_oferta_to_proforma(oid):
        """Generate proforma invoice from offer (EUR→RON conversion)"""
        try:
            oferta = Oferta.query.get_or_404(oid)
            from curs_service import get_curs_today
            
            # Get exchange rate
            curs, curs_bnr = get_curs_today('EUR')
            if curs is None:
                return jsonify({'error': 'Cursul valutar nu este disponibil. Verificați conexiunea la BNR.'}), 400
            
            # EUR amounts from oferta
            subtotal_eur = oferta.subtotal or 0
            tva_val_eur = oferta.tva_valoare or 0
            total_eur = oferta.total or 0
            
            # Convert to RON
            subtotal_ron = round(subtotal_eur * curs, 2)
            tva_val_ron = round(tva_val_eur * curs, 2)
            total_ron = round(total_eur * curs, 2)
            
            # Next proforma number (separate sequence)
            last = Factura.query.filter_by(tip='proforma').order_by(Factura.numar.desc()).first()
            next_nr = (last.numar + 1) if last else 1
            
            serie_pf = Setari.get('serie_proforma', 'PF') or 'PF'
            scadenta_str = Setari.get('scadenta_factura_zile', '30') or '30'
            scadenta_zile = int(scadenta_str)
            
            proforma = Factura(
                tip='proforma',
                serie=serie_pf, numar=next_nr,
                client_id=oferta.client_id, oferta_id=oferta.id,
                data_factura=date.today(),
                data_scadenta=date.today() + timedelta(days=scadenta_zile),
                subtotal=subtotal_ron, tva_procent=oferta.tva_procent or 19,
                tva_valoare=tva_val_ron, total=total_ron,
                moneda='RON',
                subtotal_eur=subtotal_eur, tva_valoare_eur=tva_val_eur,
                total_eur=total_eur, curs_valutar=curs,
            )
            db.session.add(proforma)
            db.session.flush()
            
            for lo in oferta.linii:
                pret_eur = lo.pret_final if hasattr(lo, 'pret_final') else lo.pret_unitar
                val_eur = lo.valoare_linie if hasattr(lo, 'valoare_linie') else (lo.cantitate * lo.pret_unitar)
                lf = LinieFactura(
                    factura_id=proforma.id,
                    denumire=f'{lo.cod_complet} - {lo.denumire}' if hasattr(lo, 'cod_complet') else lo.denumire,
                    um='buc', cantitate=lo.cantitate,
                    pret_unitar=round(pret_eur * curs, 2),
                    valoare=round(val_eur * curs, 2),
                )
                db.session.add(lf)
            
            db.session.commit()
            return jsonify({
                'success': True, 'factura_id': proforma.id,
                'numar': proforma.numar_complet,
                'curs': curs, 'total_eur': total_eur, 'total_ron': total_ron
            })
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': str(e)}), 500

    # ════════════════════════════════════════════════════════════
    # KEY FLOW: COMANDĂ → FACTURĂ FISCALĂ
    # ════════════════════════════════════════════════════════════
    @app.route('/api/comanda/<int:cid>/factura', methods=['POST'])
    @login_required
    def api_comanda_to_factura(cid):
        """Generate fiscal invoice from order (EUR→RON conversion)"""
        try:
            comanda = Comanda.query.get_or_404(cid)
            from curs_service import get_curs_today
            
            # Get exchange rate
            curs, curs_bnr = get_curs_today('EUR')
            if curs is None:
                return jsonify({'error': 'Cursul valutar nu este disponibil. Verificați conexiunea la BNR.'}), 400
            
            # EUR amounts from comanda
            subtotal_eur = comanda.subtotal or 0
            tva_val_eur = comanda.tva_valoare or 0
            total_eur = comanda.total or 0
            
            # Convert to RON
            subtotal_ron = round(subtotal_eur * curs, 2)
            tva_val_ron = round(tva_val_eur * curs, 2)
            total_ron = round(total_eur * curs, 2)
            
            # Next fiscal number (separate sequence)
            last = Factura.query.filter_by(tip='fiscala').order_by(Factura.numar.desc()).first()
            next_nr = (last.numar + 1) if last else 1
            
            serie = Setari.get('serie_factura', 'HSL') or 'HSL'
            scadenta_str = Setari.get('scadenta_factura_zile', '30') or '30'
            scadenta_zile = int(scadenta_str)
            
            factura = Factura(
                tip='fiscala',
                serie=serie, numar=next_nr,
                client_id=comanda.client_id, comanda_id=comanda.id,
                data_factura=date.today(),
                data_scadenta=date.today() + timedelta(days=scadenta_zile),
                subtotal=subtotal_ron, tva_procent=comanda.tva_procent or 19,
                tva_valoare=tva_val_ron, total=total_ron,
                moneda='RON',
                subtotal_eur=subtotal_eur, tva_valoare_eur=tva_val_eur,
                total_eur=total_eur, curs_valutar=curs,
            )
            db.session.add(factura)
            db.session.flush()
            
            for lc in comanda.linii:
                pret_eur = lc.pret_unitar or 0
                val_eur = lc.valoare_linie or 0
                lf = LinieFactura(
                    factura_id=factura.id, denumire=f'{lc.cod} - {lc.denumire}',
                    um=lc.um, cantitate=lc.cantitate,
                    pret_unitar=round(pret_eur * curs, 2),
                    valoare=round(val_eur * curs, 2),
                )
                db.session.add(lf)
            
            # Auto-mark as incasata if proforma was already paid (Flux 1)
            if comanda.oferta_sursa:
                proforme = Factura.query.filter_by(
                    oferta_id=comanda.oferta_sursa.id, tip='proforma'
                ).all()
                if any(p.status == 'incasata' for p in proforme):
                    factura.status = 'incasata'
            
            db.session.commit()
            return jsonify({
                'success': True, 'factura_id': factura.id, 'numar': factura.numar_complet,
                'auto_incasata': factura.status == 'incasata',
                'curs': curs, 'total_eur': total_eur, 'total_ron': total_ron
            })
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': str(e)}), 500

    # ════════════════════════════════════════════════════════════
    # CHECK PROFORMA STATUS (for comanda workflow)
    # ════════════════════════════════════════════════════════════
    @app.route('/api/comanda/<int:cid>/check-plata')
    @login_required
    def api_comanda_check_plata(cid):
        """Check if proforma for this order's offer is confirmed or paid"""
        comanda = Comanda.query.get_or_404(cid)
        if comanda.oferta_sursa:
            proforme = Factura.query.filter_by(
                oferta_id=comanda.oferta_sursa.id, tip='proforma'
            ).all()
            platita = any(p.status == 'incasata' for p in proforme)
            confirmata = any(p.status in ('confirmata', 'incasata') for p in proforme)
            return jsonify({
                'platita': platita,
                'confirmata': confirmata,
                'proforme': [{
                    'id': p.id, 'numar': p.numar_complet,
                    'status': p.status, 'total': p.total
                } for p in proforme]
            })
        return jsonify({'platita': False, 'confirmata': False, 'proforme': []})

    # ════════════════════════════════════════════════════════════
    # FACTURI
    # ════════════════════════════════════════════════════════════
    @app.route('/admin/facturi')
    @module_required('facturi')
    def facturi_list():
        proforme = Factura.query.filter_by(tip='proforma').order_by(Factura.data_factura.desc()).all()
        fiscale = Factura.query.filter_by(tip='fiscala').order_by(Factura.data_factura.desc()).all()
        toate = Factura.query.order_by(Factura.data_factura.desc()).all()
        return render_template('admin/facturi.html', proforme=proforme, fiscale=fiscale, toate=toate)

    @app.route('/admin/facturi/<int:fid>')
    @module_required('facturi')
    def factura_detail(fid):
        f = Factura.query.get_or_404(fid)
        return render_template('admin/factura_detail.html', factura=f)

    # ════════════════════════════════════════════════════════════
    # ȘTERGERE DOCUMENTE
    # ════════════════════════════════════════════════════════════
    @app.route('/api/oferta/<int:oid>/sterge', methods=['POST'])
    @login_required
    def api_oferta_sterge(oid):
        if not current_user.has_access('oferte'):
            return jsonify({'error': 'Acces interzis'}), 403
        o = Oferta.query.get_or_404(oid)
        if o.comanda:
            return jsonify({'error': 'Oferta are comandă asociată. Șterge comanda mai întâi.'}), 400
        numar = o.numar
        client_name = o.client.nume if o.client else ''
        db.session.delete(o)
        AuditLog.log('oferta', oid, numar, 'sters', f'Ofertă ștearsă (client: {client_name})', current_user.id)
        db.session.commit()
        return jsonify({'success': True})

    @app.route('/api/comanda/<int:cid>/sterge', methods=['POST'])
    @login_required
    def api_comanda_sterge(cid):
        if not current_user.has_access('comenzi'):
            return jsonify({'error': 'Acces interzis'}), 403
        c = Comanda.query.get_or_404(cid)
        numar = c.numar
        client_name = c.client.nume if c.client else ''
        # Delete related activities
        Activitate.query.filter_by(comanda_id=c.id).delete()
        # Delete related stock movements
        MiscareStoc.query.filter_by(comanda_id=c.id).delete()
        # Delete related invoices and their lines
        for f in Factura.query.filter_by(comanda_id=c.id).all():
            LinieFactura.query.filter_by(factura_id=f.id).delete()
            db.session.delete(f)
        # Reset oferta status if came from one
        if c.oferta_id:
            oferta = Oferta.query.get(c.oferta_id)
            if oferta and oferta.status == 'comanda':
                oferta.status = 'acceptata'
        db.session.delete(c)
        AuditLog.log('comanda', cid, numar, 'sters', f'Comandă ștearsă (client: {client_name})', current_user.id)
        db.session.commit()
        return jsonify({'success': True})

    @app.route('/api/factura/<int:fid>/status', methods=['POST'])
    @login_required
    def api_factura_status(fid):
        f = Factura.query.get_or_404(fid)
        d = request.get_json()
        f.status = d.get('status', f.status)
        db.session.commit()
        return jsonify({'success': True})

    @app.route('/api/factura/<int:fid>/sterge', methods=['POST'])
    @login_required
    def api_factura_sterge(fid):
        if not current_user.has_access('facturi'):
            return jsonify({'error': 'Acces interzis'}), 403
        f = Factura.query.get_or_404(fid)
        numar = f'{f.serie}/{f.numar}'
        client_name = f.client.nume if f.client else ''
        db.session.delete(f)
        AuditLog.log('factura', fid, numar, 'sters', f'Factură ștearsă (client: {client_name})', current_user.id)
        db.session.commit()
        return jsonify({'success': True})

    @app.route('/api/activitate/<int:aid>/sterge', methods=['POST'])
    @login_required
    def api_activitate_sterge(aid):
        if not (current_user.has_access('activitati_manage')):
            return jsonify({'error': 'Acces interzis'}), 403
        a = Activitate.query.get_or_404(aid)
        db.session.delete(a)
        db.session.commit()
        return jsonify({'success': True})

    # ════════════════════════════════════════════════════════════
    # ACTIVITĂȚI (Task Management)
    # ════════════════════════════════════════════════════════════
    @app.route('/admin/activitati')
    @module_required('activitati')
    def activitati_list():
        status = request.args.get('status', '')
        asignat = request.args.get('asignat', '')
        query = Activitate.query
        if current_user.doar_proprii:
            query = query.filter(
                db.or_(Activitate.asignat_id == current_user.id,
                       Activitate.creat_de_id == current_user.id))
        if status:
            query = query.filter_by(status=status)
        if asignat:
            query = query.filter_by(asignat_id=int(asignat))
        activitati = query.order_by(
            db.case((Activitate.prioritate == 'urgenta', 0),(Activitate.prioritate == 'ridicata', 1),
                     (Activitate.prioritate == 'normala', 2), else_=3),
            Activitate.deadline.asc().nullslast(),
            Activitate.data_creare.desc()
        ).all()
        utilizatori = Utilizator.query.filter_by(activ=True).all()
        return render_template('admin/activitati.html', activitati=activitati,
                             utilizatori=utilizatori, status_filter=status, asignat_filter=asignat)

    @app.route('/admin/activitati/nou', methods=['GET','POST'])
    @module_required('activitati_manage')
    def activitate_nou():
        if request.method == 'POST':
            a = Activitate(
                titlu=request.form.get('titlu','').strip(),
                descriere=request.form.get('descriere','').strip(),
                tip_id=int(request.form.get('tip_id') or 0) or None,
                prioritate=request.form.get('prioritate','normala'),
                comanda_id=int(request.form.get('comanda_id') or 0) or None,
                client_id=int(request.form.get('client_id') or 0) or None,
                asignat_id=int(request.form.get('asignat_id') or 0) or None,
                creat_de_id=current_user.id,
                deadline=datetime.strptime(request.form['deadline'],'%Y-%m-%d').date() if request.form.get('deadline') else None,
            )
            db.session.add(a)
            db.session.commit()
            flash(f'Activitate "{a.titlu}" creată!', 'success')
            return redirect(url_for('activitati_list'))
        comenzi = Comanda.query.filter(Comanda.status.notin_(['finalizata','anulat'])).order_by(Comanda.data_creare.desc()).all()
        clienti = Client.query.filter_by(activ=True).order_by(Client.nume).all()
        utilizatori = Utilizator.query.filter_by(activ=True).all()
        tipuri = TipActivitate.query.filter_by(activ=True).order_by(TipActivitate.ordine).all()
        return render_template('admin/activitate_form.html', act=None,
                             comenzi=comenzi, clienti=clienti, utilizatori=utilizatori, tipuri=tipuri)

    @app.route('/admin/activitati/<int:aid>')
    @module_required('activitati')
    def activitate_detail(aid):
        a = Activitate.query.get_or_404(aid)
        utilizatori = Utilizator.query.filter_by(activ=True).all()
        return render_template('admin/activitate_detail.html', act=a, utilizatori=utilizatori)

    @app.route('/admin/activitati/<int:aid>/edit', methods=['GET','POST'])
    @module_required('activitati_manage')
    def activitate_edit(aid):
        a = Activitate.query.get_or_404(aid)
        if request.method == 'POST':
            a.titlu = request.form.get('titlu', a.titlu).strip()
            a.descriere = request.form.get('descriere','').strip()
            a.tip_id = int(request.form.get('tip_id') or 0) or None
            a.prioritate = request.form.get('prioritate', a.prioritate)
            a.comanda_id = int(request.form.get('comanda_id') or 0) or None
            a.client_id = int(request.form.get('client_id') or 0) or None
            a.asignat_id = int(request.form.get('asignat_id') or 0) or None
            a.deadline = datetime.strptime(request.form['deadline'],'%Y-%m-%d').date() if request.form.get('deadline') else None
            db.session.commit()
            flash('Activitate actualizată!', 'success')
            return redirect(url_for('activitate_detail', aid=aid))
        comenzi = Comanda.query.filter(Comanda.status.notin_(['finalizata','anulat'])).order_by(Comanda.data_creare.desc()).all()
        clienti = Client.query.filter_by(activ=True).order_by(Client.nume).all()
        utilizatori = Utilizator.query.filter_by(activ=True).all()
        tipuri = TipActivitate.query.filter_by(activ=True).order_by(TipActivitate.ordine).all()
        return render_template('admin/activitate_form.html', act=a,
                             comenzi=comenzi, clienti=clienti, utilizatori=utilizatori, tipuri=tipuri)

    @app.route('/api/activitate/<int:aid>/status', methods=['POST'])
    @login_required
    def api_activitate_status(aid):
        if not (current_user.has_access('activitati_status') or current_user.has_access('activitati_manage')):
            return jsonify({'error': 'Nu ai permisiunea de a modifica statusul'}), 403
        a = Activitate.query.get_or_404(aid)
        new_status = request.json.get('status')
        if new_status in dict(Activitate.STATUSES):
            a.status = new_status
            if new_status == 'in_lucru' and not a.data_start:
                a.data_start = datetime.now(timezone.utc)
            elif new_status == 'finalizat':
                a.data_finalizare = datetime.now(timezone.utc)
            db.session.commit()
            return jsonify({'success': True})
        return jsonify({'error': 'Status invalid'}), 400

    @app.route('/api/activitate/<int:aid>/comentariu', methods=['POST'])
    @login_required
    def api_activitate_comentariu(aid):
        a = Activitate.query.get_or_404(aid)
        mesaj = request.json.get('mesaj', '').strip()
        if not mesaj:
            return jsonify({'error': 'Mesaj gol'}), 400
        c = ComentariuActivitate(activitate_id=aid, utilizator_id=current_user.id, mesaj=mesaj)
        db.session.add(c)
        db.session.commit()
        return jsonify({'success': True, 'id': c.id, 'user': current_user.nume_complet,
                       'data': c.data_creare.strftime('%d.%m.%Y %H:%M')})

    @app.route('/api/activitate/<int:aid>/assign', methods=['POST'])
    @login_required
    def api_activitate_assign(aid):
        if not current_user.has_access('activitati_manage'):
            return jsonify({'error': 'Nu ai permisiunea de a modifica activități'}), 403
        a = Activitate.query.get_or_404(aid)
        a.asignat_id = int(request.json.get('asignat_id') or 0) or None
        db.session.commit()
        return jsonify({'success': True})

    # ════════════════════════════════════════════════════════════
    # RAPORT MARJĂ / PROFITABILITATE
    # ════════════════════════════════════════════════════════════

    @app.route('/admin/raport-marja')
    @module_required('facturi')
    def raport_marja():
        from sqlalchemy import func

        # Period filter (same system as dashboard)
        period = request.args.get('period', 'luna_curenta')
        custom_start = request.args.get('start', '')
        custom_end = request.args.get('end', '')
        today_date = date.today()

        if period == 'luna_curenta':
            p_start = today_date.replace(day=1)
            p_end = (today_date.replace(day=28) + timedelta(days=4)).replace(day=1)
            p_label = today_date.strftime('%B %Y')
        elif period == 'saptamana_curenta':
            p_start = today_date - timedelta(days=today_date.weekday())
            p_end = p_start + timedelta(days=7)
            p_label = f'{p_start.strftime("%d.%m")} – {(p_end - timedelta(days=1)).strftime("%d.%m.%Y")}'
        elif period == 'ultima_luna':
            first_this = today_date.replace(day=1)
            p_end = first_this
            p_start = (first_this - timedelta(days=1)).replace(day=1)
            p_label = p_start.strftime('%B %Y')
        elif period == 'an_curent':
            p_start = date(today_date.year, 1, 1)
            p_end = date(today_date.year + 1, 1, 1)
            p_label = str(today_date.year)
        elif period == 'an_trecut':
            p_start = date(today_date.year - 1, 1, 1)
            p_end = date(today_date.year, 1, 1)
            p_label = str(today_date.year - 1)
        elif period == 'custom' and custom_start and custom_end:
            try:
                p_start = datetime.strptime(custom_start, '%Y-%m-%d').date()
                p_end = datetime.strptime(custom_end, '%Y-%m-%d').date() + timedelta(days=1)
                p_label = f'{p_start.strftime("%d.%m.%Y")} – {(p_end - timedelta(days=1)).strftime("%d.%m.%Y")}'
            except:
                p_start = today_date.replace(day=1)
                p_end = (today_date.replace(day=28) + timedelta(days=4)).replace(day=1)
                p_label = today_date.strftime('%B %Y'); period = 'luna_curenta'
        else:
            p_start = today_date.replace(day=1)
            p_end = (today_date.replace(day=28) + timedelta(days=4)).replace(day=1)
            p_label = today_date.strftime('%B %Y'); period = 'luna_curenta'

        # Filters
        filter_client = request.args.get('client_id', '', type=str)
        filter_agent = request.args.get('agent_id', '', type=str)
        filter_produs = request.args.get('produs', '').strip()
        view_by = request.args.get('view', 'comenzi')  # comenzi | produse | clienti | agenti

        # Base query: comenzi in period (exclude anulat)
        cmd_q = Comanda.query.filter(
            Comanda.status != 'anulat',
            Comanda.data_comanda >= p_start, Comanda.data_comanda < p_end
        )
        if filter_client: cmd_q = cmd_q.filter(Comanda.client_id == int(filter_client))
        if filter_agent: cmd_q = cmd_q.filter(Comanda.creat_de_id == int(filter_agent))
        if current_user.doar_proprii: cmd_q = cmd_q.filter(Comanda.creat_de_id == current_user.id)

        comenzi = cmd_q.order_by(Comanda.data_comanda.desc()).all()

        # Build acquisition price map: cod_intern -> avg price
        pret_achizitie_map = {}
        for row in db.session.query(StocProdus.cod_intern, func.avg(StocProdus.pret_achizitie_mediu)).group_by(StocProdus.cod_intern).all():
            pret_achizitie_map[row[0]] = row[1] or 0
        # Also check historical NIR lines for codes no longer in stock
        for row in db.session.query(LinieNIR.cod_intern, func.avg(LinieNIR.pret_achizitie)).group_by(LinieNIR.cod_intern).all():
            if row[0] not in pret_achizitie_map:
                pret_achizitie_map[row[0]] = row[1] or 0

        # Calculate margins
        raport_comenzi = []
        raport_produse = {}
        raport_clienti = {}
        raport_agenti = {}
        totals = {'vanzare': 0, 'cost': 0, 'marja': 0}

        for cmd in comenzi:
            cmd_vanzare = 0
            cmd_cost = 0
            for lc in cmd.linii:
                if filter_produs and filter_produs.lower() not in (lc.cod or '').lower():
                    continue
                vanz = lc.valoare_linie
                cost_unitar = pret_achizitie_map.get(lc.cod, 0)
                cost = lc.cantitate * cost_unitar
                marja = vanz - cost
                cmd_vanzare += vanz
                cmd_cost += cost

                # Per produs
                key = lc.cod or 'N/A'
                if key not in raport_produse:
                    raport_produse[key] = {'cod': key, 'denumire': lc.denumire or '', 'cant': 0, 'vanzare': 0, 'cost': 0}
                raport_produse[key]['cant'] += lc.cantitate
                raport_produse[key]['vanzare'] += vanz
                raport_produse[key]['cost'] += cost

            if filter_produs and cmd_vanzare == 0:
                continue

            cmd_marja = cmd_vanzare - cmd_cost
            raport_comenzi.append({
                'cmd': cmd, 'vanzare': cmd_vanzare, 'cost': cmd_cost,
                'marja': cmd_marja, 'marja_pct': (cmd_marja / cmd_vanzare * 100) if cmd_vanzare else 0
            })
            totals['vanzare'] += cmd_vanzare
            totals['cost'] += cmd_cost
            totals['marja'] += cmd_marja

            # Per client
            cl_name = cmd.client.nume if cmd.client else 'Fără client'
            cl_id = cmd.client_id or 0
            if cl_id not in raport_clienti:
                raport_clienti[cl_id] = {'nume': cl_name, 'comenzi': 0, 'vanzare': 0, 'cost': 0}
            raport_clienti[cl_id]['comenzi'] += 1
            raport_clienti[cl_id]['vanzare'] += cmd_vanzare
            raport_clienti[cl_id]['cost'] += cmd_cost

            # Per agent
            ag_name = cmd.creat_de.nume_complet if cmd.creat_de else 'Necunoscut'
            ag_id = cmd.creat_de_id or 0
            if ag_id not in raport_agenti:
                raport_agenti[ag_id] = {'nume': ag_name, 'comenzi': 0, 'vanzare': 0, 'cost': 0}
            raport_agenti[ag_id]['comenzi'] += 1
            raport_agenti[ag_id]['vanzare'] += cmd_vanzare
            raport_agenti[ag_id]['cost'] += cmd_cost

        totals['marja_pct'] = (totals['marja'] / totals['vanzare'] * 100) if totals['vanzare'] else 0

        # Sort
        produse_sorted = sorted(raport_produse.values(), key=lambda x: x['vanzare'] - x['cost'], reverse=True)
        clienti_sorted = sorted(raport_clienti.values(), key=lambda x: x['vanzare'] - x['cost'], reverse=True)
        agenti_sorted = sorted(raport_agenti.values(), key=lambda x: x['vanzare'] - x['cost'], reverse=True)

        clienti_all = Client.query.filter_by(activ=True).order_by(Client.nume).all()
        agenti_all = Utilizator.query.filter_by(activ=True).order_by(Utilizator.nume_complet).all()

        return render_template('admin/raport_marja.html',
                             period=period, p_label=p_label,
                             p_start=p_start.strftime('%Y-%m-%d'),
                             p_end=(p_end - timedelta(days=1)).strftime('%Y-%m-%d'),
                             view_by=view_by,
                             filter_client=filter_client, filter_agent=filter_agent, filter_produs=filter_produs,
                             raport_comenzi=raport_comenzi, produse_sorted=produse_sorted,
                             clienti_sorted=clienti_sorted, agenti_sorted=agenti_sorted,
                             totals=totals,
                             clienti_all=clienti_all, agenti_all=agenti_all)

    # ════════════════════════════════════════════════════════════
    # WMS - MIȘCĂRI STOC
    # ════════════════════════════════════════════════════════════
    @app.route('/admin/wms')
    @module_required('wms')
    def wms_dashboard():
        stoc = StocProdus.query.order_by(StocProdus.ultima_miscare.desc()).limit(50).all()
        niruri = NIR.query.order_by(NIR.data_creare.desc()).limit(10).all()
        miscari = MiscareStoc.query.order_by(MiscareStoc.data.desc()).limit(20).all()
        return render_template('admin/wms.html', stoc=stoc, niruri=niruri, miscari=miscari)

    # --- FURNIZORI ---
    @app.route('/admin/wms/furnizori')
    @module_required('wms')
    def furnizori_list():
        furnizori = Furnizor.query.order_by(Furnizor.nume).all()
        return render_template('admin/wms_furnizori.html', furnizori=furnizori)

    @app.route('/admin/wms/furnizori/nou', methods=['GET','POST'])
    @module_required('wms')
    def furnizor_nou():
        if request.method == 'POST':
            f = Furnizor(nume=request.form.get('nume','').strip(),
                         cui=request.form.get('cui','').strip(),
                         contact=request.form.get('contact','').strip(),
                         telefon=request.form.get('telefon','').strip(),
                         email=request.form.get('email','').strip())
            db.session.add(f)
            db.session.commit()
            flash(f'Furnizor "{f.nume}" creat!', 'success')
            return redirect(url_for('furnizori_list'))
        return render_template('admin/wms_furnizor_form.html', furn=None)

    @app.route('/admin/wms/furnizori/<int:fid>', methods=['GET','POST'])
    @module_required('wms')
    def furnizor_edit(fid):
        f = Furnizor.query.get_or_404(fid)
        if request.method == 'POST':
            f.nume = request.form.get('nume', f.nume).strip()
            f.cui = request.form.get('cui','').strip()
            f.contact = request.form.get('contact','').strip()
            f.telefon = request.form.get('telefon','').strip()
            f.email = request.form.get('email','').strip()
            f.activ = 'activ' in request.form
            db.session.commit()
            flash('Furnizor actualizat!', 'success')
            return redirect(url_for('furnizori_list'))
        return render_template('admin/wms_furnizor_form.html', furn=f)

    # --- CELULE DEPOZIT ---
    @app.route('/admin/wms/celule')
    @module_required('wms')
    def celule_list():
        celule = CelulaDepozit.query.order_by(CelulaDepozit.cod).all()
        # Stock per cell
        stoc_per_celula = {}
        for s in StocProdus.query.filter(StocProdus.celula_id != None, StocProdus.cantitate > 0).all():
            stoc_per_celula.setdefault(s.celula_id, []).append(s)
        return render_template('admin/wms_celule.html', celule=celule, stoc_per_celula=stoc_per_celula)

    @app.route('/api/wms/celula', methods=['POST'])
    @login_required
    def api_celula_create():
        d = request.get_json()
        cod = d.get('cod','').strip()
        if not cod:
            return jsonify({'error': 'Cod obligatoriu'}), 400
        c = CelulaDepozit(cod=cod, zona=d.get('zona',''), raft=d.get('raft',''),
                          nivel=d.get('nivel',''), descriere=d.get('descriere',''),
                          barcode=d.get('barcode', cod))
        db.session.add(c)
        db.session.commit()
        return jsonify({'success': True, 'id': c.id})

    @app.route('/api/wms/celula/<int:cid>', methods=['DELETE'])
    @login_required
    def api_celula_delete(cid):
        c = CelulaDepozit.query.get_or_404(cid)
        db.session.delete(c)
        db.session.commit()
        return jsonify({'success': True})

    # --- MAPARE CODURI (API auto-suggest) ---
    @app.route('/api/wms/mapare-suggest')
    @login_required
    def api_mapare_suggest():
        """Suggest internal code from supplier code or EAN"""
        cod_furn = request.args.get('cod_furnizor', '').strip()
        cod_ean = request.args.get('cod_ean', '').strip()
        furnizor_id = request.args.get('furnizor_id', '')
        results = []
        if cod_furn:
            q = MapareCod.query.filter(MapareCod.cod_furnizor == cod_furn)
            if furnizor_id:
                q = q.filter(MapareCod.furnizor_id == int(furnizor_id))
            results = q.all()
        if not results and cod_ean:
            results = MapareCod.query.filter(MapareCod.cod_ean == cod_ean).all()
        if results:
            m = results[0]
            return jsonify({'found': True, 'cod_intern': m.cod_intern,
                           'denumire_furnizor': m.denumire_furnizor})
        return jsonify({'found': False})

    # --- NIR ---
    @app.route('/admin/wms/niruri')
    @module_required('wms')
    def niruri_list():
        niruri = NIR.query.order_by(NIR.data_creare.desc()).all()
        return render_template('admin/wms_niruri.html', niruri=niruri)

    @app.route('/admin/wms/nir/nou', methods=['GET','POST'])
    @module_required('wms')
    def nir_nou():
        furnizori = Furnizor.query.filter_by(activ=True).order_by(Furnizor.nume).all()
        if request.method == 'POST':
            nr = f"NIR-{datetime.now().strftime('%Y%m%d-%H%M%S')}"
            nir = NIR(
                numar=nr,
                furnizor_id=int(request.form.get('furnizor_id') or 0) or None,
                numar_factura_furnizor=request.form.get('numar_factura','').strip(),
                data_factura_furnizor=datetime.strptime(request.form['data_factura'],'%Y-%m-%d').date() if request.form.get('data_factura') else None,
                observatii=request.form.get('observatii','').strip(),
                creat_de_id=current_user.id,
            )
            db.session.add(nir)
            db.session.flush()

            # Parse lines
            idx = 0
            while f'cod_intern_{idx}' in request.form:
                cod_intern = request.form.get(f'cod_intern_{idx}','').strip()
                if cod_intern:
                    l = LinieNIR(
                        nir_id=nir.id, ordine=idx,
                        cod_furnizor=request.form.get(f'cod_furnizor_{idx}','').strip() or None,
                        cod_ean=request.form.get(f'cod_ean_{idx}','').strip() or None,
                        denumire_furnizor=request.form.get(f'den_furnizor_{idx}','').strip(),
                        cod_intern=cod_intern,
                        denumire_intern=request.form.get(f'den_intern_{idx}','').strip(),
                        um=request.form.get(f'um_{idx}','buc'),
                        cantitate=float(request.form.get(f'cant_{idx}') or 1),
                        pret_achizitie=float(request.form.get(f'pret_{idx}') or 0),
                    )
                    db.session.add(l)

                    # Auto-save mapping if supplier code or EAN provided
                    if l.cod_furnizor or l.cod_ean:
                        existing = MapareCod.query.filter(
                            MapareCod.cod_intern == cod_intern,
                            db.or_(
                                db.and_(MapareCod.cod_furnizor == l.cod_furnizor, l.cod_furnizor != None),
                                db.and_(MapareCod.cod_ean == l.cod_ean, l.cod_ean != None)
                            )
                        ).first()
                        if not existing:
                            m = MapareCod(furnizor_id=nir.furnizor_id,
                                         cod_furnizor=l.cod_furnizor, cod_ean=l.cod_ean,
                                         cod_intern=cod_intern, denumire_furnizor=l.denumire_furnizor)
                            db.session.add(m)
                idx += 1

            nir.recalculeaza()
            nir.status = 'scriptic'
            db.session.commit()
            flash(f'NIR {nir.numar} creat!', 'success')
            return redirect(url_for('nir_detail', nid=nir.id))
        # Get internal products + variants for autocomplete
        produse = ProdusConfig.query.filter_by(activ=True).order_by(ProdusConfig.cod).all()
        produse_list = []  # [{cod, denumire, variant}, ...]
        for p in produse:
            vc = p.variante_config
            variants = vc.get('variants', [])
            if variants:
                # Has variants: ONLY show full variant codes (prefix+suffix), skip base
                for v in variants:
                    suffix = v.get('code') or v.get('suffix', '')
                    if suffix:
                        # Build full code: if suffix already starts with prefix, use as-is
                        full_code = suffix if suffix.startswith(p.cod) else p.cod + suffix
                        combo = v.get('params') or v.get('combination') or {}
                        desc_parts = [p.denumire] + [str(val) for val in combo.values()]
                        produse_list.append({'cod': full_code, 'denumire': ' / '.join(desc_parts), 'variant': True})
            else:
                # No variants: show base product normally
                produse_list.append({'cod': p.cod, 'denumire': p.denumire, 'variant': False})
        return render_template('admin/wms_nir_form.html', nir=None, furnizori=furnizori, produse=produse_list)

    @app.route('/admin/wms/nir/<int:nid>')
    @module_required('wms')
    def nir_detail(nid):
        nir = NIR.query.get_or_404(nid)
        celule = CelulaDepozit.query.filter_by(activ=True).order_by(CelulaDepozit.cod).all()
        return render_template('admin/wms_nir_detail.html', nir=nir, celule=celule)

    @app.route('/api/wms/nir/<int:nid>/confirma-scriptic', methods=['POST'])
    @login_required
    def api_nir_confirma(nid):
        """Confirm NIR and add products to stock (scriptic)"""
        nir = NIR.query.get_or_404(nid)
        if nir.status not in ('draft', 'scriptic'):
            return jsonify({'error': 'NIR deja procesat'}), 400

        for l in nir.linii:
            # Update or create stock entry (no cell yet - scriptic only)
            stoc = StocProdus.query.filter_by(cod_intern=l.cod_intern, celula_id=None).first()
            if stoc:
                # Weighted average price
                total_val = stoc.cantitate * stoc.pret_achizitie_mediu + l.cantitate * l.pret_achizitie
                total_qty = stoc.cantitate + l.cantitate
                stoc.pret_achizitie_mediu = total_val / total_qty if total_qty else 0
                stoc.cantitate += l.cantitate
                stoc.ultima_miscare = datetime.now(timezone.utc)
            else:
                stoc = StocProdus(cod_intern=l.cod_intern, denumire=l.denumire_intern,
                                  cantitate=l.cantitate, pret_achizitie_mediu=l.pret_achizitie)
                db.session.add(stoc)

            # Log movement
            ms = MiscareStoc(tip='intrare_nir', cod_produs=l.cod_intern,
                            denumire_produs=l.denumire_intern, cantitate=l.cantitate,
                            nir_id=nir.id, numar_document=nir.numar,
                            utilizator_id=current_user.id)
            db.session.add(ms)

        nir.status = 'in_verificare'
        AuditLog.log('nir', nir.id, nir.numar, 'confirmat_scriptic',
                     f'Confirmat scriptic: {len(nir.linii)} linii', current_user.id)
        db.session.commit()
        return jsonify({'success': True})
    @login_required
    def api_nir_sterge(nid):
        nir = NIR.query.get_or_404(nid)
        numar = nir.numar
        furnizor_name = nir.furnizor.nume if nir.furnizor else ''
        db.session.delete(nir)
        AuditLog.log('nir', nid, numar, 'sters', f'NIR șters (furnizor: {furnizor_name})', current_user.id)
        db.session.commit()
        return jsonify({'success': True})

    # --- PRODUSE NECATALOGATE ---
    @app.route('/admin/wms/necatalogate')
    @module_required('wms')
    def wms_necatalogate():
        """Show products in stock that don't exist in the product catalog (base or variant)"""
        # Collect all known codes: base products (without variants) + full variant codes
        coduri_catalog = set()
        for p in ProdusConfig.query.all():
            vc = p.variante_config
            variants = vc.get('variants', [])
            if variants:
                for v in variants:
                    suffix = v.get('code') or v.get('suffix', '')
                    if suffix:
                        full_code = suffix if suffix.startswith(p.cod) else p.cod + suffix
                        coduri_catalog.add(full_code)
            else:
                coduri_catalog.add(p.cod)

        stoc_all = db.session.query(
            StocProdus.cod_intern,
            db.func.max(StocProdus.denumire).label('denumire'),
            db.func.sum(StocProdus.cantitate).label('cantitate_totala'),
            db.func.max(StocProdus.pret_achizitie_mediu).label('pret'),
        ).group_by(StocProdus.cod_intern).all()
        necatalogate = [s for s in stoc_all if s.cod_intern not in coduri_catalog]
        produse_catalog = ProdusConfig.query.filter_by(activ=True).order_by(ProdusConfig.cod).all()
        return render_template('admin/wms_necatalogate.html',
                              necatalogate=necatalogate, produse_catalog=produse_catalog)

    @app.route('/api/wms/remap-cod', methods=['POST'])
    @login_required
    def api_wms_remap_cod():
        """Remap a temporary/uncatalogued code to a real internal code"""
        d = request.get_json()
        cod_vechi = d.get('cod_vechi', '').strip()
        cod_nou = d.get('cod_nou', '').strip()
        if not cod_vechi or not cod_nou:
            return jsonify({'error': 'Ambele coduri sunt obligatorii'}), 400
        if cod_vechi == cod_nou:
            return jsonify({'error': 'Codurile sunt identice'}), 400

        # Update stoc
        stocuri = StocProdus.query.filter_by(cod_intern=cod_vechi).all()
        for s in stocuri:
            # Check if destination exists
            existing = StocProdus.query.filter_by(cod_intern=cod_nou, celula_id=s.celula_id).first()
            if existing:
                existing.cantitate += s.cantitate
                existing.ultima_miscare = datetime.now(timezone.utc)
                db.session.delete(s)
            else:
                s.cod_intern = cod_nou
                # Get denumire from catalog if exists
                prod = ProdusConfig.query.filter_by(cod=cod_nou).first()
                if prod:
                    s.denumire = prod.denumire

        # Update NIR lines
        LinieNIR.query.filter_by(cod_intern=cod_vechi).update({'cod_intern': cod_nou})
        # Update miscari stoc
        MiscareStoc.query.filter_by(cod_produs=cod_vechi).update({'cod_produs': cod_nou})
        # Update mapari
        MapareCod.query.filter_by(cod_intern=cod_vechi).update({'cod_intern': cod_nou})

        db.session.commit()
        return jsonify({'success': True, 'updated': len(stocuri)})

    @app.route('/api/wms/nir/linie/<int:lid>/verifica', methods=['POST'])
    @login_required
    def api_nir_linie_verifica(lid):
        """Add a partial verification to a NIR line (qty + cell)"""
        l = LinieNIR.query.get_or_404(lid)
        if l.nir.status not in ('in_verificare',):
            return jsonify({'error': 'NIR-ul nu este în verificare fizică'}), 400

        d = request.get_json()
        cant = float(d.get('cantitate', 0))
        if cant <= 0:
            return jsonify({'error': 'Cantitatea trebuie să fie > 0'}), 400

        celula_id = int(d.get('celula_id')) if d.get('celula_id') else None

        # Add verification entry
        ver = VerificareNIR(
            linie_nir_id=l.id,
            cantitate=cant,
            celula_id=celula_id,
            verificat_de_id=current_user.id
        )
        db.session.add(ver)

        # Move stock to specific cell
        if celula_id:
            stoc_nealocat = StocProdus.query.filter_by(cod_intern=l.cod_intern, celula_id=None).first()
            if stoc_nealocat:
                stoc_nealocat.cantitate = max(0, stoc_nealocat.cantitate - cant)
                if stoc_nealocat.cantitate == 0:
                    db.session.delete(stoc_nealocat)
            stoc_celula = StocProdus.query.filter_by(cod_intern=l.cod_intern, celula_id=celula_id).first()
            if stoc_celula:
                total_val = stoc_celula.cantitate * stoc_celula.pret_achizitie_mediu + cant * l.pret_achizitie
                total_qty = stoc_celula.cantitate + cant
                stoc_celula.pret_achizitie_mediu = total_val / total_qty if total_qty else 0
                stoc_celula.cantitate += cant
                stoc_celula.ultima_miscare = datetime.now(timezone.utc)
            else:
                stoc_celula = StocProdus(cod_intern=l.cod_intern, denumire=l.denumire_intern,
                                         celula_id=celula_id, cantitate=cant,
                                         pret_achizitie_mediu=l.pret_achizitie)
                db.session.add(stoc_celula)

        # Check if all lines complete -> auto-finish NIR
        nir = l.nir
        db.session.flush()  # ensure verificari are saved for property calc
        v, t = nir.progres_verificare
        if v >= t:
            nir.status = 'verificat'

        db.session.commit()
        vn, tn = nir.progres_verificare
        return jsonify({
            'success': True,
            'verificat_linii': vn, 'total_linii': tn,
            'linie_cant_verificata': l.cantitate_verificata,
            'linie_cant_asteptata': l.cantitate,
            'linie_complet': l.verificat_complet,
            'linie_rest': l.rest_de_verificat,
            'nir_status': nir.status
        })

    # ════════════════════════════════════════════════════════════
    # WMS - PICKING & LIVRARE
    # ════════════════════════════════════════════════════════════

    @app.route('/admin/wms/pickings')
    @module_required('wms')
    def picking_list():
        pickings = Picking.query.order_by(Picking.data_creare.desc()).all()
        return render_template('admin/wms_pickings.html', pickings=pickings)

    @app.route('/api/wms/picking/genereaza/<int:cid>', methods=['POST'])
    @login_required
    def api_picking_genereaza(cid):
        """Generate picking from a Comanda"""
        cmd = Comanda.query.get_or_404(cid)
        if cmd.status not in ('confirmata', 'productie', 'gata'):
            return jsonify({'error': 'Comanda nu este într-un status valid pentru picking'}), 400
        # Check if active picking already exists
        existing = Picking.query.filter_by(comanda_id=cid).filter(Picking.status.notin_(['anulat','livrat'])).first()
        if existing:
            return jsonify({'error': f'Există deja picking activ: {existing.numar}', 'picking_id': existing.id}), 400

        ts = datetime.now(timezone.utc).strftime('%Y%m%d-%H%M%S')
        pick = Picking(numar=f'PICK-{ts}', comanda_id=cid, creat_de_id=current_user.id)
        db.session.add(pick)
        db.session.flush()

        # Generate lines from comanda lines + find best cell for each
        for i, lc in enumerate(cmd.linii):
            cod = lc.cod or ''
            if not cod:
                continue
            # Find stock locations for this product, ordered by qty desc
            stocuri = StocProdus.query.filter(
                StocProdus.cod_intern == cod,
                StocProdus.celula_id != None,
                StocProdus.cantitate > 0
            ).order_by(StocProdus.cantitate.desc()).all()

            celula_id = stocuri[0].celula_id if stocuri else None
            stoc_disp = stocuri[0].cantitate if stocuri else 0

            lp = LiniePicking(
                picking_id=pick.id, ordine=i, cod_intern=cod,
                denumire=lc.denumire, um=lc.um,
                cantitate_ceruta=lc.cantitate,
                celula_sursa_id=celula_id,
                stoc_disponibil=stoc_disp
            )
            db.session.add(lp)

        db.session.commit()
        AuditLog.log('picking', pick.id, pick.numar, 'creat', f'Picking generat din comandă {cmd.numar}', current_user.id)
        AuditLog.log('comanda', cmd.id, cmd.numar, 'picking_generat', f'Picking {pick.numar} generat', current_user.id)
        db.session.commit()
        return jsonify({'success': True, 'picking_id': pick.id, 'numar': pick.numar})

    @app.route('/admin/wms/picking/<int:pid>')
    @module_required('wms')
    def picking_detail(pid):
        pick = Picking.query.get_or_404(pid)
        celule = CelulaDepozit.query.order_by(CelulaDepozit.cod).all()
        return render_template('admin/wms_picking_detail.html', picking=pick, celule=celule)

    @app.route('/api/wms/picking/<int:pid>/start', methods=['POST'])
    @login_required
    def api_picking_start(pid):
        pick = Picking.query.get_or_404(pid)
        if pick.status != 'nou':
            return jsonify({'error': 'Picking-ul nu este nou'}), 400
        pick.status = 'in_pregatire'
        db.session.commit()
        return jsonify({'success': True})

    @app.route('/api/wms/picking/linie/<int:lid>/prelua', methods=['POST'])
    @login_required
    def api_picking_linie_prelua(lid):
        """Pick a line: scan product from cell, decrease stock"""
        lp = LiniePicking.query.get_or_404(lid)
        pick = lp.picking
        if pick.status not in ('in_pregatire',):
            return jsonify({'error': 'Picking-ul nu este în pregătire'}), 400

        d = request.get_json()
        cant = float(d.get('cantitate', lp.cantitate_ceruta))
        celula_id = int(d.get('celula_id')) if d.get('celula_id') else lp.celula_sursa_id

        if cant <= 0:
            return jsonify({'error': 'Cantitate invalidă'}), 400

        # Decrease stock from cell
        if celula_id:
            stoc = StocProdus.query.filter_by(cod_intern=lp.cod_intern, celula_id=celula_id).first()
            if stoc:
                if stoc.cantitate < cant:
                    return jsonify({'error': f'Stoc insuficient în celulă: {stoc.cantitate} disponibil'}), 400
                stoc.cantitate -= cant
                stoc.ultima_miscare = datetime.now(timezone.utc)
                if stoc.cantitate == 0:
                    db.session.delete(stoc)
                # Log movement
                m = MiscareStoc(tip='iesire_comanda', cod_produs=lp.cod_intern, cantitate=cant,
                                celula_id=celula_id, comanda_id=pick.comanda_id,
                                utilizator_id=current_user.id, observatii=f'Picking {pick.numar}')
                db.session.add(m)
            else:
                return jsonify({'error': 'Nu există stoc pe această celulă'}), 400

        lp.preluata = True
        lp.cantitate_preluata = cant
        lp.celula_efectiva_id = celula_id
        lp.preluat_de_id = current_user.id
        lp.data_preluare = datetime.now(timezone.utc)

        # Check if all picked
        db.session.flush()
        p_done, p_total = pick.progres
        if p_done >= p_total:
            pick.status = 'complet'
            pick.data_finalizare = datetime.now(timezone.utc)

        db.session.commit()
        return jsonify({
            'success': True, 'preluate': p_done, 'total': p_total,
            'picking_status': pick.status
        })

    @app.route('/api/wms/picking/<int:pid>/nota-livrare', methods=['POST'])
    @login_required
    def api_picking_nota_livrare(pid):
        """Generate Nota de Livrare from completed picking"""
        pick = Picking.query.get_or_404(pid)
        if pick.status != 'complet':
            return jsonify({'error': 'Picking-ul nu este complet'}), 400
        # Check if already has nota
        if pick.nota_livrare:
            return jsonify({'error': 'Nota de livrare există deja', 'nota_id': pick.nota_livrare.id}), 400

        cmd = pick.comanda
        ts = datetime.now(timezone.utc).strftime('%Y%m%d-%H%M%S')
        nota = NotaLivrare(
            numar=f'NL-{ts}', picking_id=pick.id, comanda_id=cmd.id,
            client_id=cmd.client_id, adresa_livrare=cmd.adresa_livrare or '',
            creat_de_id=current_user.id
        )
        db.session.add(nota)

        # Update picking & comanda status
        pick.status = 'livrat'
        cmd.status = 'livrata'
        cmd.data_livrare_efectiva = date.today()
        db.session.commit()
        AuditLog.log('comanda', cmd.id, cmd.numar, 'livrata', f'Notă livrare {nota.numar} generată', current_user.id)
        db.session.commit()
        return jsonify({'success': True, 'nota_id': nota.id, 'numar': nota.numar})

    @app.route('/admin/wms/note-livrare')
    @module_required('wms')
    def note_livrare_list():
        note = NotaLivrare.query.order_by(NotaLivrare.data_creare.desc()).all()
        return render_template('admin/wms_note_livrare_list.html', note=note)

    @app.route('/admin/wms/nota-livrare/<int:nid>')
    @module_required('wms')
    def nota_livrare_detail(nid):
        nota = NotaLivrare.query.get_or_404(nid)
        return render_template('admin/wms_nota_livrare.html', nota=nota)

    @app.route('/admin/wms/nota-livrare/<int:nid>/pdf')
    @module_required('wms')
    def nota_livrare_pdf(nid):
        nota = NotaLivrare.query.get_or_404(nid)
        pick = nota.picking
        cmd = nota.comanda
        client = nota.client

        html = render_template('admin/wms_nota_livrare_pdf.html', nota=nota, picking=pick, comanda=cmd, client=client)

        try:
            from xhtml2pdf import pisa
            from io import BytesIO
            buf = BytesIO()
            pisa_status = pisa.CreatePDF(html, dest=buf, encoding='utf-8')
            if pisa_status.err:
                return html  # fallback to HTML
            buf.seek(0)
            response = app.make_response(buf.read())
            response.headers['Content-Type'] = 'application/pdf'
            response.headers['Content-Disposition'] = f'inline; filename=NL-{nota.numar}.pdf'
            return response
        except ImportError:
            # xhtml2pdf not installed, return HTML with print hint
            return html + '<script>window.print()</script>'

    # ════════════════════════════════════════════════════════════
    # WMS - TRANSFER ÎNTRE CELULE
    # ════════════════════════════════════════════════════════════

    @app.route('/admin/wms/transfer')
    @module_required('wms')
    def wms_transfer():
        celule = CelulaDepozit.query.order_by(CelulaDepozit.cod).all()
        stocuri = StocProdus.query.filter(StocProdus.celula_id != None, StocProdus.cantitate > 0).order_by(StocProdus.cod_intern).all()
        return render_template('admin/wms_transfer.html', celule=celule, stocuri=stocuri)

    @app.route('/api/wms/transfer', methods=['POST'])
    @login_required
    def api_wms_transfer():
        d = request.get_json()
        cod = d.get('cod_intern', '').strip()
        sursa_id = int(d.get('celula_sursa_id'))
        dest_id = int(d.get('celula_destinatie_id'))
        cant = float(d.get('cantitate', 0))

        if not cod or cant <= 0:
            return jsonify({'error': 'Date invalide'}), 400
        if sursa_id == dest_id:
            return jsonify({'error': 'Sursa și destinația nu pot fi aceeași celulă'}), 400

        stoc_sursa = StocProdus.query.filter_by(cod_intern=cod, celula_id=sursa_id).first()
        if not stoc_sursa or stoc_sursa.cantitate < cant:
            disp = stoc_sursa.cantitate if stoc_sursa else 0
            return jsonify({'error': f'Stoc insuficient: {disp} disponibil'}), 400

        # Decrease source
        stoc_sursa.cantitate -= cant
        stoc_sursa.ultima_miscare = datetime.now(timezone.utc)
        if stoc_sursa.cantitate == 0:
            db.session.delete(stoc_sursa)

        # Increase destination
        stoc_dest = StocProdus.query.filter_by(cod_intern=cod, celula_id=dest_id).first()
        if stoc_dest:
            total_val = stoc_dest.cantitate * stoc_dest.pret_achizitie_mediu + cant * stoc_sursa.pret_achizitie_mediu
            total_qty = stoc_dest.cantitate + cant
            stoc_dest.pret_achizitie_mediu = total_val / total_qty if total_qty else 0
            stoc_dest.cantitate += cant
            stoc_dest.ultima_miscare = datetime.now(timezone.utc)
        else:
            stoc_dest = StocProdus(cod_intern=cod, denumire=stoc_sursa.denumire,
                                    celula_id=dest_id, cantitate=cant,
                                    pret_achizitie_mediu=stoc_sursa.pret_achizitie_mediu)
            db.session.add(stoc_dest)

        # Log
        m = MiscareStoc(tip='transfer', cod_produs=cod, cantitate=cant,
                         celula_id=sursa_id, celula_destinatie_id=dest_id,
                         utilizator_id=current_user.id,
                         observatii=f'Transfer {cant} {cod}')
        db.session.add(m)
        db.session.commit()

        sursa_cel = CelulaDepozit.query.get(sursa_id)
        dest_cel = CelulaDepozit.query.get(dest_id)
        return jsonify({'success': True,
                         'message': f'{cant:.0f} × {cod}: {sursa_cel.cod} → {dest_cel.cod}'})

    # ════════════════════════════════════════════════════════════
    # WMS - ALERTE STOC MINIM
    # ════════════════════════════════════════════════════════════

    @app.route('/admin/wms/alerte-stoc')
    @module_required('wms')
    def wms_alerte_stoc():
        praguri = StocMinim.query.filter_by(activ=True).order_by(StocMinim.cod_intern).all()
        # Build all product codes for dropdown
        produse_list = []
        for p in ProdusConfig.query.filter_by(activ=True).order_by(ProdusConfig.cod).all():
            vc = p.variante_config
            variants = vc.get('variants', [])
            if variants:
                for v in variants:
                    suffix = v.get('code') or v.get('suffix', '')
                    if suffix:
                        full_code = suffix if suffix.startswith(p.cod) else p.cod + suffix
                        combo = v.get('params') or v.get('combination') or {}
                        desc = p.denumire + ' / ' + ' / '.join(str(val) for val in combo.values())
                        produse_list.append({'cod': full_code, 'den': desc})
            else:
                produse_list.append({'cod': p.cod, 'den': p.denumire})
        return render_template('admin/wms_alerte_stoc.html', praguri=praguri, produse=produse_list)

    @app.route('/api/wms/stoc-minim', methods=['POST'])
    @login_required
    def api_stoc_minim_save():
        d = request.get_json()
        cod = d.get('cod_intern', '').strip()
        prag = float(d.get('prag_minim', 0))
        den = d.get('denumire', '')
        if not cod or prag < 0:
            return jsonify({'error': 'Date invalide'}), 400
        existing = StocMinim.query.filter_by(cod_intern=cod).first()
        if existing:
            existing.prag_minim = prag
            existing.denumire = den or existing.denumire
            existing.activ = True
        else:
            sm = StocMinim(cod_intern=cod, denumire=den, prag_minim=prag)
            db.session.add(sm)
        db.session.commit()
        return jsonify({'success': True})

    @app.route('/api/wms/stoc-minim/<int:sid>/delete', methods=['POST'])
    @login_required
    def api_stoc_minim_delete(sid):
        sm = StocMinim.query.get_or_404(sid)
        db.session.delete(sm)
        db.session.commit()
        return jsonify({'success': True})

    # ════════════════════════════════════════════════════════════
    # NOMENCLATOR PRODUSE (read-only view of cfg_produse)
    # ════════════════════════════════════════════════════════════
    @app.route('/admin/nomenclator')
    @module_required('nomenclator')
    def nomenclator_list():
        cat_id = request.args.get('categorie', '', type=str)
        search = request.args.get('q', '').strip()
        categorii = CategorieProdus.query.order_by(CategorieProdus.ordine).all()

        query = ProdusConfig.query.filter_by(activ=True)
        if cat_id:
            query = query.join(ProdusCategorie).filter(ProdusCategorie.categorie_id == int(cat_id))
        if search:
            query = query.filter(db.or_(
                ProdusConfig.cod.ilike(f'%{search}%'),
                ProdusConfig.denumire.ilike(f'%{search}%')
            ))
        produse = query.order_by(ProdusConfig.denumire).all()

        accesorii_query = Accesoriu.query.filter_by(activ=True)
        if search:
            accesorii_query = accesorii_query.filter(db.or_(
                Accesoriu.cod.ilike(f'%{search}%'),
                Accesoriu.denumire.ilike(f'%{search}%')
            ))
        accesorii = accesorii_query.order_by(Accesoriu.denumire).all()
        cat_accesorii = CategorieAccesoriu.query.order_by(CategorieAccesoriu.ordine).all()

        # Build stock lookup: cod -> total qty
        stoc_map = {}
        for s in db.session.query(StocProdus.cod_intern, db.func.sum(StocProdus.cantitate)).group_by(StocProdus.cod_intern).all():
            stoc_map[s[0]] = s[1]

        return render_template('admin/nomenclator.html', produse=produse, accesorii=accesorii,
                             categorii=categorii, cat_accesorii=cat_accesorii,
                             cat_filter=cat_id, search=search, stoc_map=stoc_map)

    @app.route('/admin/nomenclator/produs/<int:pid>')
    @module_required('nomenclator')
    def nomenclator_produs(pid):
        p = ProdusConfig.query.get_or_404(pid)
        from sqlalchemy import and_
        compats = db.session.query(Accesoriu, AccesoriuCompat.status).join(
            AccesoriuCompat, and_(AccesoriuCompat.accesoriu_id == Accesoriu.id,
                                   AccesoriuCompat.produs_id == p.id)
        ).filter(Accesoriu.activ == True).all()

        # Stock per variant/base
        stoc_items = []
        vc = p.variante_config
        variants = vc.get('variants', [])
        if variants:
            for v in variants:
                suffix = v.get('code') or v.get('suffix', '')
                if suffix:
                    full_code = suffix if suffix.startswith(p.cod) else p.cod + suffix
                    combo = v.get('params') or v.get('combination') or {}
                    desc = ' / '.join(str(val) for val in combo.values())
                    stocuri = StocProdus.query.filter_by(cod_intern=full_code).all()
                    total = sum(s.cantitate for s in stocuri)
                    stoc_items.append({'cod': full_code, 'desc': desc, 'total': total, 'celule': stocuri})
        else:
            stocuri = StocProdus.query.filter_by(cod_intern=p.cod).all()
            total = sum(s.cantitate for s in stocuri)
            stoc_items.append({'cod': p.cod, 'desc': '', 'total': total, 'celule': stocuri})

        return render_template('admin/nomenclator_produs.html', produs=p, compats=compats, stoc_items=stoc_items)

    # ════════════════════════════════════════════════════════════
    # NOMENCLATOR - IMPORT EXCEL
    # ════════════════════════════════════════════════════════════
    @app.route('/admin/nomenclator/import', methods=['GET', 'POST'])
    @module_required('nomenclator')
    def nomenclator_import():
        """Import products, parameters, variants & accessories from Excel template"""
        if request.method == 'GET':
            return render_template('admin/nomenclator_import.html')

        f = request.files.get('file')
        if not f or not f.filename.endswith('.xlsx'):
            flash('Încărcați un fișier .xlsx valid.', 'danger')
            return redirect('/admin/nomenclator/import')

        import openpyxl
        try:
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        except Exception as e:
            flash(f'Eroare la citirea fișierului: {e}', 'danger')
            return redirect('/admin/nomenclator/import')

        def _bool(val):
            if not val: return False
            return str(val).strip().upper() in ('DA', 'YES', 'TRUE', '1', 'Y')

        def _str(val):
            return str(val).strip() if val else ''

        results = {
            'produse':    {'added': 0, 'skipped': 0, 'errors': []},
            'parametri':  {'added': 0, 'errors': []},
            'variante':   {'added': 0, 'errors': []},
            'accesorii':  {'added': 0, 'skipped': 0, 'errors': []},
        }

        # ── 1. Import Produse (bază) ──
        produs_map = {}  # cod -> ProdusConfig object
        if 'Produse' in wb.sheetnames:
            for i, row in enumerate(wb['Produse'].iter_rows(min_row=3, values_only=True), 3):
                if not row or not row[0] or not _str(row[0]):
                    continue
                cod = _str(row[0])
                denumire = _str(row[1])
                if not denumire:
                    results['produse']['errors'].append(f'Rând {i}: cod "{cod}" fără denumire')
                    continue
                existing = ProdusConfig.query.filter_by(cod=cod).first()
                if existing:
                    produs_map[cod] = existing
                    results['produse']['skipped'] += 1
                    continue
                try:
                    p = ProdusConfig(
                        cod=cod, denumire=denumire,
                        descriere=_str(row[2]) if len(row) > 2 else '',
                        pret=float(row[3]) if len(row) > 3 and row[3] else 0,
                        um=_str(row[4]).lower() or 'buc' if len(row) > 4 else 'buc',
                    )
                    db.session.add(p); db.session.flush()
                    cat_name = _str(row[5]) if len(row) > 5 else ''
                    if cat_name:
                        cat = CategorieProdus.query.filter(CategorieProdus.nume.ilike(cat_name)).first()
                        if not cat:
                            mx = db.session.query(db.func.max(CategorieProdus.ordine)).scalar() or 0
                            cat = CategorieProdus(nume=cat_name, ordine=mx + 1)
                            db.session.add(cat); db.session.flush()
                        db.session.add(ProdusCategorie(produs_id=p.id, categorie_id=cat.id))
                    db.session.commit()
                    produs_map[cod] = p
                    results['produse']['added'] += 1
                except Exception as e:
                    db.session.rollback()
                    results['produse']['errors'].append(f'Rând {i}: {e}')

        # ── 2. Import Parametri ──
        if 'Parametri' in wb.sheetnames:
            # Group by produs cod
            param_rows = {}
            for i, row in enumerate(wb['Parametri'].iter_rows(min_row=3, values_only=True), 3):
                if not row or not row[0] or not _str(row[0]):
                    continue
                cod = _str(row[0])
                if cod not in param_rows:
                    param_rows[cod] = []
                param_rows[cod].append((i, row))

            for cod, rows in param_rows.items():
                p = produs_map.get(cod) or ProdusConfig.query.filter_by(cod=cod).first()
                if not p:
                    for i, _ in rows:
                        results['parametri']['errors'].append(f'Rând {i}: produs "{cod}" nu există')
                    continue

                params_list = list(p.parametri_config)  # existing params
                existing_keys = {param.get('key') for param in params_list}

                for i, row in rows:
                    try:
                        name = _str(row[1])
                        key = _str(row[2])
                        tip = _str(row[3]).lower() if len(row) > 3 else 'dropdown'
                        if not name or not key:
                            results['parametri']['errors'].append(f'Rând {i}: nume sau cheie lipsă')
                            continue
                        if key in existing_keys:
                            continue  # skip duplicate param key

                        param = {
                            'name': name,
                            'key': key,
                            'type': tip if tip in ('dropdown', 'dimension') else 'dropdown',
                            'affects_price': _bool(row[5]) if len(row) > 5 else True,
                        }

                        if tip == 'dropdown':
                            mode = _str(row[4]).lower() if len(row) > 4 else 'fix'
                            param['mode'] = mode if mode in ('fix', 'mixt', 'completabil') else 'fix'
                            vals_str = _str(row[6]) if len(row) > 6 else ''
                            param['values'] = [v.strip() for v in vals_str.split('|') if v.strip()] if vals_str else []
                        else:  # dimension
                            dims_str = _str(row[7]) if len(row) > 7 else ''
                            has_canate = _bool(row[8]) if len(row) > 8 else False
                            custom_dim = _bool(row[9]) if len(row) > 9 else True
                            param['has_canate'] = has_canate
                            param['custom_dimensions'] = custom_dim
                            dims = []
                            if dims_str:
                                for d in dims_str.split('|'):
                                    d = d.strip()
                                    if not d: continue
                                    parts = d.split(',')
                                    dim = {'value': parts[0].strip()}
                                    if len(parts) > 1 and parts[1].strip():
                                        dim['canate'] = parts[1].strip()
                                    dims.append(dim)
                            param['predefined_dimensions'] = dims

                        params_list.append(param)
                        existing_keys.add(key)
                        results['parametri']['added'] += 1
                    except Exception as e:
                        results['parametri']['errors'].append(f'Rând {i}: {e}')

                try:
                    p.parametri_config = params_list
                    db.session.commit()
                except Exception as e:
                    db.session.rollback()
                    results['parametri']['errors'].append(f'Produs {cod}: eroare salvare parametri - {e}')

        # ── 3. Import Variante ──
        if 'Variante' in wb.sheetnames:
            ws_var = wb['Variante']
            # Read header row to find Param: columns
            header_row = list(ws_var.iter_rows(min_row=1, max_row=1, values_only=True))[0]
            param_cols = {}  # col_index -> param_key
            for ci, h in enumerate(header_row):
                if h and str(h).strip().startswith('Param:'):
                    param_key = str(h).strip()[6:].strip()
                    param_cols[ci] = param_key

            # Group variants by produs cod
            var_rows = {}
            for i, row in enumerate(ws_var.iter_rows(min_row=3, values_only=True), 3):
                if not row or not row[0] or not _str(row[0]):
                    continue
                cod = _str(row[0])
                if cod not in var_rows:
                    var_rows[cod] = []
                var_rows[cod].append((i, row))

            for cod, rows in var_rows.items():
                p = produs_map.get(cod) or ProdusConfig.query.filter_by(cod=cod).first()
                if not p:
                    for i, _ in rows:
                        results['variante']['errors'].append(f'Rând {i}: produs "{cod}" nu există')
                    continue

                vc = p.variante_config or {}
                variants = list(vc.get('variants', []))
                existing_suffixes = {v.get('suffix', '') for v in variants}

                # Determine variant_parameters from Param: columns (exclude 'canate' - it's part of dimension)
                var_param_keys = [pk for pk in param_cols.values() if pk != 'canate']

                for i, row in rows:
                    try:
                        var_code = _str(row[1])
                        price = float(row[2]) if len(row) > 2 and row[2] else 0
                        if not var_code:
                            results['variante']['errors'].append(f'Rând {i}: cod variantă lipsă')
                            continue

                        # Build suffix: variant code is the full code, suffix = what comes after base cod
                        # If var_code starts with base cod, suffix is the rest; otherwise use as-is
                        suffix = var_code
                        if var_code.startswith(p.cod):
                            suffix = var_code[len(p.cod):]
                        # Ensure suffix starts with a separator if not empty
                        if suffix and not suffix[0] in ('.', '-', '_', ' '):
                            suffix = '.' + suffix

                        if suffix in existing_suffixes:
                            continue

                        # Build combination dict from Param: columns
                        combination = {}
                        for ci, pkey in param_cols.items():
                            if ci < len(row) and row[ci] is not None:
                                val = row[ci]
                                if pkey == 'canate':
                                    val = str(int(val)) if isinstance(val, (int, float)) else str(val).strip()
                                elif isinstance(val, (int, float)):
                                    val = str(int(val)) if val == int(val) else str(val)
                                else:
                                    val = str(val).strip()
                                combination[pkey] = val

                        variant = {'suffix': suffix, 'combination': combination, 'pret': price}
                        variants.append(variant)
                        existing_suffixes.add(suffix)
                        results['variante']['added'] += 1
                    except Exception as e:
                        results['variante']['errors'].append(f'Rând {i}: {e}')

                try:
                    vc['variants'] = variants
                    vc['use_variants'] = True
                    # Set variant_parameters if not already set
                    if not vc.get('variant_parameters'):
                        vc['variant_parameters'] = var_param_keys
                    p.variante_config = vc
                    db.session.commit()
                except Exception as e:
                    db.session.rollback()
                    results['variante']['errors'].append(f'Produs {cod}: eroare salvare variante - {e}')

        # ── 4. Import Accesorii ──
        if 'Accesorii' in wb.sheetnames:
            for i, row in enumerate(wb['Accesorii'].iter_rows(min_row=3, values_only=True), 3):
                if not row or not row[0] or not _str(row[0]):
                    continue
                cod = _str(row[0])
                denumire = _str(row[1])
                if not denumire:
                    results['accesorii']['errors'].append(f'Rând {i}: cod "{cod}" fără denumire')
                    continue
                if Accesoriu.query.filter_by(cod=cod).first():
                    results['accesorii']['skipped'] += 1
                    continue
                try:
                    tip_val = _str(row[3]).lower() if len(row) > 3 and row[3] else 'accesoriu'
                    if tip_val not in ('accesoriu', 'sticla', 'panou'):
                        tip_val = 'accesoriu'
                    pret_mode = _str(row[5]).lower() if len(row) > 5 and row[5] else 'fix'
                    if pret_mode not in ('fix', 'per_mp', 'per_ml'):
                        pret_mode = 'fix'
                    a = Accesoriu(
                        cod=cod, denumire=denumire,
                        descriere=_str(row[2]) if len(row) > 2 else '',
                        tip=tip_val,
                        pret=float(row[4]) if len(row) > 4 and row[4] else 0,
                        pret_mode=pret_mode,
                        um=_str(row[6]).lower() or 'buc' if len(row) > 6 else 'buc',
                        poate_standalone=not _bool(row[8]) if len(row) > 8 and _str(row[8]).upper() == 'NU' else True,
                    )
                    cat_name = _str(row[7]) if len(row) > 7 else ''
                    if cat_name:
                        cat = CategorieAccesoriu.query.filter(CategorieAccesoriu.nume.ilike(cat_name)).first()
                        if not cat:
                            mx = db.session.query(db.func.max(CategorieAccesoriu.ordine)).scalar() or 0
                            cat = CategorieAccesoriu(nume=cat_name, ordine=mx + 1)
                            db.session.add(cat); db.session.flush()
                        a.categorie_id = cat.id
                    db.session.add(a); db.session.commit()
                    results['accesorii']['added'] += 1
                except Exception as e:
                    db.session.rollback()
                    results['accesorii']['errors'].append(f'Rând {i}: {e}')

        wb.close()
        return render_template('admin/nomenclator_import.html', results=results)

    @app.route('/admin/nomenclator/import/template')
    @module_required('nomenclator')
    def nomenclator_import_template():
        """Download import template Excel"""
        import os
        template_path = os.path.join(app.static_folder, 'template_import_nomenclator.xlsx')
        if os.path.exists(template_path):
            return send_file(template_path, as_attachment=True, download_name='template_import_nomenclator.xlsx')
        # Generate on the fly
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO

        wb = openpyxl.Workbook()
        hf = Font(bold=True, color='FFFFFF', size=11, name='Arial')
        hfill = PatternFill('solid', fgColor='1D2F34')
        ha = Alignment(horizontal='center', vertical='center', wrap_text=True)
        hint_font = Font(italic=True, color='888888', size=9, name='Arial')

        # Produse sheet
        ws = wb.active; ws.title = "Produse"
        for ci, (hdr, w, hint) in enumerate([
            ('Cod *',15,'Cod unic produs'), ('Denumire *',40,'Numele complet'),
            ('Descriere',50,'Opțional'), ('Preț (EUR)',12,'Preț de bază'),
            ('UM',8,'buc/mp/ml'), ('Categorie',20,'Se creează automat')
        ], 1):
            c = ws.cell(row=1, column=ci, value=hdr); c.font=hf; c.fill=hfill; c.alignment=ha
            ws.cell(row=2, column=ci, value=hint).font = hint_font
            ws.column_dimensions[chr(64+ci)].width = w
        ws.freeze_panes = 'A3'

        # Accesorii sheet
        ws2 = wb.create_sheet("Accesorii")
        for ci, (hdr, w, hint) in enumerate([
            ('Cod *',15,'Cod unic'), ('Denumire *',40,'Numele complet'),
            ('Descriere',40,'Opțional'), ('Tip',12,'accesoriu/sticla/panou'),
            ('Preț (EUR)',12,'Preț'), ('Mod Preț',12,'fix/per_mp/per_ml'),
            ('UM',8,'buc/mp/ml'), ('Categorie',20,'Categorie accesorii'),
            ('Standalone',12,'DA/NU')
        ], 1):
            c = ws2.cell(row=1, column=ci, value=hdr); c.font=hf; c.fill=hfill; c.alignment=ha
            ws2.cell(row=2, column=ci, value=hint).font = hint_font
            col_letter = chr(64+ci) if ci <= 9 else chr(64+ci)
            ws2.column_dimensions[chr(64+ci) if ci < 10 else 'I'].width = w
        ws2.freeze_panes = 'A3'

        buf = BytesIO()
        wb.save(buf); buf.seek(0)
        return send_file(buf, as_attachment=True, download_name='template_import_nomenclator.xlsx',
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    # ════════════════════════════════════════════════════════════
    # CONFIGURATOR PAGE
    # ════════════════════════════════════════════════════════════
    @app.route('/configurator')
    @module_required('configurator')
    def configurator_page():
        cat_produse = [{'id': c.id, 'nume': c.nume} for c in CategorieProdus.query.order_by(CategorieProdus.ordine).all()]
        cat_accesorii = [{'id': c.id, 'nume': c.nume} for c in CategorieAccesoriu.query.order_by(CategorieAccesoriu.ordine).all()]
        clienti = Client.query.filter_by(activ=True).order_by(Client.nume).all()
        return render_template('configurator/index.html', cat_produse=cat_produse, cat_accesorii=cat_accesorii, clienti=clienti)

    @app.route('/configurator/admin')
    @module_required('cfg_admin')
    def cfg_admin_panel():
        return render_template('admin/cfg_panel.html')

    # ════════════════════════════════════════════════════════════
    # CONFIGURATOR CRUD APIs
    # ════════════════════════════════════════════════════════════

    # --- Products ---
    @app.route('/api/cfg/produse', methods=['GET'])
    @login_required
    def api_cfg_produse_list():
        produse = ProdusConfig.query.filter_by(activ=True).order_by(ProdusConfig.cod).all()
        show_cost = current_user.has_access('preturi_furnizor')
        return jsonify([{
            'id': p.id, 'cod': p.cod, 'denumire': p.denumire, 'pret': p.pret,
            'um': p.um, 'descriere': p.descriere,
            **({'pret_furnizor': p.pret_furnizor, 'discount_furnizor': p.discount_furnizor,
                'adaos_hsl': p.adaos_hsl, 'pret_achizitie': p.pret_achizitie,
                'pret_minim_vanzare': p.pret_minim_vanzare} if show_cost else {}),
            'categories': [pc.categorie_id for pc in p.categorii],
            'parametri_config': p.parametri_config, 'variante_config': p.variante_config,
        } for p in produse])

    @app.route('/api/cfg/produse', methods=['POST'])
    @login_required
    def api_cfg_produs_create():
        d = request.get_json()
        p = ProdusConfig(cod=d['cod'], denumire=d['denumire'], pret=d.get('pret',0),
                         um=d.get('um','buc'), descriere=d.get('descriere',''),
                         pret_furnizor=d.get('pret_furnizor',0),
                         discount_furnizor=d.get('discount_furnizor',0),
                         adaos_hsl=d.get('adaos_hsl',0))
        db.session.add(p)
        db.session.flush()
        for cat_id in d.get('categories', []):
            db.session.add(ProdusCategorie(produs_id=p.id, categorie_id=cat_id))
        db.session.commit()
        return jsonify({'success': True, 'id': p.id})

    @app.route('/api/cfg/produse/<int:pid>', methods=['PUT'])
    @login_required
    def api_cfg_produs_update(pid):
        p = ProdusConfig.query.get_or_404(pid)
        d = request.get_json()
        for k in ['cod','denumire','pret','um','descriere','pret_furnizor','discount_furnizor','adaos_hsl']:
            if k in d: setattr(p, k, d[k])
        if 'categories' in d:
            ProdusCategorie.query.filter_by(produs_id=pid).delete()
            for cat_id in d['categories']:
                db.session.add(ProdusCategorie(produs_id=pid, categorie_id=cat_id))
        db.session.commit()
        return jsonify({'success': True})

    @app.route('/api/cfg/produse/<int:pid>', methods=['DELETE'])
    @login_required
    def api_cfg_produs_delete(pid):
        p = ProdusConfig.query.get_or_404(pid)
        p.activ = False
        db.session.commit()
        return jsonify({'success': True})

    # --- Liste de Prețuri (Șabloane) ---
    @app.route('/api/cfg/liste-preturi', methods=['GET'])
    @login_required
    def api_cfg_liste_preturi_list():
        sabloane = SablonListaPreturi.query.filter_by(activ=True).order_by(SablonListaPreturi.nume).all()
        return jsonify([{
            'id': s.id, 'nume': s.nume, 'descriere': s.descriere,
            'discount_global': s.discount_global, 'reguli': s.reguli,
            'nr_clienti': Client.query.filter_by(sablon_pret_id=s.id).count(),
        } for s in sabloane])

    @app.route('/api/cfg/liste-preturi', methods=['POST'])
    @login_required
    def api_cfg_lista_preturi_create():
        d = request.get_json()
        s = SablonListaPreturi(nume=d['nume'], discount_global=d.get('discount_global',0),
                               descriere=d.get('descriere',''))
        s.reguli = d.get('reguli', {})
        db.session.add(s)
        db.session.commit()
        return jsonify({'success': True, 'id': s.id})

    @app.route('/api/cfg/liste-preturi/<int:sid>', methods=['PUT'])
    @login_required
    def api_cfg_lista_preturi_update(sid):
        s = SablonListaPreturi.query.get_or_404(sid)
        d = request.get_json()
        for k in ['nume','discount_global','descriere']:
            if k in d: setattr(s, k, d[k])
        if 'reguli' in d: s.reguli = d['reguli']
        db.session.commit()
        return jsonify({'success': True})

    @app.route('/api/cfg/liste-preturi/<int:sid>', methods=['DELETE'])
    @login_required
    def api_cfg_lista_preturi_delete(sid):
        s = SablonListaPreturi.query.get_or_404(sid)
        s.activ = False
        db.session.commit()
        return jsonify({'success': True})

    @app.route('/api/cfg/produse/<int:pid>/config', methods=['GET'])
    @login_required
    def api_cfg_produs_config_get(pid):
        p = ProdusConfig.query.get_or_404(pid)
        return jsonify({'parameter_types': p.parametri_config})

    @app.route('/api/cfg/produse/<int:pid>/config', methods=['PUT'])
    @login_required
    def api_cfg_produs_config_set(pid):
        p = ProdusConfig.query.get_or_404(pid)
        d = request.get_json()
        p.parametri_config = d.get('parameter_types', [])
        db.session.commit()
        return jsonify({'success': True})

    @app.route('/api/cfg/produse/<int:pid>/variante', methods=['GET'])
    @login_required
    def api_cfg_produs_variante_get(pid):
        p = ProdusConfig.query.get_or_404(pid)
        vc = p.variante_config or {}
        return jsonify(vc)

    @app.route('/api/cfg/produse/<int:pid>/variante', methods=['PUT'])
    @login_required
    def api_cfg_produs_variante_set(pid):
        p = ProdusConfig.query.get_or_404(pid)
        d = request.get_json()
        p.variante_config = d
        db.session.commit()
        return jsonify({'success': True})

    # --- Accessories ---
    @app.route('/api/cfg/accesorii', methods=['GET'])
    @login_required
    def api_cfg_accesorii_list():
        accs = Accesoriu.query.filter_by(activ=True).order_by(Accesoriu.cod).all()
        return jsonify([{
            'id': a.id, 'cod': a.cod, 'denumire': a.denumire, 'pret': a.pret,
            'pret_mode': a.pret_mode, 'um': a.um, 'tip': a.tip,
            'poate_standalone': a.poate_standalone, 'descriere': a.descriere,
            'categorie_id': a.categorie_id,
            'pret_furnizor': a.pret_furnizor, 'discount_furnizor': a.discount_furnizor,
            'adaos_hsl': a.adaos_hsl,
            'compatibilitati': [{'produs_id': c.produs_id, 'status': c.status, 'variante_compat': c.variante_compat} for c in a.compatibilitati]
        } for a in accs])

    @app.route('/api/cfg/accesorii', methods=['POST'])
    @login_required
    def api_cfg_accesoriu_create():
        d = request.get_json()
        a = Accesoriu(cod=d['cod'], denumire=d['denumire'], pret=d.get('pret',0),
                      pret_mode=d.get('pret_mode','fix'), um=d.get('um','buc'),
                      tip=d.get('tip','accesoriu'), poate_standalone=d.get('poate_standalone',True),
                      descriere=d.get('descriere',''), categorie_id=d.get('categorie_id'),
                      pret_furnizor=d.get('pret_furnizor',0),
                      discount_furnizor=d.get('discount_furnizor',0),
                      adaos_hsl=d.get('adaos_hsl',0))
        db.session.add(a)
        db.session.flush()
        for comp in d.get('compatibilitati', []):
            ac = AccesoriuCompat(accesoriu_id=a.id, produs_id=comp['produs_id'], status=comp.get('status','optional'))
            ac.variante_compat = comp.get('variante_compat', [])
            db.session.add(ac)
        db.session.commit()
        return jsonify({'success': True, 'id': a.id})

    @app.route('/api/cfg/accesorii/<int:aid>', methods=['PUT'])
    @login_required
    def api_cfg_accesoriu_update(aid):
        a = Accesoriu.query.get_or_404(aid)
        d = request.get_json()
        for k in ['cod','denumire','pret','pret_mode','um','tip','poate_standalone','descriere','categorie_id',
                  'pret_furnizor','discount_furnizor','adaos_hsl']:
            if k in d: setattr(a, k, d[k])
        if 'compatibilitati' in d:
            AccesoriuCompat.query.filter_by(accesoriu_id=aid).delete()
            for comp in d['compatibilitati']:
                ac = AccesoriuCompat(accesoriu_id=aid, produs_id=comp['produs_id'], status=comp.get('status','optional'))
                ac.variante_compat = comp.get('variante_compat', [])
                db.session.add(ac)
        db.session.commit()
        return jsonify({'success': True})

    @app.route('/api/cfg/accesorii/<int:aid>', methods=['DELETE'])
    @login_required
    def api_cfg_accesoriu_delete(aid):
        a = Accesoriu.query.get_or_404(aid)
        a.activ = False
        db.session.commit()
        return jsonify({'success': True})

    # --- Categories ---
    @app.route('/api/cfg/categorii-produse', methods=['GET'])
    @login_required
    def api_cfg_cat_produse_list():
        return jsonify([{'id':c.id,'nume':c.nume,'ordine':c.ordine} for c in CategorieProdus.query.order_by(CategorieProdus.ordine).all()])

    @app.route('/api/cfg/categorii-produse', methods=['POST'])
    @login_required
    def api_cfg_cat_produse_create():
        d = request.get_json()
        mx = db.session.query(db.func.max(CategorieProdus.ordine)).scalar() or 0
        c = CategorieProdus(nume=d['nume'], ordine=mx+1)
        db.session.add(c); db.session.commit()
        return jsonify({'success': True, 'id': c.id})

    @app.route('/api/cfg/categorii-produse/<int:cid>', methods=['PUT'])
    @login_required
    def api_cfg_cat_produse_update(cid):
        c = CategorieProdus.query.get_or_404(cid)
        d = request.get_json()
        if 'nume' in d: c.nume = d['nume']
        db.session.commit()
        return jsonify({'success': True})

    @app.route('/api/cfg/categorii-produse/<int:cid>', methods=['DELETE'])
    @login_required
    def api_cfg_cat_produse_delete(cid):
        c = CategorieProdus.query.get_or_404(cid)
        db.session.delete(c); db.session.commit()
        return jsonify({'success': True})

    @app.route('/api/cfg/categorii-accesorii', methods=['GET'])
    @login_required
    def api_cfg_cat_acc_list():
        return jsonify([{'id':c.id,'nume':c.nume,'ordine':c.ordine} for c in CategorieAccesoriu.query.order_by(CategorieAccesoriu.ordine).all()])

    @app.route('/api/cfg/categorii-accesorii', methods=['POST'])
    @login_required
    def api_cfg_cat_acc_create():
        d = request.get_json()
        mx = db.session.query(db.func.max(CategorieAccesoriu.ordine)).scalar() or 0
        c = CategorieAccesoriu(nume=d['nume'], ordine=mx+1)
        db.session.add(c); db.session.commit()
        return jsonify({'success': True, 'id': c.id})

    @app.route('/api/cfg/categorii-accesorii/<int:cid>', methods=['PUT'])
    @login_required
    def api_cfg_cat_acc_update(cid):
        c = CategorieAccesoriu.query.get_or_404(cid)
        d = request.get_json()
        if 'nume' in d: c.nume = d['nume']
        db.session.commit()
        return jsonify({'success': True})

    @app.route('/api/cfg/categorii-accesorii/<int:cid>', methods=['DELETE'])
    @login_required
    def api_cfg_cat_acc_delete(cid):
        c = CategorieAccesoriu.query.get_or_404(cid)
        db.session.delete(c); db.session.commit()
        return jsonify({'success': True})

    # --- Configurator runtime APIs (used by configurator page) ---
    @app.route('/api/cfg/configurator/produse')
    @login_required
    def api_cfg_runtime_produse():
        """Full product data for configurator UI"""
        produse = ProdusConfig.query.filter_by(activ=True).all()
        result = []
        for p in produse:
            accs = []
            compats = AccesoriuCompat.query.filter_by(produs_id=p.id).all()
            for c in compats:
                a = Accesoriu.query.get(c.accesoriu_id)
                if a and a.activ:
                    accs.append({'id':a.id,'cod':a.cod,'denumire':a.denumire,'pret':a.pret,
                                'pret_mode':a.pret_mode,'um':a.um,'status':c.status,
                                'variante_compat':c.variante_compat})
            # Produse legate
            links = ProdusLegatura.query.filter(
                db.or_(ProdusLegatura.produs_a_id == p.id, ProdusLegatura.produs_b_id == p.id)
            ).all()
            legaturi = []
            for lk in links:
                other_id = lk.produs_b_id if lk.produs_a_id == p.id else lk.produs_a_id
                other = ProdusConfig.query.get(other_id)
                if other and other.activ:
                    is_a = lk.produs_a_id == p.id
                    principal_is_this = (lk.principal == 'a') == is_a
                    dim_is_this = (lk.dim_source == 'a') == is_a
                    ps = lk.price_source or 'both'
                    price_source_val = 'both' if ps == 'both' else ('this' if (ps == 'a') == is_a else 'other')
                    legaturi.append({
                        'link_id': lk.id,
                        'produs_id': other.id, 'cod': other.cod, 'denumire': other.denumire,
                        'pret': other.pret, 'um': other.um,
                        'parameter_types': other.parametri_config,
                        'variante': other.variante_config,
                        'compatibilitati': lk.compatibilitati,
                        'principal': 'this' if principal_is_this else 'other',
                        'dim_source': 'this' if dim_is_this else 'other',
                        'price_source': price_source_val,
                    })
            result.append({
                'id':p.id,'cod':p.cod,'denumire':p.denumire,'pret':p.pret,'um':p.um,
                'pret_furnizor':p.pret_furnizor,'discount_furnizor':p.discount_furnizor,
                'adaos_hsl':p.adaos_hsl,'pret_achizitie':p.pret_achizitie,
                'pret_minim_vanzare':p.pret_minim_vanzare,
                'categories':[pc.categorie_id for pc in p.categorii],
                'parameter_types':p.parametri_config,
                'variante':p.variante_config,
                'accessories':accs,
                'legaturi': legaturi,
            })
        return jsonify(result)

    @app.route('/api/cfg/configurator/match-varianta', methods=['POST'])
    @login_required
    def api_cfg_match_varianta():
        d = request.get_json()
        prod_cod = d.get('cod')
        params = d.get('params', {})
        p = ProdusConfig.query.filter_by(cod=prod_cod).first()
        if not p:
            return jsonify({'matched': False})
        vc = p.variante_config or {}
        variants = vc.get('variants', [])

        # Get variant_parameters to know which params matter for matching
        var_params = vc.get('variant_parameters', [])

        best_match = None
        best_score = -1

        for v in variants:
            v_combo = v.get('combination') or v.get('params') or {}
            score = 0
            mismatch = False

            # Compare all params that both sides have
            all_keys = set(list(v_combo.keys()) + list(params.keys()))
            for key in all_keys:
                v_val = str(v_combo.get(key, '')).strip()
                p_val = str(params.get(key, '')).strip()
                if not v_val and not p_val:
                    continue
                if v_val and p_val:
                    # Normalize: collapse whitespace for comparison
                    if v_val.replace(' ', '') == p_val.replace(' ', ''):
                        score += 1
                    else:
                        mismatch = True
                        break
                # One side has value, other doesn't - skip (don't penalize)

            if not mismatch and score > best_score and score > 0:
                best_score = score
                best_match = v

        if best_match:
            suffix = best_match.get('suffix', '') or best_match.get('code', '')
            price = best_match.get('pret') if best_match.get('pret') is not None else best_match.get('price', p.pret)
            # Variant-level pricing (fallback to product defaults)
            v_pf = best_match.get('pret_furnizor') if best_match.get('pret_furnizor') is not None else p.pret_furnizor
            v_df = best_match.get('discount_furnizor') if best_match.get('discount_furnizor') is not None else p.discount_furnizor
            v_adaos = best_match.get('adaos_hsl') if best_match.get('adaos_hsl') is not None else p.adaos_hsl
            v_achiz = v_pf * (1 - (v_df or 0) / 100)
            v_minim = v_achiz * (1 + (v_adaos or 0) / 100)
            return jsonify({
                'matched': True, 'variant_code': suffix, 'price': price,
                'pret_furnizor': v_pf, 'discount_furnizor': v_df, 'adaos_hsl': v_adaos,
                'pret_achizitie': round(v_achiz, 2), 'pret_minim_vanzare': round(v_minim, 2),
            })
        return jsonify({'matched': False, 'price': p.pret,
            'pret_furnizor': p.pret_furnizor, 'discount_furnizor': p.discount_furnizor,
            'adaos_hsl': p.adaos_hsl, 'pret_achizitie': round(p.pret_achizitie, 2),
            'pret_minim_vanzare': round(p.pret_minim_vanzare, 2),
        })

    @app.route('/api/cfg/configurator/all-accesorii')
    @login_required
    def api_cfg_all_accesorii():
        accs = Accesoriu.query.filter_by(activ=True, poate_standalone=True).all()
        return jsonify([{'id':a.id,'cod':a.cod,'denumire':a.denumire,'pret':a.pret,
                        'pret_mode':a.pret_mode,'um':a.um,'categorie_id':a.categorie_id} for a in accs])

    # ════════════════════════════════════════════════════════════
    # CONFIGURATOR - CLIENT PRICE LIST
    # ════════════════════════════════════════════════════════════
    @app.route('/api/cfg/configurator/client-preturi/<int:client_id>')
    @login_required
    def api_cfg_client_preturi(client_id):
        """Get price adjustments for a client based on their price list template"""
        client = Client.query.get_or_404(client_id)
        if not client.sablon_pret_id:
            return jsonify({'has_sablon': False})
        sablon = SablonListaPreturi.query.get(client.sablon_pret_id)
        if not sablon or not sablon.activ:
            return jsonify({'has_sablon': False})
        # Build per-product and per-variant discount map
        produse = ProdusConfig.query.filter_by(activ=True).all()
        discounts = {}
        for p in produse:
            cat_ids = [pc.categorie_id for pc in p.categorii]
            disc = sablon.get_discount(p.id, cat_ids)
            pret_client = p.pret * (1 - disc / 100)
            pret_min = p.pret_minim_vanzare
            if pret_client < pret_min:
                pret_client = pret_min
            # Per-product base discount
            prod_info = {
                'discount': disc,
                'pret_lista': p.pret,
                'pret_client': round(pret_client, 2),
                'pret_minim': round(pret_min, 2),
                'variante': {},
            }
            # Per-variant discounts
            vc = p.variante_config or {}
            for v in vc.get('variants', []):
                suffix = v.get('suffix', '')
                v_pret = v.get('pret') or p.pret
                v_disc = sablon.get_discount(p.id, cat_ids, suffix)
                v_pret_client = v_pret * (1 - v_disc / 100)
                # Variant-level minim
                v_pf = v.get('pret_furnizor') if v.get('pret_furnizor') is not None else p.pret_furnizor
                v_df = v.get('discount_furnizor') if v.get('discount_furnizor') is not None else p.discount_furnizor
                v_ad = v.get('adaos_hsl') if v.get('adaos_hsl') is not None else p.adaos_hsl
                v_achiz = v_pf * (1 - (v_df or 0) / 100)
                v_min = v_achiz * (1 + (v_ad or 0) / 100)
                if v_pret_client < v_min:
                    v_pret_client = v_min
                prod_info['variante'][suffix] = {
                    'discount': v_disc,
                    'pret_lista': v_pret,
                    'pret_client': round(v_pret_client, 2),
                    'pret_minim': round(v_min, 2),
                }
            discounts[str(p.id)] = prod_info
        return jsonify({
            'has_sablon': True,
            'sablon_nume': sablon.nume,
            'discount_global': sablon.discount_global,
            'discounts': discounts
        })

    # ════════════════════════════════════════════════════════════
    # CONFIGURATOR - PRODUSE LEGATE (many-to-many)
    # ════════════════════════════════════════════════════════════
    @app.route('/api/cfg/produse/<int:pid>/legaturi', methods=['GET'])
    @login_required
    def api_cfg_produs_legaturi_get(pid):
        p = ProdusConfig.query.get_or_404(pid)
        links = ProdusLegatura.query.filter(
            db.or_(ProdusLegatura.produs_a_id == pid, ProdusLegatura.produs_b_id == pid)
        ).all()
        result = []
        for lk in links:
            is_a = lk.produs_a_id == pid
            other = lk.produs_b if is_a else lk.produs_a
            # Translate a/b to this/other perspective
            principal_is_this = (lk.principal == 'a') == is_a
            dim_is_this = (lk.dim_source == 'a') == is_a
            price_src = lk.price_source or 'both'
            if price_src == 'both':
                price_source_val = 'both'
            else:
                price_source_val = 'this' if (price_src == 'a') == is_a else 'other'
            result.append({
                'id': lk.id,
                'produs_id': other.id, 'produs_cod': other.cod, 'produs_denumire': other.denumire,
                'compatibilitati': lk.compatibilitati,
                'principal': 'this' if principal_is_this else 'other',
                'dim_source': 'this' if dim_is_this else 'other',
                'price_source': price_source_val,
            })
        return jsonify(result)

    @app.route('/api/cfg/produse/<int:pid>/legaturi', methods=['POST'])
    @login_required
    def api_cfg_produs_legaturi_add(pid):
        d = request.get_json()
        other_id = d.get('produs_id')
        if not other_id or other_id == pid:
            return jsonify({'error': 'Produs invalid'}), 400
        # Check if link already exists
        existing = ProdusLegatura.query.filter(
            db.or_(
                db.and_(ProdusLegatura.produs_a_id == pid, ProdusLegatura.produs_b_id == other_id),
                db.and_(ProdusLegatura.produs_a_id == other_id, ProdusLegatura.produs_b_id == pid),
            )
        ).first()
        if existing:
            return jsonify({'error': 'Legătura există deja'}), 400
        lk = ProdusLegatura(produs_a_id=pid, produs_b_id=other_id)
        db.session.add(lk)
        db.session.commit()
        return jsonify({'success': True, 'id': lk.id})

    @app.route('/api/cfg/legaturi/<int:lid>', methods=['DELETE'])
    @login_required
    def api_cfg_legatura_delete(lid):
        lk = ProdusLegatura.query.get_or_404(lid)
        db.session.delete(lk)
        db.session.commit()
        return jsonify({'success': True})

    @app.route('/api/cfg/legaturi/<int:lid>/compatibilitati', methods=['PUT'])
    @login_required
    def api_cfg_legatura_compat_set(lid):
        lk = ProdusLegatura.query.get_or_404(lid)
        d = request.get_json()
        lk.compatibilitati = d.get('compatibilitati', [])
        db.session.commit()
        return jsonify({'success': True})

    @app.route('/api/cfg/legaturi/<int:lid>/settings', methods=['PUT'])
    @login_required
    def api_cfg_legatura_settings(lid):
        lk = ProdusLegatura.query.get_or_404(lid)
        d = request.get_json()
        # caller_pid tells us which side is "this"
        caller_pid = d.get('caller_pid')
        is_a = lk.produs_a_id == caller_pid
        # Convert this/other to a/b
        if d.get('principal') == 'this':
            lk.principal = 'a' if is_a else 'b'
        elif d.get('principal') == 'other':
            lk.principal = 'b' if is_a else 'a'
        if d.get('dim_source') == 'this':
            lk.dim_source = 'a' if is_a else 'b'
        elif d.get('dim_source') == 'other':
            lk.dim_source = 'b' if is_a else 'a'
        ps = d.get('price_source')
        if ps == 'both':
            lk.price_source = 'both'
        elif ps == 'this':
            lk.price_source = 'a' if is_a else 'b'
        elif ps == 'other':
            lk.price_source = 'b' if is_a else 'a'
        db.session.commit()
        return jsonify({'success': True})

    # ════════════════════════════════════════════════════════════
    # SAVE OFERTA (from configurator)
    # ════════════════════════════════════════════════════════════
    @app.route('/api/oferta/save', methods=['POST'])
    @login_required
    def api_oferta_save():
        d = request.get_json()
        tva_rate = float(Setari.get('tva_rate', '19'))
        edit_id = d.get('oferta_id')

        if edit_id:
            # UPDATE existing offer
            oferta = Oferta.query.get_or_404(edit_id)
            oferta.client_id = d.get('client_id')
            oferta.valabilitate_zile = d.get('valabilitate_zile', 30)
            oferta.data_expirare = date.today() + timedelta(days=d.get('valabilitate_zile', 30))
            oferta.discount_mode = d.get('discount_mode', 'individual')
            oferta.discount_global = d.get('discount_global', 0)
            oferta.observatii = d.get('observatii', '') or oferta.observatii
            # Delete old lines
            LinieOferta.query.filter_by(oferta_id=oferta.id).delete()
        else:
            # CREATE new offer
            nr = f"OF-{datetime.now().strftime('%Y%m%d-%H%M%S-%f')}"
            oferta = Oferta(
                numar=nr, client_id=d.get('client_id'),
                valabilitate_zile=d.get('valabilitate_zile', 30),
                data_oferta=date.today(),
                data_expirare=date.today() + timedelta(days=d.get('valabilitate_zile', 30)),
                discount_mode=d.get('discount_mode', 'individual'),
                discount_global=d.get('discount_global', 0),
                moneda='EUR', tva_procent=tva_rate,
                observatii=d.get('observatii', ''),
                creat_de_id=current_user.id,
            )
            db.session.add(oferta)
            db.session.flush()

        for idx, item in enumerate(d.get('items', [])):
            linie = LinieOferta(
                oferta_id=oferta.id, ordine=idx, tip=item.get('tip', 'Produs'),
                cod=item.get('cod', ''), denumire=item.get('denumire', ''),
                um=item.get('um', 'buc'), dimensiune=item.get('dimensiune', ''),
                cantitate=item.get('cantitate', 1), pret_catalog=item.get('pret_catalog', 0),
                discount_adaos=item.get('discount_adaos', 0), pret_final=item.get('pret_final', 0),
                is_sub_line=item.get('is_sub_line', False),
                parent_cod=item.get('parent_cod', ''),
            )
            linie.parametri = item.get('parametri', {})
            linie.accesorii = item.get('accesorii', [])
            db.session.add(linie)

        oferta.recalculeaza()
        db.session.commit()
        return jsonify({'success': True, 'oferta_id': oferta.id, 'numar': oferta.numar_display})

    # ════════════════════════════════════════════════════════════
    # EXCEL EXPORT (Oferta)
    # ════════════════════════════════════════════════════════════
    @app.route('/api/oferta/<int:oid>/excel')
    @login_required
    def api_oferta_excel(oid):
        if not HAS_OPENPYXL:
            return jsonify({'error': 'openpyxl not installed'}), 500
        oferta = Oferta.query.get_or_404(oid)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f'Oferta {oferta.numar}'

        # ── Styles matching HSL template ──
        arial8 = Font(name='Arial', size=8)
        arial8b = Font(name='Arial', size=8, bold=True)
        arial9b = Font(name='Arial', size=9, bold=True)
        hsl_green = 'FF56B347'       # header green
        hsl_green_light = 'FFC4F0BA'  # product row light green
        fill_header = PatternFill('solid', fgColor=hsl_green)
        fill_prod = PatternFill('solid', fgColor=hsl_green_light)
        fill_subtotal = PatternFill('solid', fgColor=hsl_green)
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left = Alignment(horizontal='left', vertical='center', wrap_text=True)
        right = Alignment(horizontal='right', vertical='center')
        num_fmt = '#,##0.00'
        thin_border = Border(
            bottom=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'))

        # ── Column widths ──
        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 55
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 16
        ws.column_dimensions['F'].width = 18

        # ── Header section ──
        client_name = oferta.client.nume if oferta.client else 'N/A'
        contact = ''
        if oferta.client and oferta.client.persoana_contact:
            contact = oferta.client.persoana_contact

        ws['A1'] = f'Catre: {client_name}'
        ws['A1'].font = arial9b
        ws['A1'].alignment = left
        if contact:
            ws['A3'] = f'Attn: {contact}'
            ws['A3'].font = arial9b
            ws['A3'].alignment = left

        ws.merge_cells('A5:F5')
        ws['A5'] = 'OFERTA'
        ws['A5'].font = arial9b
        ws['A5'].alignment = Alignment(horizontal='center')

        # ── Table header row ──
        row = 8
        headers = [('Simbol', 14), ('Descriere', 55), ('Dimensiuni gol zid', 16),
                   ('Buc', 8), ('Pret EUR fara TVA', 16), ('Valoare EUR fara TVA', 18)]
        for col, (h, _) in enumerate(headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = arial8b
            c.fill = fill_header
            c.alignment = center
            c.border = thin_border

        # ── Build product groups ──
        # Collect main lines (not sub-lines) and their sub-lines
        main_lines = []
        for l in oferta.linii:
            if not l.is_sub_line:
                main_lines.append(l)

        # Pre-load product descriptions for lookup
        prod_descrieri = {}
        all_prods = ProdusConfig.query.filter_by(activ=True).all()
        for p in all_prods:
            if p.descriere:
                prod_descrieri[p.cod] = p.descriere

        subtotal_rows = []  # track subtotal row numbers for grand total

        for ml in main_lines:
            row += 1
            prod_start_row = row

            # Product main line (bold, green bg on A and D)
            c = ws.cell(row=row, column=1, value=ml.cod)
            c.font = arial8b; c.fill = fill_prod; c.alignment = Alignment(horizontal='left', wrap_text=True); c.border = thin_border

            # Use descriere if available, otherwise denumire
            # Try exact code match, then base code (before dot for variants like PFG.6012A)
            desc = prod_descrieri.get(ml.cod, '')
            if not desc and '.' in (ml.cod or ''):
                base_code = ml.cod.split('.')[0]
                desc = prod_descrieri.get(base_code, '')
            display_name = desc or ml.denumire
            # Replace [parametru] placeholders with actual values
            if display_name and ml.parametri:
                for key, val in ml.parametri.items():
                    if val:
                        if key == 'canate':
                            v = str(val)
                            canate_str = v + (' canat' if v == '1' else ' canate')
                            display_name = display_name.replace('[canate]', canate_str)
                        else:
                            display_name = display_name.replace(f'[{key}]', str(val))
            c = ws.cell(row=row, column=2, value=display_name)
            c.font = arial8b; c.alignment = left; c.border = thin_border

            dim = ml.dimensiune or ''
            if not dim and ml.parametri:
                dim = ml.parametri.get('dimensiune', '')
            c = ws.cell(row=row, column=3, value=dim)
            c.font = arial8b; c.alignment = center; c.border = thin_border

            c = ws.cell(row=row, column=4, value=ml.cantitate)
            c.font = arial8b; c.fill = fill_prod; c.alignment = center; c.border = thin_border

            # Price with formula reference
            c = ws.cell(row=row, column=5, value=ml.pret_final)
            c.font = arial8b; c.alignment = center; c.number_format = num_fmt; c.border = thin_border

            # Value = D * E
            c = ws.cell(row=row, column=6)
            c.value = f'=D{row}*E{row}'
            c.font = arial8b; c.alignment = center; c.number_format = num_fmt; c.border = thin_border

            # Sub-lines: linked products
            sub_lines = [l for l in oferta.linii if l.is_sub_line and l.parent_cod == ml.cod and l.tip == 'Produs']
            for sl in sub_lines:
                row += 1
                ws.cell(row=row, column=1).border = thin_border
                c = ws.cell(row=row, column=2, value=sl.denumire)
                c.font = arial8b; c.alignment = left; c.border = thin_border

                sl_dim = sl.dimensiune or ''
                if not sl_dim and sl.parametri:
                    sl_dim = sl.parametri.get('dimensiune', '')
                c = ws.cell(row=row, column=3, value=sl_dim)
                c.font = arial8b; c.alignment = center; c.border = thin_border

                c = ws.cell(row=row, column=4, value=sl.cantitate)
                c.font = arial8b; c.alignment = center; c.border = thin_border

                c = ws.cell(row=row, column=5, value=sl.pret_final)
                c.font = arial8b; c.alignment = center; c.number_format = num_fmt; c.border = thin_border

                c = ws.cell(row=row, column=6)
                c.value = f'=D{row}*E{row}'
                c.font = arial8b; c.alignment = center; c.number_format = num_fmt; c.border = thin_border

            # Accessories (standard = no price, just "standard")
            for acc in ml.accesorii:
                row += 1
                is_std = acc.get('is_standard', False)
                pret_acc = acc.get('pret_final', acc.get('pret', 0))
                has_price = not is_std and pret_acc and pret_acc > 0

                ws.cell(row=row, column=1).border = thin_border
                c = ws.cell(row=row, column=2, value=acc.get('denumire', ''))
                c.font = arial8b if has_price else arial8
                c.alignment = left; c.border = thin_border

                ws.cell(row=row, column=3).border = thin_border

                c = ws.cell(row=row, column=4)
                c.value = f'=D{prod_start_row}'
                c.font = arial8b if has_price else arial8
                c.alignment = center; c.border = thin_border

                if has_price:
                    c = ws.cell(row=row, column=5, value=pret_acc)
                    c.font = arial8b; c.alignment = center; c.number_format = num_fmt; c.border = thin_border
                    c = ws.cell(row=row, column=6)
                    c.value = f'=E{row}*D{row}'
                    c.font = arial8b; c.alignment = center; c.number_format = num_fmt; c.border = thin_border
                else:
                    c = ws.cell(row=row, column=5, value='standard')
                    c.font = arial8; c.alignment = center; c.border = thin_border
                    c = ws.cell(row=row, column=6, value='standard')
                    c.font = arial8; c.alignment = center; c.border = thin_border

            # Sub-line accessories
            acc_sub_lines = [l for l in oferta.linii if l.is_sub_line and l.parent_cod == ml.cod and l.tip == 'Accesoriu']
            for sl in acc_sub_lines:
                row += 1
                has_price = sl.pret_final and sl.pret_final > 0

                ws.cell(row=row, column=1).border = thin_border
                c = ws.cell(row=row, column=2, value=sl.denumire)
                c.font = arial8b if has_price else arial8
                c.alignment = left; c.border = thin_border

                ws.cell(row=row, column=3).border = thin_border

                c = ws.cell(row=row, column=4)
                c.value = f'=D{prod_start_row}'
                c.font = arial8b if has_price else arial8
                c.alignment = center; c.border = thin_border

                if has_price:
                    c = ws.cell(row=row, column=5, value=sl.pret_final)
                    c.font = arial8b; c.alignment = center; c.number_format = num_fmt; c.border = thin_border
                    c = ws.cell(row=row, column=6)
                    c.value = f'=E{row}*D{row}'
                    c.font = arial8b; c.alignment = center; c.number_format = num_fmt; c.border = thin_border
                else:
                    c = ws.cell(row=row, column=5, value='standard')
                    c.font = arial8; c.alignment = center; c.border = thin_border
                    c = ws.cell(row=row, column=6, value='standard')
                    c.font = arial8; c.alignment = center; c.border = thin_border

            # ── Subtotal row (green bar) ──
            row += 1
            ws.cell(row=row, column=1).border = thin_border
            c = ws.cell(row=row, column=2)
            c.value = f'=A{prod_start_row}'
            c.font = arial8b; c.fill = fill_subtotal; c.alignment = right; c.border = thin_border

            c = ws.cell(row=row, column=3, value='subtotal produse')
            c.font = arial8b; c.fill = fill_subtotal; c.alignment = right; c.border = thin_border

            c = ws.cell(row=row, column=4)
            c.value = f'=D{prod_start_row}'
            c.font = arial8b; c.fill = fill_subtotal; c.alignment = center; c.border = thin_border

            c = ws.cell(row=row, column=5)
            c.value = f'=F{row}/D{row}'
            c.font = arial8b; c.fill = fill_subtotal; c.alignment = center; c.number_format = num_fmt; c.border = thin_border

            c = ws.cell(row=row, column=6)
            c.value = f'=SUM(F{prod_start_row}:F{row-1})'
            c.font = arial8b; c.fill = fill_subtotal; c.alignment = center; c.number_format = num_fmt; c.border = thin_border

            subtotal_rows.append(row)

        # ── Total Bucati ──
        row += 1
        qty_refs = '+'.join([f'D{r}' for r in subtotal_rows])
        c = ws.cell(row=row, column=2, value='Total Bucati')
        c.font = arial8b; c.alignment = left
        c = ws.cell(row=row, column=4)
        c.value = f'={qty_refs}' if qty_refs else 0
        c.font = arial8b; c.fill = fill_prod; c.alignment = center

        # ── Total Produse ──
        row += 1
        val_refs = '+'.join([f'F{r}' for r in subtotal_rows])
        c = ws.cell(row=row, column=2, value='Total Produse')
        c.font = arial8b; c.fill = fill_subtotal; c.alignment = left
        c = ws.cell(row=row, column=4)
        c.value = f'=D{row-1}'
        c.font = arial8b; c.fill = fill_subtotal; c.alignment = center
        c = ws.cell(row=row, column=6)
        c.value = f'={val_refs}' if val_refs else 0
        c.font = arial8b; c.fill = fill_subtotal; c.alignment = center; c.number_format = num_fmt

        # ── Total General ──
        row += 1
        c = ws.cell(row=row, column=2, value='Total General EUR fara TVA')
        c.font = arial8b; c.fill = fill_subtotal; c.alignment = left
        c = ws.cell(row=row, column=6)
        c.value = f'=F{row-1}'
        c.font = arial8b; c.fill = fill_subtotal; c.alignment = center; c.number_format = num_fmt

        # ── Footer notes ──
        row += 2
        notes = [
            ('Oferta nu cuprinde lucrari pentru ajustarea golurilor de montaj, demontat usi existente, confectionar cadre.', False),
            ('', False),
            ('Preturile sunt exprimate in EURO si nu contin TVA.', False),
            ('', False),
            ('Oferta nu include transport si montaj.', True),
            ('', False),
            (f'Termen de livrare: conform conditiilor agreate', True),
            ('', False),
            (f'Termen de plata: conform conditiilor agreate', True),
        ]
        for text, bold_flag in notes:
            row += 1
            c = ws.cell(row=row, column=1, value=text)
            c.font = arial8b if bold_flag else arial8
            c.alignment = left

        # ── Signature ──
        row += 2
        ws.cell(row=row, column=1, value='Cu respect,').font = arial8
        agent = oferta.creat_de
        row += 2
        ws.cell(row=row, column=1, value=agent.nume_complet if agent else '').font = arial8
        row += 1
        ws.cell(row=row, column=1).font = arial8
        row += 1
        ws.cell(row=row, column=1, value=Setari.get('company_name', 'HSL Solutions SRL')).font = arial8
        row += 1
        ws.cell(row=row, column=1, value=Setari.get('company_address', '')).font = arial8
        if agent and agent.telefon:
            row += 1
            ws.cell(row=row, column=1, value=f'Mob: {agent.telefon}').font = arial8
        ws.cell(row=row+1, column=1, value=f'Tel: {Setari.get("company_phone", "")}').font = arial8
        ws.cell(row=row+2, column=1, value=f'E-mail: {agent.email if agent else Setari.get("company_email", "")}').font = arial8
        ws.cell(row=row+3, column=1, value='website: www.hsl.ro').font = arial8

        # ── Print setup ──
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True

        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        return send_file(buf, as_attachment=True,
                        download_name=f'Oferta_{oferta.numar}.xlsx',
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    # ════════════════════════════════════════════════════════════
    # SETARI API
    # ════════════════════════════════════════════════════════════
    @app.route('/api/setari', methods=['GET'])
    @login_required
    def api_setari_get():
        keys = ['company_name','company_address','company_phone','company_email',
                'tva_rate','serie_factura','serie_proforma','scadenta_factura_zile','curs_multiplicator']
        return jsonify({k: Setari.get(k, '') for k in keys})

    @app.route('/api/setari', methods=['POST'])
    @module_required('utilizatori')
    def api_setari_save():
        d = request.get_json()
        for k, v in d.items():
            Setari.set_val(k, v)
        return jsonify({'success': True})

    # ════════════════════════════════════════════════════════════
    # ADMIN - ROLURI
    # ════════════════════════════════════════════════════════════
    @app.route('/admin/roluri')
    @module_required('utilizatori')
    def roluri_list():
        roluri = Rol.query.order_by(Rol.id).all()
        return render_template('admin/roluri.html', roluri=roluri, MODULES=MODULES)

    @app.route('/admin/roluri/nou', methods=['GET', 'POST'])
    @module_required('utilizatori')
    def rol_nou():
        if request.method == 'POST':
            r = Rol(
                nume=request.form.get('nume', '').strip(),
                descriere=request.form.get('descriere', '').strip(),
                doar_proprii='doar_proprii' in request.form
            )
            perm = {}
            for key, label, icon, group in MODULES:
                if key in request.form:
                    perm[key] = True
            r.set_permisiuni(perm)
            db.session.add(r)
            db.session.commit()
            flash(f'Rol "{r.nume}" creat!', 'success')
            return redirect(url_for('roluri_list'))
        return render_template('admin/rol_form.html', rol=None, MODULES=MODULES)

    @app.route('/admin/roluri/<int:rid>', methods=['GET', 'POST'])
    @module_required('utilizatori')
    def rol_edit(rid):
        r = Rol.query.get_or_404(rid)
        if request.method == 'POST':
            if r.is_system:
                flash('Rolul Admin nu poate fi modificat.', 'error')
                return redirect(url_for('roluri_list'))
            r.nume = request.form.get('nume', r.nume).strip()
            r.descriere = request.form.get('descriere', '').strip()
            r.doar_proprii = 'doar_proprii' in request.form
            perm = {}
            for key, label, icon, group in MODULES:
                if key in request.form:
                    perm[key] = True
            r.set_permisiuni(perm)
            db.session.commit()
            flash(f'Rol "{r.nume}" actualizat!', 'success')
            return redirect(url_for('roluri_list'))
        return render_template('admin/rol_form.html', rol=r, MODULES=MODULES)

    @app.route('/admin/roluri/<int:rid>/sterge', methods=['POST'])
    @module_required('utilizatori')
    def rol_sterge(rid):
        r = Rol.query.get_or_404(rid)
        if r.is_system:
            flash('Rolul Admin nu poate fi șters.', 'error')
        elif r.utilizatori.count() > 0:
            flash(f'Rolul "{r.nume}" are utilizatori asociați. Mută-i mai întâi.', 'error')
        else:
            db.session.delete(r)
            db.session.commit()
            flash(f'Rol "{r.nume}" șters.', 'success')
        return redirect(url_for('roluri_list'))

    # ════════════════════════════════════════════════════════════
    # ADMIN - UTILIZATORI
    # ════════════════════════════════════════════════════════════
    @app.route('/admin/utilizatori')
    @module_required('utilizatori')
    def utilizatori_list():
        users = Utilizator.query.order_by(Utilizator.data_creare.desc()).all()
        return render_template('admin/utilizatori.html', utilizatori=users)

    @app.route('/api/utilizatori/status')
    @login_required
    def api_utilizatori_status():
        users = Utilizator.query.filter_by(activ=True).all()
        now = datetime.now(timezone.utc)
        result = []
        for u in users:
            if u.last_seen:
                ls = u.last_seen
                if ls.tzinfo is None:
                    ls = ls.replace(tzinfo=timezone.utc)
                delta = (now - ls).total_seconds()
                if delta < 300:  # 5 min
                    status = 'online'
                elif delta < 900:  # 15 min
                    status = 'away'
                else:
                    status = 'offline'
            else:
                status = 'offline'
                delta = None
            result.append({
                'id': u.id, 'nume': u.nume_complet, 'username': u.username,
                'rol': u.rol_nume, 'status': status,
                'current_page': u.current_page or '',
                'last_seen': u.last_seen.isoformat() if u.last_seen else None,
                'seconds_ago': int(delta) if delta is not None else None,
            })
        return jsonify(result)

    @app.route('/admin/utilizatori/nou', methods=['GET', 'POST'])
    @module_required('utilizatori')
    def utilizator_nou():
        roluri = Rol.query.order_by(Rol.id).all()
        if request.method == 'POST':
            username = request.form.get('username', '').strip()
            email = request.form.get('email', '').strip()
            if Utilizator.query.filter_by(username=username).first():
                flash('Username deja existent!', 'error')
                return render_template('admin/utilizator_form.html', user=None, roluri=roluri)
            u = Utilizator(
                username=username, email=email or f'{username}@hsl.ro',
                nume_complet=request.form.get('nume_complet', '').strip(),
                telefon=request.form.get('telefon', '').strip(),
                comision_procent=float(request.form.get('comision_procent') or 0.75),
                rol_id=int(request.form.get('rol_id', 0)) or None,
                activ='activ' in request.form
            )
            u.set_password(request.form.get('password', 'parola123'))
            db.session.add(u)
            db.session.commit()
            flash(f'Utilizator {u.username} creat!', 'success')
            return redirect(url_for('utilizatori_list'))
        return render_template('admin/utilizator_form.html', user=None, roluri=roluri)

    @app.route('/admin/utilizatori/<int:uid>', methods=['GET', 'POST'])
    @module_required('utilizatori')
    def utilizator_edit(uid):
        u = Utilizator.query.get_or_404(uid)
        roluri = Rol.query.order_by(Rol.id).all()
        if request.method == 'POST':
            u.username = request.form.get('username', u.username).strip()
            u.email = request.form.get('email', u.email).strip()
            u.nume_complet = request.form.get('nume_complet', u.nume_complet).strip()
            u.telefon = request.form.get('telefon', '').strip()
            u.comision_procent = float(request.form.get('comision_procent') or 0.75)
            u.rol_id = int(request.form.get('rol_id', 0)) or None
            u.activ = 'activ' in request.form
            pw = request.form.get('password', '').strip()
            if pw:
                u.set_password(pw)
            db.session.commit()
            flash(f'Utilizator {u.username} actualizat!', 'success')
            return redirect(url_for('utilizatori_list'))
        return render_template('admin/utilizator_form.html', user=u, roluri=roluri)

    @app.route('/admin/utilizatori/<int:uid>/toggle', methods=['POST'])
    @module_required('utilizatori')
    def utilizator_toggle(uid):
        u = Utilizator.query.get_or_404(uid)
        if u.is_admin:
            flash('Nu poți dezactiva contul de administrator!', 'error')
        else:
            u.activ = not u.activ
            db.session.commit()
            flash(f'Contul {u.username} a fost {"activat" if u.activ else "dezactivat"}.', 'success')
        return redirect(url_for('utilizator_edit', uid=uid))

    # ════════════════════════════════════════════════════════════
    # ADMIN - TIPURI ACTIVITĂȚI & ȘABLOANE
    # ════════════════════════════════════════════════════════════
    @app.route('/admin/activitati/config')
    @module_required('activitati_manage')
    def activitati_config():
        tipuri = TipActivitate.query.order_by(TipActivitate.ordine).all()
        sabloane = SablonActivitate.query.order_by(SablonActivitate.data_creare.desc()).all()
        return render_template('admin/activitati_config.html', tipuri=tipuri, sabloane=sabloane)

    # --- Tipuri CRUD (AJAX) ---
    @app.route('/api/tip-activitate', methods=['POST'])
    @login_required
    def api_tip_activitate_create():
        d = request.get_json()
        t = TipActivitate(nume=d.get('nume','').strip(), culoare=d.get('culoare','#6c757d'),
                          ordine=d.get('ordine', 0))
        db.session.add(t)
        db.session.commit()
        return jsonify({'success': True, 'id': t.id})

    @app.route('/api/tip-activitate/<int:tid>', methods=['PUT'])
    @login_required
    def api_tip_activitate_update(tid):
        t = TipActivitate.query.get_or_404(tid)
        d = request.get_json()
        t.nume = d.get('nume', t.nume).strip()
        t.culoare = d.get('culoare', t.culoare)
        t.ordine = d.get('ordine', t.ordine)
        t.activ = d.get('activ', t.activ)
        db.session.commit()
        return jsonify({'success': True})

    @app.route('/api/tip-activitate/<int:tid>', methods=['DELETE'])
    @login_required
    def api_tip_activitate_delete(tid):
        t = TipActivitate.query.get_or_404(tid)
        db.session.delete(t)
        db.session.commit()
        return jsonify({'success': True})

    # --- Șabloane CRUD ---
    @app.route('/admin/sabloane/nou', methods=['GET','POST'])
    @module_required('activitati_manage')
    def sablon_nou():
        tipuri = TipActivitate.query.filter_by(activ=True).order_by(TipActivitate.ordine).all()
        if request.method == 'POST':
            s = SablonActivitate(
                nume=request.form.get('nume','').strip(),
                descriere=request.form.get('descriere','').strip(),
                trigger=request.form.get('trigger','manual'),
                activ='activ' in request.form,
            )
            db.session.add(s)
            db.session.flush()
            # Parse lines
            idx = 0
            while f'linie_titlu_{idx}' in request.form:
                titlu = request.form.get(f'linie_titlu_{idx}','').strip()
                if titlu:
                    l = LinieSablon(
                        sablon_id=s.id, titlu=titlu,
                        descriere=request.form.get(f'linie_desc_{idx}','').strip(),
                        tip_id=int(request.form.get(f'linie_tip_{idx}') or 0) or None,
                        prioritate=request.form.get(f'linie_prio_{idx}','normala'),
                        ordine=idx,
                    )
                    db.session.add(l)
                idx += 1
            db.session.commit()
            flash(f'Șablon "{s.nume}" creat!', 'success')
            return redirect(url_for('activitati_config'))
        return render_template('admin/sablon_form.html', sablon=None, tipuri=tipuri)

    @app.route('/admin/sabloane/<int:sid>', methods=['GET','POST'])
    @module_required('activitati_manage')
    def sablon_edit(sid):
        s = SablonActivitate.query.get_or_404(sid)
        tipuri = TipActivitate.query.filter_by(activ=True).order_by(TipActivitate.ordine).all()
        if request.method == 'POST':
            s.nume = request.form.get('nume', s.nume).strip()
            s.descriere = request.form.get('descriere','').strip()
            s.trigger = request.form.get('trigger', s.trigger)
            s.activ = 'activ' in request.form
            # Clear old lines and re-add
            LinieSablon.query.filter_by(sablon_id=s.id).delete()
            idx = 0
            while f'linie_titlu_{idx}' in request.form:
                titlu = request.form.get(f'linie_titlu_{idx}','').strip()
                if titlu:
                    l = LinieSablon(
                        sablon_id=s.id, titlu=titlu,
                        descriere=request.form.get(f'linie_desc_{idx}','').strip(),
                        tip_id=int(request.form.get(f'linie_tip_{idx}') or 0) or None,
                        prioritate=request.form.get(f'linie_prio_{idx}','normala'),
                        ordine=idx,
                    )
                    db.session.add(l)
                idx += 1
            db.session.commit()
            flash(f'Șablon "{s.nume}" actualizat!', 'success')
            return redirect(url_for('activitati_config'))
        return render_template('admin/sablon_form.html', sablon=s, tipuri=tipuri)

    @app.route('/admin/sabloane/<int:sid>/sterge', methods=['POST'])
    @module_required('activitati_manage')
    def sablon_sterge(sid):
        s = SablonActivitate.query.get_or_404(sid)
        db.session.delete(s)
        db.session.commit()
        flash(f'Șablon "{s.nume}" șters.', 'success')
        return redirect(url_for('activitati_config'))

    @app.route('/api/sablon/<int:sid>/aplica', methods=['POST'])
    @login_required
    def api_sablon_aplica(sid):
        """Manually apply a template"""
        s = SablonActivitate.query.get_or_404(sid)
        d = request.get_json() or {}
        created = s.aplica(
            comanda_id=int(d.get('comanda_id') or 0) or None,
            client_id=int(d.get('client_id') or 0) or None,
            creat_de_id=current_user.id,
        )
        db.session.commit()
        return jsonify({'success': True, 'count': len(created)})

    # ════════════════════════════════════════════════════════════
    # DB INIT
    # ════════════════════════════════════════════════════════════
    with app.app_context():
        db.create_all()
        # Create system Admin role if not exists
        admin_rol = Rol.query.filter_by(is_system=True).first()
        if not admin_rol:
            admin_rol = Rol(nume='Admin', descriere='Acces complet la toate modulele', is_system=True, doar_proprii=False)
            db.session.add(admin_rol)
            db.session.flush()
        # Create default Agent role
        agent_rol = Rol.query.filter_by(nume='Agent Vânzări').first()
        if not agent_rol:
            agent_rol = Rol(nume='Agent Vânzări', descriere='Acces la CRM, Configurator, Oferte, Comenzi', doar_proprii=True)
            agent_rol.set_permisiuni({'dashboard': True, 'crm': True, 'configurator': True, 'oferte': True, 'comenzi': True, 'activitati': True, 'activitati_status': True})
            db.session.add(agent_rol)
            db.session.flush()
        # Create default users
        if Utilizator.query.count() == 0:
            admin = Utilizator(username='admin', email='admin@hsl.ro', nume_complet='Administrator', rol_id=admin_rol.id)
            admin.set_password('admin123')
            db.session.add(admin)
            agent = Utilizator(username='robert', email='robert@hsl.ro', nume_complet='Robert', rol_id=agent_rol.id, telefon='0721000000')
            agent.set_password('robert123')
            db.session.add(agent)
        db.session.commit()
        # Default settings
        defaults = {'company_name':'HSL Solutions SRL','company_address':'','company_phone':'',
                    'company_email':'','tva_rate':'19','serie_factura':'HSL','serie_proforma':'PF','scadenta_factura_zile':'30'}
        for k, v in defaults.items():
            if not Setari.query.filter_by(cheie=k).first():
                db.session.add(Setari(cheie=k, valoare=v))
        db.session.commit()

    return app


app = create_app()

# ═══ MAIL SYNC SCHEDULER ═══
def start_mail_scheduler(app):
    """Start background scheduler for automatic mail sync"""
    try:
        from apscheduler.schedulers.background import BackgroundScheduler
        from mail_service import sync_inbox
    except ImportError:
        print("  ⚠️  APScheduler not installed - mail auto-sync disabled")
        print("     Run: pip install apscheduler")
        return None

    def sync_all_accounts():
        with app.app_context():
            from models import ContMail
            conturi = ContMail.query.filter_by(activ=True).all()
            for cont in conturi:
                try:
                    n = sync_inbox(cont, max_results=20)
                    if n and n > 0:
                        print(f"  📬 Sync {cont.email}: {n} mesaje noi")
                except Exception as e:
                    print(f"  ❌ Sync error {cont.email}: {e}")

    scheduler = BackgroundScheduler(daemon=True)
    scheduler.add_job(sync_all_accounts, 'interval', minutes=2, id='mail_sync',
                      max_instances=1, coalesce=True)
    scheduler.start()
    print("  📧 Mail auto-sync: la fiecare 2 minute")
    return scheduler

app = create_app()

if __name__ == '__main__':
    print("=" * 60)
    print("  🏭 HSL Solutions ERP v3.0")
    print("  📍 http://localhost:5000")
    print("  📍 Dashboard:     http://localhost:5000/admin")
    print("  📍 Configurator:  http://localhost:5000/configurator")
    print("  📍 Config Admin:  http://localhost:5000/configurator/admin")
    print("  🔑 Login: admin / admin123")
    mail_scheduler = start_mail_scheduler(app)
    print("=" * 60)
    app.run(debug=True, host='0.0.0.0', port=5000, use_reloader=False)
